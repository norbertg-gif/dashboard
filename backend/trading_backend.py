#!/usr/bin/env python3
"""
Trading Dashboard Backend — port 8766
pip install fastapi uvicorn yfinance pandas requests
"""

import asyncio
import gc, json, os, math, threading
import re
import hashlib
import hmac
from io import BytesIO
import numpy as np
from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED
_executor = ThreadPoolExecutor(max_workers=2)
from pathlib import Path
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from fastapi import FastAPI, Query, HTTPException, Request, Header
from fastapi.middleware.cors import CORSMiddleware
import yfinance as yf
import pandas as pd
import requests
import urllib.parse
import time as _time_module

# ── RETRY HELPER ──────────────────────────────────────────────────────────────
def fetch_with_retry(url: str, headers: dict = None, timeout: int = 10,
                     retries: int = 3, backoff: float = 1.5) -> requests.Response:
    """GET s exponential backoff. Vyvolá výnimku ak všetky pokusy zlyhajú."""
    last_exc = None
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=headers or {}, timeout=timeout)
            resp.raise_for_status()
            return resp
        except requests.exceptions.Timeout as e:
            last_exc = e
            print(f"  Timeout ({attempt+1}/{retries}): {url[:60]}")
        except requests.exceptions.HTTPError as e:
            # 4xx — neopakuj (napr. 404, 401), 5xx — opakuj
            if e.response is not None and e.response.status_code < 500:
                raise
            last_exc = e
            print(f"  HTTP {e.response.status_code} ({attempt+1}/{retries}): {url[:60]}")
        except Exception as e:
            last_exc = e
            print(f"  Error ({attempt+1}/{retries}): {e}")
        if attempt < retries - 1:
            _time_module.sleep(backoff ** attempt)
    raise last_exc
import uvicorn
# scipy/sklearn importované lazy vnútri funkcií (train_ml_model, optimalizácia
# váh) — ušetrí ~60–80 MB RAM pri štarte; načítajú sa až keď ich treba.

BASE_DIR = Path(__file__).resolve().parent
APP_ROOT = BASE_DIR.parent if BASE_DIR.name == "backend" else BASE_DIR
FRONTEND_DIR = APP_ROOT / "frontend"
DATA_ROOT = Path(os.getenv("DATA_DIR", str(APP_ROOT))).resolve()
DATA_ROOT.mkdir(parents=True, exist_ok=True)
app = FastAPI(title="Trading Dashboard API")


# ?? Memory profile / feature flags ???????????????????????????????????????????
def _env_bool(name: str, default: bool) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return raw.strip().lower() in ("1", "true", "yes", "on")

DASH_MEMORY_PROFILE = os.getenv("DASH_MEMORY_PROFILE", "low").strip().lower()
_LOW_MEMORY = DASH_MEMORY_PROFILE in ("low", "512", "512mb", "render-free")

# Heavy analytical layers. Low profile keeps the dashboard usable under 512 MB:
# core charts/portfolio/signals remain, but fit-heavy diagnostics run only when enabled.
ENABLE_PREDICTIVE_ML = _env_bool("ENABLE_PREDICTIVE_ML", not _LOW_MEMORY)
ENABLE_PREDICTIVE_HMM = _env_bool("ENABLE_PREDICTIVE_HMM", not _LOW_MEMORY)
ENABLE_SIGNAL_CONTEXT_BACKFILL = _env_bool("ENABLE_SIGNAL_CONTEXT_BACKFILL", not _LOW_MEMORY)
ENABLE_SIGNAL_ANALYTICS = _env_bool("ENABLE_SIGNAL_ANALYTICS", True)
ENABLE_MARKET_BREADTH = _env_bool("ENABLE_MARKET_BREADTH", not _LOW_MEMORY)
ENABLE_MASSIVE_SP500 = _env_bool("ENABLE_MASSIVE_SP500", not _LOW_MEMORY)
ENABLE_MASSIVE_MARKET = _env_bool("ENABLE_MASSIVE_MARKET", True)


def _process_memory_mb() -> dict:
    """Best-effort process memory snapshot without psutil dependency."""
    rss_mb = None
    source = None
    try:
        import resource
        rss = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
        # Linux reports KB, macOS reports bytes. Render is Linux, keep guard anyway.
        rss_mb = rss / 1024 if rss < 10_000_000 else rss / (1024 * 1024)
        source = "resource.ru_maxrss"
    except Exception:
        try:
            import tracemalloc
            if not tracemalloc.is_tracing():
                tracemalloc.start()
            current, peak = tracemalloc.get_traced_memory()
            return {"rss_mb": None, "tracemalloc_current_mb": round(current / 1024 / 1024, 1),
                    "tracemalloc_peak_mb": round(peak / 1024 / 1024, 1), "source": "tracemalloc"}
        except Exception:
            pass
    return {"rss_mb": round(rss_mb, 1) if rss_mb is not None else None, "source": source}


@app.get("/api/admin/memory")
def admin_memory():
    return {
        "memory": _process_memory_mb(),
        "profile": DASH_MEMORY_PROFILE,
        "feature_flags": {
            "predictive_ml": ENABLE_PREDICTIVE_ML,
            "predictive_hmm": ENABLE_PREDICTIVE_HMM,
            "signal_context_backfill": ENABLE_SIGNAL_CONTEXT_BACKFILL,
            "signal_analytics": ENABLE_SIGNAL_ANALYTICS,
            "market_breadth": ENABLE_MARKET_BREADTH,
            "massive_sp500": ENABLE_MASSIVE_SP500,
            "massive_market": ENABLE_MASSIVE_MARKET,
        },
        "cache_sizes": {
            "model": len(_MODEL_CACHE) if "_MODEL_CACHE" in globals() else None,
            "backtest": len(_BACKTEST_CACHE) if "_BACKTEST_CACHE" in globals() else None,
            "yf": len(_YF_CACHE) if "_YF_CACHE" in globals() else None,
            "disk_mem": len(_cache_mem) if "_cache_mem" in globals() else None,
            "positions": len(_positions_cache) if "_positions_cache" in globals() else None,
        },
    }

# ── Public API token (pre Claude / externý prístup bez Basic Auth) ────────────
PUBLIC_API_TOKEN = os.getenv("PUBLIC_API_TOKEN", "")
PUBLIC_RATE_LIMIT_WINDOW = int(os.getenv("PUBLIC_RATE_LIMIT_WINDOW", "60"))
PUBLIC_RATE_LIMIT_MAX = int(os.getenv("PUBLIC_RATE_LIMIT_MAX", "30"))
_public_rate_lock = threading.Lock()
_public_rate: dict[str, list[float]] = {}

# ══ PREDICTIVE CHART — functions ══════════════════════════════
# ── Indicators ────────────────────────────────────────────────────────────────
# NOTE: calc_ema/calc_rsi/calc_macd/calc_ichimoku/calc_stoch_rsi/calc_adx are
# defined ONCE in the OHLCV+INDIKÁTORY section below (CLAUDE.md pitfall —
# a duplicate set here used to shadow-break ADX unpacking in add_indicators).

def calc_atr(df, period=14):
    hl = df["High"] - df["Low"]
    hc = (df["High"] - df["Close"].shift()).abs()
    lc = (df["Low"]  - df["Close"].shift()).abs()
    tr = pd.concat([hl, hc, lc], axis=1).max(axis=1)
    return tr.ewm(com=period - 1, adjust=False).mean()

def add_indicators(df):
    df = df.copy()
    df["ema10"]      = calc_ema(df["Close"], 10)
    df["ema20"]      = calc_ema(df["Close"], 20)
    df["rsi"]        = calc_rsi(df["Close"])
    df["atr"]        = calc_atr(df)
    df["macd"], df["macd_sig"], df["macd_hist"] = calc_macd(df["Close"])
    df["vol_ma"]     = df["Volume"].rolling(10).mean()
    df["vol_ratio"]  = df["Volume"] / df["vol_ma"]
    _t, _k, _sa, _sb, _ch = calc_ichimoku(df)
    df["ichi_tenkan"] = pd.to_numeric(_t,  errors='coerce')
    df["ichi_kijun"]  = pd.to_numeric(_k,  errors='coerce')
    df["ichi_sa"]     = pd.to_numeric(_sa, errors='coerce')
    df["ichi_sb"]     = pd.to_numeric(_sb, errors='coerce')
    df["ichi_chikou"] = pd.to_numeric(_ch, errors='coerce')
    _sk, _sd = calc_stoch_rsi(df["Close"])
    df["stoch_k"] = pd.to_numeric(_sk, errors='coerce')
    df["stoch_d"] = pd.to_numeric(_sd, errors='coerce')
    _adx_df = calc_adx(df)
    df["adx"]      = pd.to_numeric(_adx_df["adx"],      errors='coerce')
    df["di_plus"]  = pd.to_numeric(_adx_df["di_plus"],  errors='coerce')
    df["di_minus"] = pd.to_numeric(_adx_df["di_minus"], errors='coerce')
    # Momentum & candle features (from ML feature set)
    df["ret_1"] = df["Close"].pct_change(1)
    df["ret_3"] = df["Close"].pct_change(3)
    df["ret_5"] = df["Close"].pct_change(5)
    df["body"]  = (df["Close"] - df["Open"]) / df["Open"]
    df["range"] = (df["High"] - df["Low"]) / df["Close"]
    df["volatility"] = df["ret_1"].rolling(10).std()
    df["ema20_dist"] = (df["Close"] - df["ema20"]) / df["Close"]
    # ROC (4-period momentum) + pozícia v 52-period high/low rozsahu (0=dno, 1=vrchol).
    # Na weekly grafe = ~1 mesiac momentum a 1-ročný rozsah; min_periods drží feature
    # živú aj na kratšej histórii (inak by dropna zožralo prvých 52 riadkov).
    df["roc_4"] = df["Close"].pct_change(4)
    _roll_max = df["High"].rolling(52, min_periods=20).max()
    _roll_min = df["Low"].rolling(52, min_periods=20).min()
    df["pos_52w"] = (df["Close"] - _roll_min) / (_roll_max - _roll_min).replace(0, np.nan)
    return df

# ── ML confidence model ──────────────────────────────────────────────────────

WEIGHTS_LOG     = DATA_ROOT / "predictive_weights_log.json"
SIGNALS_LOG     = DATA_ROOT / "predictive_signals_log.json"
SIGNALS_ARCHIVE = DATA_ROOT / "predictive_signals_archive.json"
SCANNER_NOTES_FILE = DATA_ROOT / "scanner_notes.json"
DEFAULT_WEIGHTS = {"ema": 0.20, "rsi": 0.10, "macd": 0.20, "vol": 0.15, "ichi": 0.25, "stoch": 0.10}


def load_weights_log() -> dict:
    if WEIGHTS_LOG.exists():
        try:
            return json.loads(WEIGHTS_LOG.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def save_weights_log(log: dict):
    # Prunovanie: vyhoď tickery ktoré neboli optimalizované posledných 180 dní
    cutoff = str((datetime.now(timezone.utc) - timedelta(days=180)).date())
    pruned = {t: v for t, v in log.items()
              if isinstance(v, dict) and str(v.get("optimized_at", "9999")) >= cutoff}
    WEIGHTS_LOG.write_text(json.dumps(pruned, indent=2, ensure_ascii=False), encoding="utf-8")

def load_signals_log() -> dict:
    if SIGNALS_LOG.exists():
        try:
            return json.loads(SIGNALS_LOG.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def _archive_pruned_signals(archived: dict):
    """Append pruned (>90d) signal entries to the archive file instead of losing them.
    Archive grows append-only; needed for per-regime analytics (20-30 signals per regime)."""
    if not archived:
        return
    try:
        existing = {}
        if SIGNALS_ARCHIVE.exists():
            try:
                existing = json.loads(SIGNALS_ARCHIVE.read_text(encoding="utf-8"))
            except Exception:
                print("  WARN: signals archive corrupted, starting fresh (old file kept as .bak)")
                try:
                    SIGNALS_ARCHIVE.rename(SIGNALS_ARCHIVE.with_suffix(".json.bak"))
                except Exception:
                    pass
        for ticker, entries in archived.items():
            existing.setdefault(ticker, {}).update(entries)
        SIGNALS_ARCHIVE.write_text(json.dumps(existing, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception as e:
        print(f"  WARN: failed to archive pruned signals: {e}")

def load_signals_archive() -> dict:
    if SIGNALS_ARCHIVE.exists():
        try:
            return json.loads(SIGNALS_ARCHIVE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def save_signals_log(log: dict):
    # Prunovanie: zachovaj len záznamy z posledných 90 dní; staršie presuň do archívu
    cutoff = (datetime.now(timezone.utc) - timedelta(days=90)).date()
    pruned = {}
    archived = {}
    for ticker, entries in log.items():
        if not isinstance(entries, dict):
            continue
        kept = {d: v for d, v in entries.items() if d >= str(cutoff)}
        old = {d: v for d, v in entries.items() if d < str(cutoff)}
        if kept:
            pruned[ticker] = kept
        if old:
            archived[ticker] = old
    _archive_pruned_signals(archived)
    SIGNALS_LOG.write_text(json.dumps(pruned, indent=2, ensure_ascii=False), encoding="utf-8")

ML_FEATURES = ["ret_1", "ret_3", "ret_5", "body", "range",
               "volatility", "ema20_dist", "rsi", "macd_hist", "vol_ratio",
               "roc_4", "pos_52w"]

# ── Model cache (ML + HMM) ────────────────────────────────────────────────────
# train_ml_model (RandomForest 3-fold) aj detect_market_regime (GaussianHMM) sa
# inak fitujú pri každom /api/chart requeste. Keď sa cache_key (ticker+timeframe+
# posledná sviečka) nezmení, výsledok je identický — tak ho cacheujeme do najbližšej
# uzavretej sviečky. ML ukladá len (acc, bull_prob), nie samotný model (šetrí RAM
# na Render free tieri); regime ukladá celý dict (malý).
_MODEL_CACHE: dict = {}
_MODEL_CACHE_LOCK = threading.Lock()
_MODEL_CACHE_MAX = 48 if _LOW_MEMORY else 256


def _model_cache_get(kind: str, cache_key):
    if not cache_key:
        return None
    with _MODEL_CACHE_LOCK:
        return _MODEL_CACHE.get((kind, cache_key))


def _model_cache_put(kind: str, cache_key, value):
    if not cache_key:
        return
    with _MODEL_CACHE_LOCK:
        if len(_MODEL_CACHE) >= _MODEL_CACHE_MAX:
            for k in list(_MODEL_CACHE.keys())[: _MODEL_CACHE_MAX // 3]:
                _MODEL_CACHE.pop(k, None)
        _MODEL_CACHE[(kind, cache_key)] = value


def train_ml_model(df, cache_key=None):
    """Walk-forward validácia: 3 expanding-window foldy namiesto jedného 70/30
    splitu — accuracy je priemer cez foldy, takže nezávisí od toho, ktorý režim
    padol do test setu. Finálny model sa trénuje na všetkých dátach.
    Returns (model, mean_test_accuracy, bull_prob_for_last_row).
    Pri cache hit je model None (nepoužíva sa downstream) — vraciame (acc, bull)."""
    cached = _model_cache_get("ml", cache_key)
    if cached is not None:
        return None, cached[0], cached[1]
    df_ml = df.copy()
    df_ml["target"] = (df_ml["Close"].shift(-1) > df_ml["Close"]).astype(int)
    df_ml = df_ml.dropna(subset=ML_FEATURES + ["target"])

    if len(df_ml) < 40:
        return None, None, 0.5

    usable = df_ml.iloc[:-1]  # last row has no future close
    X_all = usable[ML_FEATURES].values
    y_all = usable["target"].values
    n = len(usable)

    def make_model():
        from sklearn.ensemble import RandomForestClassifier
        from sklearn.calibration import CalibratedClassifierCV
        base = RandomForestClassifier(
            n_estimators=50, max_depth=4,
            min_samples_leaf=5, random_state=42, n_jobs=1
        )
        return CalibratedClassifierCV(base, cv=2)

    # Walk-forward foldy: train na [0:60%/70%/80%], test na nasledujúcich 10 %
    fold_accs = []
    for train_frac in (0.60, 0.70, 0.80):
        split = int(n * train_frac)
        test_end = int(n * (train_frac + 0.10))
        if split < 20 or test_end - split < 5:
            continue
        m = make_model()
        try:
            m.fit(X_all[:split], y_all[:split])
            fold_accs.append(m.score(X_all[split:test_end], y_all[split:test_end]))
        except Exception:
            continue

    if not fold_accs:
        return None, None, 0.5

    acc = float(np.mean(fold_accs)) * 100

    # Finálny model na všetkých dátach → predikcia pre posledný riadok
    model = make_model()
    model.fit(X_all, y_all)
    last_row = df_ml.iloc[-1][ML_FEATURES].values.reshape(1, -1)
    bull_prob = float(model.predict_proba(last_row)[0][1])

    acc_rounded = round(acc, 1)
    _model_cache_put("ml", cache_key, (acc_rounded, bull_prob))
    return model, acc_rounded, bull_prob


# ── HMM Regime Detection ──────────────────────────────────────────────────────

def detect_market_regime(df, cache_key=None) -> dict:
    """Fit a 3-state Gaussian HMM on log-returns and map states to
    bear / sideways / bull / high_volatility using mean return + volatility.

    Returns dict with: regime (str), confidence (float 0-1), states (list),
    regime_history (list of {date, regime} for last 26 bars), or error (str).
    """
    cached = _model_cache_get("hmm", cache_key)
    if cached is not None:
        return cached
    try:
        from hmmlearn.hmm import GaussianHMM
    except ImportError:
        return {"error": "hmmlearn nie je nainštalovaný (pip install hmmlearn)"}

    try:
        closes = df["Close"].values.astype(float)
        if len(closes) < 60:
            return {"error": "Nedostatok dát (min 60 sviečok)"}

        log_returns = np.diff(np.log(closes))
        X = log_returns.reshape(-1, 1)

        model = GaussianHMM(
            n_components=3,
            covariance_type="diag",
            n_iter=200,
            random_state=42,
            tol=1e-4,
        )
        model.fit(X)

        if not model.monitor_.converged:
            # Retry with more iterations — convergence failure is common on short/flat series
            model = GaussianHMM(n_components=3, covariance_type="diag", n_iter=500, random_state=0)
            model.fit(X)

        hidden_states = model.predict(X)

        # Map hidden states → semantic labels by mean return per state
        state_means = [log_returns[hidden_states == s].mean() if (hidden_states == s).any() else 0
                       for s in range(3)]
        state_vols  = [log_returns[hidden_states == s].std()  if (hidden_states == s).any() else 0
                       for s in range(3)]

        # Sort by mean return: lowest = bear, mid = sideways, highest = bull
        order = np.argsort(state_means)   # [bear_state, sideways_state, bull_state]
        label_map = {order[0]: "bear", order[1]: "sideways", order[2]: "bull"}

        # High volatility override: if current state has vol > 1.8× median → high_volatility
        median_vol = float(np.median(state_vols))
        current_raw_state = int(hidden_states[-1])
        current_vol = state_vols[current_raw_state]
        if median_vol > 0 and current_vol > median_vol * 1.8:
            current_regime = "high_volatility"
        else:
            current_regime = label_map[current_raw_state]

        # Posterior probability of current state = confidence
        log_prob, posteriors = model.score_samples(X)
        confidence = float(posteriors[-1, current_raw_state])
        regime_probabilities = {
            label_map[state]: round(float(posteriors[-1, state]), 4)
            for state in range(3)
        }

        # Last 26 bars of regime history for chart shading
        dates = df.index[-26:] if len(df) >= 26 else df.index
        states_tail = hidden_states[-(len(dates)):]
        vols_tail   = [state_vols[s] for s in states_tail]
        history = []
        for i, (d, s, v) in enumerate(zip(dates, states_tail, vols_tail)):
            lbl = "high_volatility" if (median_vol > 0 and v > median_vol * 1.8) else label_map[int(s)]
            history.append({"date": str(d.date() if hasattr(d, 'date') else d)[:10], "regime": lbl})

        result = {
            "regime":          current_regime,
            "confidence":      round(confidence, 3),
            "regime_probabilities": regime_probabilities,
            "regime_history":  history,
            "state_means_pct": [round(m * 100, 3) for m in [state_means[i] for i in order]],
            "state_vols_pct":  [round(v * 100, 3) for v in [state_vols[i]  for i in order]],
            "model": {
                "name": "GaussianHMM",
                "components": 3,
                "covariance_type": "diag",
                "random_state": int(model.random_state or 0),
                "n_iter": int(model.n_iter),
                "training_bars": int(len(closes)),
            },
            "error": None,
        }
        _model_cache_put("hmm", cache_key, result)
        return result
    except Exception as e:
        return {"error": f"HMM chyba: {type(e).__name__}: {str(e)[:80]}"}


# ── Prediction engine ─────────────────────────────────────────────────────────

ANALOG_FEATURE_COLUMNS = [
    "ema_gap",
    "ema20_dist",
    "kijun_dist",
    "kumo_pos",
    "rsi",
    "macd_hist_pct",
    "vol_ratio",
    "stoch_k",
    "adx",
    "atr_pct",
    "roc_4",
    "pos_52w",
]

# Prah pre analog override nad drift priorom — |vote| >= 0.60 znamená ~80/20
# zhodu susedov. Nižšie hodnoty merateľne zhoršujú smerovú úspešnosť (viď
# komentár v _analog_prediction).
ANALOG_OVERRIDE_VOTE = 0.60


def _analog_feature_frame(df: pd.DataFrame) -> pd.DataFrame:
    """Compact numeric fingerprint of each candle setup for nearest-neighbor matching."""
    close = pd.to_numeric(df["Close"], errors="coerce").astype(float)
    high = pd.to_numeric(df["High"], errors="coerce").astype(float)
    low = pd.to_numeric(df["Low"], errors="coerce").astype(float)
    frame = pd.DataFrame(index=df.index)

    ema10 = pd.to_numeric(df.get("ema10"), errors="coerce")
    ema20 = pd.to_numeric(df.get("ema20"), errors="coerce")
    kijun = pd.to_numeric(df.get("ichi_kijun"), errors="coerce")
    ichi_sa = pd.to_numeric(df.get("ichi_sa"), errors="coerce")
    ichi_sb = pd.to_numeric(df.get("ichi_sb"), errors="coerce")
    atr = pd.to_numeric(df.get("atr"), errors="coerce")

    frame["ema_gap"] = (ema10 - ema20) / close
    frame["ema20_dist"] = pd.to_numeric(df.get("ema20_dist"), errors="coerce")
    frame["kijun_dist"] = (close - kijun) / close

    kumo_top = pd.concat([ichi_sa, ichi_sb], axis=1).max(axis=1)
    kumo_bottom = pd.concat([ichi_sa, ichi_sb], axis=1).min(axis=1)
    kumo_width = (kumo_top - kumo_bottom).replace(0, np.nan)
    frame["kumo_pos"] = np.where(
        close > kumo_top,
        (close - kumo_top) / close,
        np.where(close < kumo_bottom, (close - kumo_bottom) / close, (close - (kumo_top + kumo_bottom) / 2) / kumo_width),
    )

    frame["rsi"] = pd.to_numeric(df.get("rsi"), errors="coerce")
    frame["macd_hist_pct"] = pd.to_numeric(df.get("macd_hist"), errors="coerce") / close
    frame["vol_ratio"] = pd.to_numeric(df.get("vol_ratio"), errors="coerce").clip(lower=0, upper=4)
    frame["stoch_k"] = pd.to_numeric(df.get("stoch_k"), errors="coerce")
    frame["adx"] = pd.to_numeric(df.get("adx"), errors="coerce")
    frame["atr_pct"] = atr / close
    frame["roc_4"] = pd.to_numeric(df.get("roc_4"), errors="coerce")
    if "pos_52w" in df.columns:
        frame["pos_52w"] = pd.to_numeric(df.get("pos_52w"), errors="coerce")
    else:
        roll_max = high.rolling(52, min_periods=20).max()
        roll_min = low.rolling(52, min_periods=20).min()
        frame["pos_52w"] = (close - roll_min) / (roll_max - roll_min).replace(0, np.nan)

    return frame.replace([np.inf, -np.inf], np.nan)


def _analog_prediction(df: pd.DataFrame, min_history: int = 60, neighbors: int = 21) -> dict | None:
    """Predict next direction from historically similar setups.

    The candidate set only uses candles whose following candle is already inside
    the provided df. During backtest df is a past-only slice, so this remains
    walk-forward and does not peek into the future.
    """
    if df is None or len(df) < min_history:
        return None
    try:
        features = _analog_feature_frame(df)
        current = features.iloc[-1]
        candidate_features = features.iloc[:-1].copy()
        closes = pd.to_numeric(df["Close"], errors="coerce").astype(float).reset_index(drop=True)
        next_returns = (closes.shift(-1) - closes) / closes
        outcomes = np.where(next_returns >= 0, 1.0, -1.0)

        work = candidate_features.copy()
        work["_ret"] = next_returns.iloc[:-1].to_numpy(dtype=float)
        work["_outcome"] = outcomes[:-1]
        needed = ANALOG_FEATURE_COLUMNS + ["_ret", "_outcome"]
        work = work[needed].dropna()
        current = current[ANALOG_FEATURE_COLUMNS]
        if current.isna().any() or len(work) < max(25, neighbors):
            return None

        x = work[ANALOG_FEATURE_COLUMNS].astype(float)
        mean = x.mean(axis=0)
        std = x.std(axis=0).replace(0, 1.0).fillna(1.0)
        xz = (x - mean) / std
        cz = ((current.astype(float) - mean) / std).to_numpy(dtype=float)
        distances = np.sqrt(((xz.to_numpy(dtype=float) - cz) ** 2).sum(axis=1))
        if not np.isfinite(distances).all():
            return None

        k = min(neighbors, len(work))
        idx = np.argsort(distances)[:k]
        selected = work.iloc[idx].copy()
        selected_dist = distances[idx]
        similarity = 1.0 / (1.0 + selected_dist)

        # Slight recency tilt: similar recent setups matter a bit more, but old
        # analogs still vote. Index positions remain from the original df.
        pos = np.array([candidate_features.index.get_loc(i) for i in selected.index], dtype=float)
        recency = 0.65 + 0.35 * ((pos + 1.0) / max(1.0, len(candidate_features)))
        vote_weights = similarity * recency
        weight_sum = float(vote_weights.sum())
        if weight_sum <= 0:
            return None

        outcomes_sel = selected["_outcome"].to_numpy(dtype=float)
        returns_sel = selected["_ret"].to_numpy(dtype=float)
        vote = float(np.sum(outcomes_sel * vote_weights) / weight_sum)
        weighted_move = float(np.sum(returns_sel * vote_weights) / weight_sum)

        # Smerové rozhodnutie: bázou je historický drift titulu (up-rate celej
        # walk-forward histórie slicu). Merané na 7854 predikciách (S&P500
        # weekly, 34 tickerov, 2013-2018): samotný susedský hlas prehráva
        # s base rate o ~2.6 pb; drift prior + override len pri silnej zhode
        # susedov (|vote| >= 0.60, ~80/20 hlasovanie) bol najlepší variant
        # (test 54.8% vs 52.2% čistého hlasu). Vote ostáva zdrojom magnitúdy.
        hist_outcomes = work["_outcome"].to_numpy(dtype=float)
        up_rate = float((hist_outcomes > 0).mean())
        drift = 2.0 * up_rate - 1.0
        if abs(vote) >= ANALOG_OVERRIDE_VOTE:
            direction = 1 if vote >= 0 else -1
            decision = "analog_override"
        else:
            direction = 1 if drift >= 0 else -1
            decision = "drift_prior"

        if weighted_move == 0 or (weighted_move > 0) != (direction > 0):
            avg_abs_move = float(np.sum(np.abs(returns_sel) * vote_weights) / weight_sum)
            weighted_move = direction * avg_abs_move * max(0.35, abs(vote), abs(drift))
        weighted_move = float(np.clip(weighted_move, -0.15, 0.15))
        if abs(weighted_move) < 0.001:
            weighted_move = direction * 0.001

        up_count = int((outcomes_sel > 0).sum())
        down_count = int((outcomes_sel < 0).sum())
        edge = abs(vote) if decision == "analog_override" else abs(drift)
        confidence = float(np.clip(0.50 + edge * 0.50, 0.50, 0.99))
        composite = float(np.clip(weighted_move / 0.10, -1.0, 1.0))
        return {
            "direction": direction,
            "vote": vote,
            "drift": round(drift, 4),
            "up_rate": round(up_rate, 4),
            "decision": decision,
            "confidence": confidence,
            "neighbors": int(k),
            "up": up_count,
            "down": down_count,
            "move_pct": weighted_move,
            "composite": composite,
            "avg_abs_move_pct": float(np.sum(np.abs(returns_sel) * vote_weights) / weight_sum),
        }
    except Exception:
        return None


def predict_next_candle(df, weights: dict = None, **kwargs):
    row        = df.iloc[-1]
    last_close = float(row["Close"])

    ema_diff   = (row["ema10"] - row["ema20"]) / last_close
    ema_signal = float(np.tanh(ema_diff * 20))

    rsi_val = float(row["rsi"])
    if rsi_val > 65:
        rsi_signal = -((rsi_val - 65) / 35)
    elif rsi_val < 35:
        rsi_signal = (35 - rsi_val) / 35
    else:
        rsi_signal = 0.0

    macd_signal = float(np.tanh(row["macd_hist"] / (last_close * 0.001 + 1e-9)))

    try:
        vol_ratio = float(row["vol_ratio"])
        if math.isnan(vol_ratio): vol_ratio = 1.0
    except (ValueError, TypeError):
        vol_ratio = 1.0
    vol_signal = float(np.clip(vol_ratio - 1.0, -0.5, 0.5))

    # Ichimoku Kumo signal
    try:
        ichi_sa = float(row["ichi_sa"])
        if math.isnan(ichi_sa): ichi_sa = last_close
    except (ValueError, TypeError):
        ichi_sa = last_close
    try:
        ichi_sb = float(row["ichi_sb"])
        if math.isnan(ichi_sb): ichi_sb = last_close
    except (ValueError, TypeError):
        ichi_sb = last_close
    kumo_top    = max(ichi_sa, ichi_sb)
    kumo_bottom = min(ichi_sa, ichi_sb)
    kumo_thick  = (kumo_top - kumo_bottom) / last_close  # cloud thickness as % of price
    if last_close > kumo_top:
        ichi_signal = min(1.0, (last_close - kumo_top) / (last_close * 0.02 + 1e-9))
    elif last_close < kumo_bottom:
        ichi_signal = max(-1.0, (last_close - kumo_bottom) / (last_close * 0.02 + 1e-9))
    else:
        # inside cloud = weak signal, bias toward center direction
        ichi_signal = (last_close - (kumo_top + kumo_bottom) / 2) / (kumo_top - kumo_bottom + 1e-9) * 0.3
    # Tenkan/Kijun cross
    tk_diff = (float(row["ichi_tenkan"]) - float(row["ichi_kijun"])) / last_close
    tk_signal = float(np.tanh(tk_diff * 30))
    # Combine kumo + tk cross
    ichi_combined = ichi_signal * 0.6 + tk_signal * 0.4

    # Stochastic RSI signal
    try:
        stoch_k = float(row["stoch_k"])
        if math.isnan(stoch_k): stoch_k = 50.0
    except (ValueError, TypeError):
        stoch_k = 50.0
    if stoch_k > 80:
        stoch_signal = -((stoch_k - 80) / 20)
    elif stoch_k < 20:
        stoch_signal = (20 - stoch_k) / 20
    else:
        stoch_signal = 0.0

    # ADX signal -- modulates confidence, not direction
    try:
        adx_val = float(row["adx"])
        if math.isnan(adx_val): adx_val = 20.0
    except (ValueError, TypeError):
        adx_val = 20.0
    # Strong trend (ADX>25) amplifies directional signals, weak trend dampens
    adx_factor = float(np.clip((adx_val - 20) / 30, -0.5, 1.0))

    # Weights — use provided or defaults
    w = weights if weights else DEFAULT_WEIGHTS
    tech_composite = (
        ema_signal    * w.get("ema",   0.20) +
        rsi_signal    * w.get("rsi",   0.10) +
        macd_signal   * w.get("macd",  0.20) +
        vol_signal    * w.get("vol",   0.15) +
        ichi_combined * w.get("ichi",  0.25) +
        stoch_signal  * w.get("stoch", 0.10)
    ) * (1.0 + adx_factor * 0.3)  # ADX scales overall confidence +/-30%
    tech_composite = float(np.clip(tech_composite, -1.0, 1.0))

    # ML confidence modulation — blend with ML bull probability if provided
    ml_bull_prob = kwargs.get("ml_bull_prob") if kwargs else None
    composite = tech_composite
    if ml_bull_prob is not None:
        # Convert prob to signal: 0.5 = neutral, 1.0 = +1, 0.0 = -1
        ml_signal = float(np.clip((ml_bull_prob - 0.5) * 2, -1.0, 1.0))
        # Blend: 70% technical composite, 30% ML signal
        composite = float(np.clip(composite * 0.70 + ml_signal * 0.30, -1.0, 1.0))

    atr        = float(row["atr"])
    pred_open  = last_close
    analog = _analog_prediction(df)
    prediction_method = "technical_composite"
    if analog:
        prediction_method = "analog_similarity"
        composite = float(analog["composite"])
        pred_close = last_close * (1 + float(analog["move_pct"]))
    else:
        pred_close = last_close * (1 + composite * 0.6) + composite * atr * 0.4
    mid        = (pred_open + pred_close) / 2
    pred_high  = mid + atr * 0.75
    pred_low   = mid - atr * 0.75

    # Technical entry zone
    kijun    = float(row["ichi_kijun"]) if not math.isnan(float(row["ichi_kijun"])) else last_close
    kumo_bot = float(min(
        row["ichi_sa"] if not math.isnan(float(row["ichi_sa"])) else last_close,
        row["ichi_sb"] if not math.isnan(float(row["ichi_sb"])) else last_close
    ))
    ema20    = float(row["ema20"])
    atr_val  = float(row["atr"])

    # Support levels weighted by reliability
    supports = {
        "kijun":    (kijun,    0.40),  # strongest ichimoku support
        "kumo_bot": (kumo_bot, 0.35),  # cloud bottom
        "ema20":    (ema20,    0.25),  # dynamic support
    }
    # Only include levels below current price (actual support)
    valid = {k: v for k, v in supports.items() if v[0] < last_close}
    if valid:
        weighted_entry = sum(v[0] * v[1] for v in valid.values()) / sum(v[1] for v in valid.values())
        entry_low  = round(min(v[0] for v in valid.values()) - atr_val * 0.25, 2)
        entry_high = round(weighted_entry + atr_val * 0.15, 2)
        entry_mid  = round(weighted_entry, 2)
    else:
        # All supports above price — use ATR-based zone below current price
        entry_mid  = round(last_close - atr_val * 0.5, 2)
        entry_low  = round(last_close - atr_val * 1.0, 2)
        entry_high = round(last_close - atr_val * 0.25, 2)

    return {
        "open":  round(pred_open,  4),
        "high":  round(pred_high,  4),
        "low":   round(pred_low,   4),
        "close": round(pred_close, 4),
        "composite": round(composite, 4),
        "method": prediction_method,
        "tech_composite": round(tech_composite, 4),
        "analog": {
            "vote": round(float(analog["vote"]), 4),
            "drift": round(float(analog.get("drift", 0)), 4),
            "up_rate": round(float(analog.get("up_rate", 0.5)) * 100, 1),
            "decision": analog.get("decision", "analog_override"),
            "confidence": round(float(analog["confidence"]) * 100, 1),
            "neighbors": int(analog["neighbors"]),
            "up": int(analog["up"]),
            "down": int(analog["down"]),
            "move_pct": round(float(analog["move_pct"]) * 100, 2),
            "avg_abs_move_pct": round(float(analog["avg_abs_move_pct"]) * 100, 2),
        } if analog else None,
        "entry_zone": {
            "low":  entry_low,
            "mid":  entry_mid,
            "high": entry_high,
            "levels": {k: round(v[0], 2) for k, v in supports.items()},
        },
        "signals": {
            "ema":   {"value": round(ema_diff * 100, 3),         "signal": round(ema_signal, 3),    "weight": w.get("ema",   0.20)},
            "rsi":   {"value": round(rsi_val, 2),                "signal": round(rsi_signal, 3),    "weight": w.get("rsi",   0.10)},
            "macd":  {"value": round(float(row["macd_hist"]), 4), "signal": round(macd_signal, 3),  "weight": w.get("macd",  0.20)},
            "atr":   {"value": round(atr, 4),                    "signal": 0.0,                     "weight": 0.0},
            "vol":   {"value": round(vol_ratio, 3),              "signal": round(vol_signal, 3),    "weight": w.get("vol",   0.15)},
            "ichi":  {"value": round(ichi_combined, 3),          "signal": round(ichi_combined, 3), "weight": w.get("ichi",  0.25)},
            "stoch": {"value": round(stoch_k, 2),                "signal": round(stoch_signal, 3),  "weight": w.get("stoch", 0.10)},
            "adx":   {"value": round(adx_val, 2),                "signal": round(adx_factor, 3),    "weight": 0.0},
            "analog": {
                "value": round(float(analog["vote"]), 3) if analog else None,
                "signal": round(float(analog["confidence"]) * (1 if analog and analog["direction"] >= 0 else -1), 3) if analog else 0.0,
                "weight": 0.0,
            },
        }
    }

# ── Weekly trend label (Donchian 20w + SMA50w + EMA10/20 konfirmácia) ────────
# Nahrádza pôvodné bool weekly_bullish za 5-stupňový label. Bool ostáva
# k dispozícii ako derivát (score >= 3) pre ML kontext a signal log
# kompatibilitu — táto funkcia je čisto prezentačná vrstva nad weekly DF.

WEEKLY_TREND_LABELS = {
    2:  {"key": "strong_up",   "label": "Strong uptrend",   "icon": "⬆⬆", "dir": "up"},
    1:  {"key": "up",          "label": "Uptrend",          "icon": "⬆",  "dir": "up"},
    0:  {"key": "range",       "label": "Range / sideways", "icon": "→",  "dir": "side"},
    -1: {"key": "down",        "label": "Downtrend",        "icon": "⬇",  "dir": "down"},
    -2: {"key": "strong_down", "label": "Strong downtrend", "icon": "⬇⬇", "dir": "down"},
}


def _weekly_trend(df_w, asof_index: int | None = None) -> dict | None:
    """Vyhodnotí dlhodobý týždenný trend pomocou 20-týždňového Donchian
    kanála + 50-týždňovej SMA + EMA10/EMA20 ako momentum konfirmácie.

    asof_index = None znamená posledná sviečka. Vracia None keď je málo dát
    (potrebujeme aspoň 20 týždňov pre Donchian)."""
    if df_w is None or len(df_w) < 20:
        return None
    idx = len(df_w) - 1 if asof_index is None else asof_index
    if idx < 19:
        return None
    window = df_w.iloc[max(0, idx - 19): idx + 1]
    if len(window) < 20:
        return None
    high20 = float(window["High"].max())
    low20  = float(window["Low"].min())
    close  = float(df_w["Close"].iloc[idx])
    rng = high20 - low20
    donchian_pos = (close - low20) / rng if rng > 0 else 0.5

    # SMA50 (long-term trend filter) — fail-soft pri kratšej histórii
    sma50 = None
    if idx >= 49:
        sma50 = float(df_w["Close"].iloc[idx - 49: idx + 1].mean())
    above_sma50 = (sma50 is not None and close > sma50)

    ema10 = float(df_w["ema10"].iloc[idx]) if "ema10" in df_w.columns else None
    ema20 = float(df_w["ema20"].iloc[idx]) if "ema20" in df_w.columns else None
    ema_bull = (ema10 is not None and ema20 is not None and ema10 > ema20)

    # 5-stupňové skóre — Donchian pozícia primárna, SMA50/EMA len ako gating
    if donchian_pos >= 0.80 and (above_sma50 or sma50 is None):
        score = 2
    elif donchian_pos >= 0.55 and (above_sma50 or sma50 is None):
        score = 1
    elif donchian_pos < 0.15 and (not above_sma50) and (not ema_bull):
        score = -2
    elif donchian_pos < 0.30 and (not above_sma50):
        score = -1
    else:
        score = 0

    meta = WEEKLY_TREND_LABELS[score]
    return {
        "score": score,
        "label": meta["label"],
        "key": meta["key"],
        "icon": meta["icon"],
        "direction": meta["dir"],
        "donchian_pos": round(donchian_pos, 3),
        "donchian_high": round(high20, 4),
        "donchian_low":  round(low20, 4),
        "above_sma50": above_sma50,
        "ema_bull": ema_bull,
        "close": round(close, 4),
        "sma50": round(sma50, 4) if sma50 is not None else None,
    }


def _weekly_trend_to_bullish(trend: dict | None) -> bool | None:
    """Derivát starého bool weekly_bullish — score >= 1 znamená Uptrend alebo
    Strong uptrend. Používa sa pre ML kontext, signal log a stará scanner UI."""
    if trend is None:
        return None
    return trend.get("score", 0) >= 1


# ── Backtesting ───────────────────────────────────────────────────────────────

def calc_raw_signals(df_slice):
    """Extract raw signals from last row without applying weights."""
    row = df_slice.iloc[-1]
    lc  = float(row["Close"])

    ema_diff   = (row["ema10"] - row["ema20"]) / lc
    ema_s      = float(np.tanh(ema_diff * 20))

    rsi_val = float(row["rsi"])
    rsi_s   = -((rsi_val-65)/35) if rsi_val>65 else (35-rsi_val)/35 if rsi_val<35 else 0.0

    macd_s = float(np.tanh(row["macd_hist"] / (lc * 0.001 + 1e-9)))

    vr    = float(row["vol_ratio"]) if not math.isnan(float(row["vol_ratio"])) else 1.0
    vol_s = float(np.clip(vr - 1.0, -0.5, 0.5))

    ichi_sa = float(row["ichi_sa"]) if not math.isnan(float(row["ichi_sa"])) else lc
    ichi_sb = float(row["ichi_sb"]) if not math.isnan(float(row["ichi_sb"])) else lc
    kumo_top    = max(ichi_sa, ichi_sb)
    kumo_bottom = min(ichi_sa, ichi_sb)
    if lc > kumo_top:
        ichi_s = min(1.0, (lc - kumo_top) / (lc * 0.02 + 1e-9))
    elif lc < kumo_bottom:
        ichi_s = max(-1.0, (lc - kumo_bottom) / (lc * 0.02 + 1e-9))
    else:
        ichi_s = (lc - (kumo_top+kumo_bottom)/2) / (kumo_top-kumo_bottom+1e-9) * 0.3
    tk_s   = float(np.tanh(((float(row["ichi_tenkan"])-float(row["ichi_kijun"]))/lc)*30))
    ichi_s = ichi_s * 0.6 + tk_s * 0.4

    try:
        stoch_k = float(row["stoch_k"])
        if math.isnan(stoch_k): stoch_k = 50.0
    except (ValueError, TypeError):
        stoch_k = 50.0
    stoch_s = -((stoch_k-80)/20) if stoch_k>80 else (20-stoch_k)/20 if stoch_k<20 else 0.0

    adx_val    = float(row["adx"]) if not math.isnan(float(row["adx"])) else 20.0
    adx_factor = float(np.clip((adx_val-20)/30, -0.5, 1.0))

    return {"ema": ema_s, "rsi": rsi_s, "macd": macd_s, "vol": vol_s,
            "ichi": ichi_s, "stoch": stoch_s, "adx_factor": adx_factor}


def optimize_weights(df) -> dict:
    """Multi-start optimization: maximize directional accuracy on first 70% of data."""
    train_end = int(len(df) * 0.70)

    # Pre-compute signals on training window
    samples = []
    for i in range(30, train_end - 1):
        try:
            sig = calc_raw_signals(df.iloc[:i])
        except Exception:
            continue
        actual_open  = float(df.iloc[i]["Open"])
        actual_close = float(df.iloc[i]["Close"])
        actual_dir   = 1 if actual_close >= actual_open else -1
        samples.append((sig, actual_dir))

    if len(samples) < 10:
        return DEFAULT_WEIGHTS.copy()

    keys = ["ema", "rsi", "macd", "vol", "ichi", "stoch"]
    n    = len(keys)
    constraints = [{"type": "eq", "fun": lambda w: np.sum(w) - 1.0}]
    bounds = [(0.0, 0.5)] * n

    def neg_accuracy(w_arr):
        correct = 0
        for sig, actual_dir in samples:
            composite = (
                sig["ema"]   * w_arr[0] +
                sig["rsi"]   * w_arr[1] +
                sig["macd"]  * w_arr[2] +
                sig["vol"]   * w_arr[3] +
                sig["ichi"]  * w_arr[4] +
                sig["stoch"] * w_arr[5]
            ) * (1.0 + sig["adx_factor"] * 0.3)
            if (composite >= 0) == (actual_dir == 1):
                correct += 1
        return -correct / len(samples)

    best_val = neg_accuracy([DEFAULT_WEIGHTS[k] for k in keys])
    best_w   = [DEFAULT_WEIGHTS[k] for k in keys]

    # Multi-start: default + 30 random starts
    rng = np.random.default_rng(42)
    starts = [[DEFAULT_WEIGHTS[k] for k in keys]]
    for _ in range(30):
        r = rng.dirichlet(np.ones(n))           # random weights summing to 1
        r = np.clip(r, 0.0, 0.5)
        r = r / r.sum()
        starts.append(r.tolist())

    from scipy.optimize import minimize  # lazy — len pri prepočte váh
    for x0 in starts:
        res = minimize(neg_accuracy, x0, method="SLSQP",
                       bounds=bounds, constraints=constraints,
                       options={"maxiter": 300, "ftol": 1e-7})
        if res.fun < best_val:
            best_val = res.fun
            best_w   = res.x.tolist()

    # Normalize and round
    arr   = np.clip(best_w, 0.0, 0.5)
    arr   = arr / arr.sum()
    opt_w = {k: round(float(v), 4) for k, v in zip(keys, arr)}
    total = sum(opt_w.values())
    opt_w = {k: round(v / total, 4) for k, v in opt_w.items()}
    return opt_w


def run_backtest(df, weights: dict = None):
    """Backtest with train/test split — hit rate reported on test 30% only."""
    train_end = int(len(df) * 0.70)
    dates = list(df.index)
    df_r  = df.reset_index(drop=True)
    results = []

    for i in range(30, len(df_r) - 1):
        slice_df = df.iloc[:i]
        try:
            pred = predict_next_candle(slice_df, weights=weights)
        except Exception:
            continue

        actual       = df_r.iloc[i]
        prev_close   = float(df_r.iloc[i-1]["Close"])
        actual_open  = float(actual["Open"])
        actual_close = float(actual["Close"])
        pred_close   = pred["close"]

        # The model predicts the next close relative to the previous closed
        # candle. Evaluate against the same baseline. Using close >= open would
        # score candle color, not price direction, and gaps can flip the result.
        pred_direction   = 1 if pred_close >= prev_close else -1
        actual_direction = 1 if actual_close >= prev_close else -1
        contrib = {k: v["signal"] * v["weight"] for k, v in pred["signals"].items()}

        results.append({
            "date":         str(dates[i])[:10],
            "prev_close":   round(prev_close, 4),
            "pred_open":    round(pred["open"], 4),
            "pred_high":    round(pred["high"], 4),
            "pred_low":     round(pred["low"], 4),
            "pred_close":   round(pred_close, 4),
            "actual_open":  round(actual_open, 4),
            "actual_close": round(actual_close, 4),
            "correct":      pred_direction == actual_direction,
            "pred_dir":      pred_direction,
            "actual_dir":   actual_direction,
            "pred_change_pct":   round((pred_close - prev_close) / prev_close * 100, 3) if prev_close else None,
            "actual_change_pct": round((actual_close - prev_close) / prev_close * 100, 3) if prev_close else None,
            "error_pct":    round(abs(pred_close - actual_close) / actual_close * 100, 3),
            "composite":    round(pred["composite"], 4),
            "method":       pred.get("method", "technical_composite"),
            "analog":       pred.get("analog"),
            "contrib":      contrib,
            "is_test":      i >= train_end,
        })

    if not results:
        return {"error": "Not enough data for backtest"}

    # Overall accuracy (all data)
    total   = len(results)
    correct = sum(1 for r in results if r["correct"])

    # Test-only accuracy and hit rate (last 30%)
    test_results = [r for r in results if r["is_test"]]
    test_total   = len(test_results)
    test_correct = sum(1 for r in test_results if r["correct"])

    ind_hit = {}
    for k in ["ema", "rsi", "macd", "vol", "ichi", "stoch"]:
        hits = count = 0
        # Use only test set for hit rate
        for r in test_results:
            sig = r["contrib"].get(k, 0)
            if abs(sig) > 0.005:
                count += 1
                if (sig > 0 and r["actual_dir"] == 1) or (sig < 0 and r["actual_dir"] == -1):
                    hits += 1
        ind_hit[k] = round(hits / count * 100, 1) if count else None

    # Base rate "vždy hore" — poctivý benchmark pre direction_accuracy.
    # Merané (2026-07): žiaden variant modelu base rate neprekonáva, smer je drift.
    up_total = sum(1 for r in results if r["actual_dir"] == 1)
    test_up = sum(1 for r in test_results if r["actual_dir"] == 1)

    return {
        "total_predictions":      total,
        "direction_accuracy":     round(correct / total * 100, 1),
        "test_accuracy":          round(test_correct / test_total * 100, 1) if test_total else None,
        "test_total":             test_total,
        "base_rate_up":           round(up_total / total * 100, 1),
        "test_base_rate_up":      round(test_up / test_total * 100, 1) if test_total else None,
        "avg_error_pct":          round(sum(r["error_pct"] for r in results) / total, 2),
        "indicator_hit_rate":     ind_hit,
        "detail":                 results,
    }


# Backtest je CPU-heavy walk-forward slučka. Držíme iba niekoľko posledných
# výsledkov; kľúč sa zmení pri novej weekly sviečke alebo zmene váh.
_BACKTEST_CACHE: dict = {}
_BACKTEST_CACHE_LOCK = threading.Lock()
_BACKTEST_CACHE_MAX = 8 if _LOW_MEMORY else 24


def run_backtest_cached(df, weights: dict = None):
    effective = weights or DEFAULT_WEIGHTS
    weight_key = tuple(sorted((str(k), round(float(v), 6)) for k, v in effective.items()))
    # Dátum + počet riadkov nestačia: väčšina tickerov má rovnaký weekly kalendár
    # a predtým preto mohli omylom zdieľať backtest iného titulu. Fingerprint
    # OHLCV je lacný pri ~100–250 riadkoch a zároveň pokryje opravené historické dáta.
    source_cols = [col for col in ("Open", "High", "Low", "Close", "Volume") if col in df.columns]
    try:
        hashed = pd.util.hash_pandas_object(df[source_cols], index=True).values.tobytes()
        data_key = hashlib.blake2b(hashed, digest_size=12).hexdigest()
    except Exception:
        data_key = hashlib.blake2b(repr((tuple(df.index), len(df))).encode("utf-8"), digest_size=12).hexdigest()
    key = (data_key, weight_key)
    with _BACKTEST_CACHE_LOCK:
        cached = _BACKTEST_CACHE.get(key)
        if cached is not None:
            return cached
    result = run_backtest(df, weights=effective)
    with _BACKTEST_CACHE_LOCK:
        if len(_BACKTEST_CACHE) >= _BACKTEST_CACHE_MAX:
            oldest = next(iter(_BACKTEST_CACHE), None)
            if oldest is not None:
                _BACKTEST_CACHE.pop(oldest, None)
        _BACKTEST_CACHE[key] = result
    return result

# ── Serialize ─────────────────────────────────────────────────────────────────

def safe_float(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    return round(float(v), 4)

def df_to_candles(df):
    candles = []
    for ts, row in df.iterrows():
        o = safe_float(row["Open"])
        h = safe_float(row["High"])
        l = safe_float(row["Low"])
        c = safe_float(row["Close"])
        if None in (o, h, l, c):
            continue
        candles.append({
            "time":   int(pd.Timestamp(ts).timestamp()),
            "open":   o, "high": h, "low": l, "close": c,
            "volume": safe_float(row.get("Volume")),
        })
    return candles

# ── Routes ────────────────────────────────────────────────────────────────────

# ══ END PREDICTIVE CHART functions ══════════════════════════

ALLOWED_CORS_ORIGINS = [
    origin.strip()
    for origin in os.getenv(
        "CORS_ALLOWED_ORIGINS",
        "https://dashboard-yvb5.onrender.com",
    ).split(",")
    if origin.strip()
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_CORS_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)

# GZip — frontend moduly a väčšie JSON odpovede sa oplatí komprimovať; malé
# JSON payloady necháva minimum_size=1000 na pokoji.
from fastapi.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=1000)

# ── Basic Auth — registruje sa pri štarte, env vars sú dostupné ──────────────
import base64 as _b64, secrets as _secrets
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import Response as _SR

class _BasicAuth(BaseHTTPMiddleware):
    COOKIE_NAME = "td_auth"

    def __init__(self, app):
        super().__init__(app)
        # Render values are read when the middleware instance is created.
        # Strip accidental spaces/newlines added while editing secret env vars.
        self._user = os.getenv("DASH_USER", "").strip()
        self._pass = os.getenv("DASH_PASS", "").strip()
        self._cookie_value = self._make_cookie_value()
        print(
            "  Basic Auth middleware: "
            f"{'zapnuta' if self._user else 'vypnuta'} "
            f"(user_len={len(self._user)}, pass_len={len(self._pass)})"
        )

    def _make_cookie_value(self):
        if not self._user or not self._pass:
            return ""
        secret = f"{self._user}:{self._pass}".encode("utf-8")
        return hmac.new(secret, b"trading-dashboard-basic-session", hashlib.sha256).hexdigest()

    def _has_session_cookie(self, request):
        cookie = request.cookies.get(self.COOKIE_NAME, "")
        return bool(self._cookie_value and hmac.compare_digest(cookie, self._cookie_value))

    async def dispatch(self, request, call_next):
        # Verejné endpointy — preskočiť Basic Auth (chránené vlastným tokenom)
        if request.url.path.startswith("/api/public/"):
            return await call_next(request)
        if not self._user:          # Auth vypnutá ak DASH_USER nie je nastavený
            return await call_next(request)
        if self._has_session_cookie(request):
            return await call_next(request)
        auth = request.headers.get("Authorization", "")
        scheme, _, encoded = auth.partition(" ")
        if scheme.lower() == "basic" and encoded:
            try:
                u, p = _b64.b64decode(encoded, validate=True).decode("utf-8").split(":", 1)
                if (_secrets.compare_digest(u, self._user) and
                        _secrets.compare_digest(p, self._pass)):
                    response = await call_next(request)
                    response.set_cookie(
                        self.COOKIE_NAME,
                        self._cookie_value,
                        max_age=60 * 60 * 24 * 30,
                        httponly=True,
                        samesite="lax",
                        secure=request.url.scheme == "https",
                    )
                    return response
            except Exception:
                pass
        return _SR(status_code=401,
                   headers={"WWW-Authenticate": 'Basic realm="Trading Dashboard"'})

app.add_middleware(_BasicAuth)

def _public_client_key(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for", "")
    if forwarded:
        return forwarded.split(",", 1)[0].strip()
    return request.client.host if request.client else "unknown"

_PUBLIC_RATE_MAX_KEYS = 1000  # strop proti rastu pamäte pri spoofovaných XFF adresách

def _check_public_rate_limit(request: Request):
    now = _time_module.time()
    cutoff = now - PUBLIC_RATE_LIMIT_WINDOW
    key = _public_client_key(request)
    with _public_rate_lock:
        # Priebežné čistenie: expirované kľúče inak ostávajú do reštartu procesu.
        for stale_key in [k for k, ts_list in _public_rate.items() if k != key and (not ts_list or ts_list[-1] < cutoff)]:
            del _public_rate[stale_key]
        if key not in _public_rate and len(_public_rate) >= _PUBLIC_RATE_MAX_KEYS:
            _public_rate.clear()
        hits = [ts for ts in _public_rate.get(key, []) if ts >= cutoff]
        if len(hits) >= PUBLIC_RATE_LIMIT_MAX:
            _public_rate[key] = hits
            raise HTTPException(status_code=429, detail="Too many requests")
        hits.append(now)
        _public_rate[key] = hits

def _public_token_from_headers(
    authorization: str | None,
    x_api_token: str | None,
) -> str:
    # Token výhradne v hlavičke — query string sa loguje (proxy, Render, história).
    if authorization and authorization.lower().startswith("bearer "):
        return authorization[7:].strip()
    if x_api_token:
        return x_api_token.strip()
    return ""

# ── PUBLIC ENDPOINT — pre Claude / externý prístup ────────────────────────────
@app.get("/api/public/portfolio")
def get_public_portfolio(
    request: Request,
    account: str = Query("1"),
    authorization: str | None = Header(None),
    x_api_token: str | None = Header(None, alias="X-API-Token"),
):
    """
    Verejný endpoint chránený tokenom (nie Basic Auth).
    Vráti aktuálne pozície + summary pre daný účet.
    Pouzitie: Authorization: Bearer <token> alebo X-API-Token hlavička.
    """
    _check_public_rate_limit(request)
    provided_token = _public_token_from_headers(authorization, x_api_token)
    if not PUBLIC_API_TOKEN or not _secrets.compare_digest(provided_token, PUBLIC_API_TOKEN):
        raise HTTPException(status_code=403, detail="Invalid token")
    if account not in ("1", "2"):
        raise HTTPException(status_code=400, detail="Invalid account")

    # Jediný zdroj pravdy: rovnaký processed snapshot ako Portfólio tab.
    # get_portfolio rieši RAM/disk cache, live fetch aj stale fallback; verejná
    # cesta už nesmie mať vlastnú summary logiku ani starý raw cache formát.
    try:
        payload = get_portfolio(account=account, refresh=0)
        cache_entry = _positions_cache.get(account) or {}
        source = "stale_cache" if payload.get("stale") else ("cache" if payload.get("cached") else "live")
        return {
            "positions": payload.get("positions", []),
            "summary": payload.get("summary", {}),
            "timestamp": cache_entry.get("ts"),
            "source": source,
        }
    except HTTPException:
        raise
    except Exception as e:
        err_type = type(e).__name__
        raise HTTPException(502, f"Portfolio nedostupné ({err_type}) — skontrolujte eToro proxy")

PRESETS_FILE = str(DATA_ROOT / "presets.json")
DAILY_INTERVALS = {"1d", "5d", "1wk", "1mo", "3mo"}
YF_HEADERS = {"User-Agent": "Mozilla/5.0"}

# ── PRESETS ───────────────────────────────────────────────────────────────────

def read_presets() -> dict:
    try:
        with open(PRESETS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}
    except Exception as e:
        raise HTTPException(500, f"Chyba citania presetov: {e}")

def write_presets(data: dict):
    try:
        with open(PRESETS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        raise HTTPException(500, f"Chyba zapisu presetov: {e}")

@app.get("/api/presets")
def get_presets():
    return read_presets()

@app.put("/api/presets/{name}")
def save_preset_put(name: str):
    return {"error": "Pouzi POST"}, 405

@app.post("/api/presets/{name}")
async def upsert_preset(name: str, request: Request):
    body = await request.json()
    presets = read_presets()
    presets[name] = body
    write_presets(presets)
    return {"ok": True, "name": name, "count": len(body)}

@app.delete("/api/presets/{name}")
def delete_preset(name: str):
    presets = read_presets()
    if name not in presets:
        raise HTTPException(404, f"Preset '{name}' neexistuje")
    del presets[name]
    write_presets(presets)
    return {"ok": True, "deleted": name}

# ── NEWS ─────────────────────────────────────────────────────────────────────

@app.get("/api/news")
def get_news(symbol: str = Query(...)):
    sym = symbol.upper().strip()
    url = (
        f"https://query2.finance.yahoo.com/v1/finance/search"
        f"?q={requests.utils.quote(sym)}&quotesCount=0&newsCount=8"
        f"&enableFuzzyQuery=false&enableCb=false"
    )
    try:
        resp = requests.get(url, headers=YF_HEADERS, timeout=6)
        resp.raise_for_status()
        news = resp.json().get("news", [])
    except Exception as e:
        raise HTTPException(502, f"News fetch zlyhalo: {e}")

    result = []
    for item in news:
        pub = item.get("providerPublishTime", 0)
        result.append({
            "title":     item.get("title", ""),
            "url":       item.get("link", ""),
            "source":    item.get("publisher", ""),
            "published": pub,
        })
    return result

# ── ETORO INTERVAL MAPPING ───────────────────────────────────────────────────

# Yahoo Finance period/interval -> eToro candlesCount + interval
YAHOO_TO_ETORO_INTERVAL = {
    "1m":  "OneMinute",
    "5m":  "FiveMinutes",
    "15m": "FiveMinutes",
    "30m": "ThirtyMinutes",
    "1h":  "OneHour",
    "4h":  "FourHours",
    "1d":  "OneDay",
    "1wk": "OneWeek",
    "1mo": "OneWeek",   # eToro nemá monthly — použijeme weekly
}

YAHOO_PERIOD_TO_COUNT = {
    # (period, interval) -> candlesCount  (eToro max = 1000)
    ("1d",  "1m"):   390,
    ("5d",  "5m"):   390,
    ("1mo", "30m"):  320,
    ("3mo", "1h"):   500,
    ("6mo", "1d"):   130,
    ("1y",  "1d"):   252,
    ("2y",  "1d"):   504,
    ("5y",  "1d"):   1000,
    ("max", "1d"):   1000,
    ("2y",  "1wk"):  104,
    ("5y",  "1wk"):  260,
    ("max", "1wk"):  1000,
    # 4h resampling
    ("1y",  "4h"):   500,
    ("6mo", "4h"):   250,
    ("2y",  "4h"):   1000,
}

AUTO_INTERVAL_TO_COUNT = {
    "1m":   390,
    "5m":   1000,
    "15m":  1000,
    "30m":  1000,
    "1h":   1000,
    "4h":   1000,
    "12h":  1000,
    "1d":   1000,
    "1wk":  1000,
    "1mo":  1000,
}

# ── ETORO INSTRUMENT SEARCH ──────────────────────────────────────────────────

# ── TICKER SEARCH ─────────────────────────────────────────────────────────────

@app.get("/api/search")
def search_ticker_yahoo(q: str = Query(..., min_length=1)):
    """Yahoo Finance autocomplete — fuzzy, rýchly, spoľahlivý."""
    url = (
        "https://query2.finance.yahoo.com/v1/finance/search"
        f"?q={requests.utils.quote(q)}&quotesCount=12&newsCount=0"
        "&enableFuzzyQuery=true&enableCb=false&enableNavLinks=false"
    )
    try:
        resp = requests.get(url, headers=YF_HEADERS, timeout=6)
        resp.raise_for_status()
        quotes = resp.json().get("quotes", [])
    except Exception as e:
        raise HTTPException(502, f"Yahoo Finance search zlyhalo: {e}")

    results = []
    for item in quotes:
        sym = item.get("symbol", "")
        if not sym:
            continue
        results.append({
            "symbol":   sym,
            "name":     item.get("shortname") or item.get("longname") or sym,
            "type":     item.get("quoteType", ""),
            "exchange": item.get("exchDisp") or item.get("exchange", ""),
        })
    return results


# ── ETORO INTEGRÁCIA ──────────────────────────────────────────────────────────

ETORO_PROXY   = "http://localhost:8765"
ETORO_PROXY_TIMEOUT = 10

import time, json as _json, threading, gzip as _gzip, hashlib as _hashlib
from pathlib import Path as _Path

# ── DISK CACHE ────────────────────────────────────────────────────────────────
CACHE_DIR = _Path(DATA_ROOT) / "cache"
CACHE_DIR.mkdir(exist_ok=True)
(CACHE_DIR / "ohlcv").mkdir(exist_ok=True)
(CACHE_DIR / "portfolio").mkdir(exist_ok=True)
_cache_mem: dict[str, tuple[float, dict]] = {}   # key → (mtime, data)
_cache_mem_lock = threading.Lock()
_CACHE_MEM_MAX = 50    # max položiek v RAM cache (znížené z 75 — RAM rezerva)
# Pevný počet lock stripes: chráni súbory pred concurrent read/write bez toho,
# aby slovník lockov rástol s každým tickerom počas života procesu.
_CACHE_FILE_LOCKS = tuple(threading.Lock() for _ in range(32))
WATCHLIST_PATH = _Path(DATA_ROOT) / "watchlist.json"
_watchlist_lock = threading.Lock()


def _cache_file_lock(path: _Path):
    digest = hashlib.blake2b(str(path).encode("utf-8"), digest_size=2).digest()
    return _CACHE_FILE_LOCKS[int.from_bytes(digest, "big") % len(_CACHE_FILE_LOCKS)]

def _normalize_watchlist_items(items):
    out = []
    seen = set()
    if not isinstance(items, list):
        return out
    for item in items:
        if isinstance(item, str):
            item = {"symbol": item}
        if not isinstance(item, dict):
            continue
        symbol = str(item.get("symbol") or "").strip().upper()
        if not symbol or symbol in seen:
            continue
        seen.add(symbol)
        clean = {"symbol": symbol}
        for key in ("name", "price", "chg", "lastUpdated", "instrumentId", "addedAt"):
            if item.get(key) is not None:
                clean[key] = item.get(key)
        out.append(clean)
    return out

def _read_watchlist_file():
    with _watchlist_lock:
        try:
            if WATCHLIST_PATH.exists():
                data = _json.loads(WATCHLIST_PATH.read_text(encoding="utf-8"))
                return _normalize_watchlist_items(data.get("items", data if isinstance(data, list) else []))
        except Exception as e:
            print(f"  watchlist read error: {e}")
    return []

def _write_watchlist_file(items):
    clean = _normalize_watchlist_items(items)
    payload = {"items": clean, "updatedAt": datetime.now(timezone.utc).isoformat()}
    with _watchlist_lock:
        tmp = WATCHLIST_PATH.with_suffix(".json.tmp")
        tmp.write_text(_json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(WATCHLIST_PATH)
    return payload

@app.get("/api/watchlist")
def get_watchlist():
    return {"items": _read_watchlist_file()}

@app.put("/api/watchlist")
def put_watchlist(body: dict):
    return _write_watchlist_file(body.get("items", []))

def cache_write(path: _Path, data: dict):
    """Ulož JSON.gz atómovo; čitateľ nikdy neuvidí rozpracovaný gzip."""
    target = _Path(str(path) + ".gz")
    tmp = target.with_name(
        f"{target.name}.{threading.get_ident()}.{_time_module.time_ns()}.tmp"
    )
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        with _cache_file_lock(path):
            with _gzip.open(tmp, "wt", encoding="utf-8") as f:
                _json.dump(data, f)
            os.replace(tmp, target)
            key = str(path)
            with _cache_mem_lock:
                _cache_mem[key] = (target.stat().st_mtime, data)
                # Evict najstaršie záznamy ak prekročíme limit
                if len(_cache_mem) > _CACHE_MEM_MAX:
                    oldest = sorted(_cache_mem, key=lambda k: _cache_mem[k][0])
                    for k in oldest[:len(_cache_mem) - _CACHE_MEM_MAX]:
                        del _cache_mem[k]
    except Exception as e:
        print(f"  cache_write error: {e}")
    finally:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass

def cache_read(path: _Path) -> dict | None:
    """Načítaj JSON.gz z disku. Vráti None ak neexistuje."""
    try:
        p = _Path(str(path) + ".gz")
        if not p.exists():
            return None
        with _cache_file_lock(path):
            if not p.exists():
                return None
            mtime = p.stat().st_mtime
            key = str(path)
            with _cache_mem_lock:
                cached = _cache_mem.get(key)
                if cached and cached[0] == mtime:
                    return cached[1]
            with _gzip.open(str(p), "rt", encoding="utf-8") as f:
                data = _json.load(f)
            with _cache_mem_lock:
                _cache_mem[key] = (mtime, data)
            return data
    except Exception as e:
        print(f"  cache_read error: {e}")
        return None

def cache_age_seconds(path: _Path) -> float:
    """Vek súboru v sekundách. inf ak neexistuje."""
    p = _Path(str(path) + ".gz")
    if not p.exists():
        return float("inf")
    return time.time() - p.stat().st_mtime

# Cache pre instrument mapping (11MB — načítame raz, ukladáme na disk)
_instruments_cache: dict = {}
_instruments_loaded = False
_instruments_lock = threading.Lock()
INSTRUMENTS_CACHE_PATH = CACHE_DIR / "instruments"
INSTRUMENTS_CACHE_TTL  = 86400  # 24 hodín

# Cache pre pozície — RAM + disk
import time
_positions_cache: dict = {}
# 24h (2026-07-09, raised from 30 min): zoznam pozícií/objednávok sa pri
# bežnom štýle (max niekoľko obchodov týždenne) mení zriedka — živé ceny a P/L
# idú aj tak nezávisle cez WS tick vrstvu (estimatePositionLivePnl), takže
# dlhší TTL neznamená zastarané čísla na obrazovke, len menej zbytočných
# eToro round-tripov pri otváraní grafov.
# Manuálne obnovenie: ⟳ v Portfóliu posiela refresh=1 a obíde cache úplne.
POSITIONS_CACHE_TTL = 86400

# Typy nástrojov ktoré chceme (akcie + ETF) — podľa InstrumentTypeID
# 1=Forex, 2=CFD, 3=Crypto, 4=Commodity, 5=Index, 6=ETF, 7=Stocks
ALLOWED_INSTRUMENT_TYPES = {5, 6, 7}   # Index, ETF, Stocks

# eToro SymbolFull → Yahoo Finance ticker (špeciálne prípady)
SYMBOL_OVERRIDES = {
    # eToro používa bez sufixu, Yahoo niekedy potrebuje iný tvar
    "GOOGL": "GOOGL",
    "GOOG":  "GOOG",
    # Doplň podľa potreby ak niečo nefunguje
}

def load_instruments():
    global _instruments_cache, _instruments_loaded
    if _instruments_loaded:
        return _instruments_cache
    with _instruments_lock:
        if _instruments_loaded:
            return _instruments_cache

        def restore_disk(payload):
            rows = payload.get("items", []) if isinstance(payload, dict) else []
            restored = {}
            for row in rows:
                try:
                    iid = int(row["instrumentId"])
                except (KeyError, TypeError, ValueError):
                    continue
                restored[iid] = {
                    "symbol": row.get("symbol", ""),
                    "name": row.get("name", ""),
                    "typeID": row.get("typeID", 0),
                    "logo": row.get("logo", ""),
                }
            return restored

        disk_payload = cache_read(INSTRUMENTS_CACHE_PATH)
        if cache_age_seconds(INSTRUMENTS_CACHE_PATH) < INSTRUMENTS_CACHE_TTL:
            restored = restore_disk(disk_payload)
            if restored:
                _instruments_cache = restored
                _instruments_loaded = True
                print(f"  Instruments z disk cache: {len(restored)} záznamov")
                return _instruments_cache

        try:
            resp = requests.get(f"{ETORO_PROXY}/instruments", timeout=20)
            resp.raise_for_status()
            items = resp.json().get("InstrumentDisplayDatas", [])
            compact = {}
            for item in items:
                iid = item["InstrumentID"]
                images = item.get("Images", [])
                logo = ""
                for img in sorted(images, key=lambda x: x.get("Width", 999) * x.get("Height", 999)):
                    uri = img.get("Uri", "")
                    if uri:
                        logo = uri
                        break
                compact[iid] = {
                    "symbol": item.get("SymbolFull", ""),
                    "name": item.get("InstrumentDisplayName", ""),
                    "typeID": item.get("InstrumentTypeID", 0),
                    "logo": logo,
                }
            if compact:
                _instruments_cache = compact
                _instruments_loaded = True
                cache_write(INSTRUMENTS_CACHE_PATH, {
                    "schema": 1,
                    "items": [dict(instrumentId=iid, **meta) for iid, meta in compact.items()],
                })
                print(f"  Instruments načítané: {len(compact)} záznamov")
        except Exception as e:
            restored = restore_disk(disk_payload)
            if restored:
                _instruments_cache = restored
                _instruments_loaded = True
                print(f"  Instruments stale disk fallback: {len(restored)} záznamov")
            else:
                print(f"  WARN: Instruments cache zlyhalo: {e}")
    return _instruments_cache


@app.get("/api/logo-map")
def get_logo_map():
    instruments = load_instruments()
    return {v["symbol"]: v["logo"] for v in instruments.values() if v.get("symbol") and v.get("logo")}


def etoro_symbol_to_yf(symbol: str) -> str:
    """Konvertuje eToro SymbolFull na Yahoo Finance ticker."""
    if symbol in SYMBOL_OVERRIDES:
        return SYMBOL_OVERRIDES[symbol]
    return symbol


@app.get("/api/etoro/positions")
def get_etoro_positions(account: str = Query("1"), refresh: int = Query(0)):
    """Kompatibilný stock/ETF pohľad nad jednotným portfóliovým snapshotom."""
    portfolio = get_portfolio(account=account, refresh=refresh)
    positions = []
    for pos in portfolio.get("positions", []):
        if pos.get("type") not in {"Stock", "ETF"}:
            continue
        positions.append({
            "instrumentId": pos.get("instrumentId"),
            "symbol": pos.get("symbol"),
            "name": pos.get("name"),
            "openDate": (pos.get("openDateTime") or "")[:10],
            "openRate": pos.get("openRate"),
            "currentRate": pos.get("currentRate"),
            "pnl": pos.get("pnl"),
            "isBuy": pos.get("isBuy", True),
            "amount": pos.get("amount"),
            "units": pos.get("units"),
            "positionID": pos.get("positionId"),
        })

    summary = dict(portfolio.get("summary") or {})
    summary["positions_count"] = len(positions)
    response = {
        "positions": positions,
        "summary": summary,
        "cached": bool(portfolio.get("cached")),
    }
    if portfolio.get("stale"):
        response["stale"] = True
    return response
@app.get("/api/etoro/accounts")
def etoro_accounts():
    """Vráti zoznam dostupných eToro účtov z proxy."""
    try:
        resp = requests.get(f"{ETORO_PROXY}/accounts", timeout=3)
        return resp.json() if resp.ok else []
    except Exception as e:
        return []

# eToro username mapping — hardcoded (user-info endpoints vyžadujú username, nie CID)
ETORO_USERNAMES = {"1": "DD1973", "2": "nelka39"}

@app.get("/api/etoro/gain")
def get_etoro_gain(account: str = Query("1")):
    """Mesacna a rocna vykonnost uctu v % — /user-info/people/{username}/gain"""
    username = ETORO_USERNAMES.get(account)
    if not username:
        raise HTTPException(400, f"Nezname ucet: {account}")
    try:
        resp = fetch_with_retry(
            f"{ETORO_PROXY}/etoro/user-info/people/{username}/gain?account={account}",
            timeout=10, retries=2
        )
        return resp.json()
    except Exception as e:
        raise HTTPException(502, f"eToro gain zlyhalo: {e}")

@app.get("/api/etoro/daily-gain")
def get_etoro_daily_gain(
    account: str = Query("1"),
    minDate: str = Query(""),
    maxDate: str = Query(""),
    gain_type: str = Query("Daily", alias="type"),
):
    """Denny alebo suhrnny % gain za obdobie — /user-info/people/{username}/daily-gain"""
    username = ETORO_USERNAMES.get(account)
    if not username:
        raise HTTPException(400, f"Nezname ucet: {account}")
    if not minDate:
        minDate = (datetime.now(timezone.utc) - timedelta(days=365)).date().isoformat()
    if not maxDate:
        maxDate = datetime.now(timezone.utc).date().isoformat()
    try:
        url = (f"{ETORO_PROXY}/etoro/user-info/people/{username}/daily-gain"
               f"?minDate={requests.utils.quote(minDate)}&maxDate={requests.utils.quote(maxDate)}"
               f"&type={gain_type}&account={account}")
        resp = fetch_with_retry(url, timeout=10, retries=2)
        return resp.json()
    except Exception as e:
        raise HTTPException(502, f"eToro daily-gain zlyhalo: {e}")

@app.get("/api/etoro/watchlists")
def get_etoro_watchlists(account: str = Query("1")):
    try:
        resp = requests.get(f"{ETORO_PROXY}/etoro/watchlists?ensureBuiltinWatchlists=true&account={account}", timeout=8)
        if not resp.ok: raise HTTPException(502, f"eToro watchlists: {resp.status_code}")
        result = []
        for wl in resp.json().get("watchlists", []):
            items = []
            for it in wl.get("items", []):
                mkt = it.get("market") or {}
                sym = mkt.get("symbolName") or ""
                items.append({"instrumentId": it.get("itemId"), "symbol": sym, "name": mkt.get("displayName") or sym})
            result.append({"id": wl.get("id"), "name": wl.get("name"), "items": items})
        return result
    except HTTPException: raise
    except Exception as e: raise HTTPException(502, str(e))


@app.post("/api/etoro/watchlists/{watchlist_id}/items")
def add_to_etoro_watchlist(watchlist_id: str, body: dict, account: str = Query("1")):
    iid = body.get("instrumentId")
    if not iid: raise HTTPException(400, "instrumentId required")
    import json as _json
    payload = _json.dumps([{"itemId": iid, "itemType": "Instrument"}]).encode()
    try:
        resp = requests.post(f"{ETORO_PROXY}/etoro/watchlists/{watchlist_id}/items?account={account}",
            data=payload, headers={"Content-Type": "application/json"}, timeout=8)
        return {"ok": resp.ok, "status": resp.status_code}
    except Exception as e: raise HTTPException(502, str(e))


@app.delete("/api/etoro/watchlists/{watchlist_id}/items/{instrument_id}")
def remove_from_etoro_watchlist(watchlist_id: str, instrument_id: int, account: str = Query("1")):
    import json as _json
    payload = _json.dumps([{"itemId": instrument_id, "itemType": "Instrument"}]).encode()
    try:
        resp = requests.delete(f"{ETORO_PROXY}/etoro/watchlists/{watchlist_id}/items?account={account}",
            data=payload, headers={"Content-Type": "application/json"}, timeout=8)
        return {"ok": resp.ok, "status": resp.status_code}
    except Exception as e: raise HTTPException(502, str(e))


@app.get("/api/etoro/rates")
def get_etoro_rates(instrument_ids: str = Query(...), account: str = Query("1")):
    try:
        resp = requests.get(
            f"{ETORO_PROXY}/etoro/market-data/instruments/rates?instrumentIds={instrument_ids}&account={account}",
            timeout=8)
        if not resp.ok: raise HTTPException(502, f"eToro rates: {resp.status_code}")
        return resp.json()
    except HTTPException: raise
    except Exception as e: raise HTTPException(502, str(e))


@app.get("/api/etoro/rates-batch")
def get_etoro_rates_batch(symbols: str = Query(""), instrument_ids: str = Query(""), account: str = Query("1")):
    """Batch live rates pre symboly alebo instrumentId. eToro limit je 100 id v requeste."""
    id_to_symbol = {}
    ids = []

    for raw_id in [x.strip() for x in instrument_ids.split(",") if x.strip()]:
        try:
            iid = int(raw_id)
            ids.append(iid)
            id_to_symbol[iid] = raw_id
        except ValueError:
            continue

    for sym in [x.strip().upper() for x in symbols.split(",") if x.strip()]:
        iid = get_instrument_id(sym, account)
        if iid is not None:
            ids.append(iid)
            id_to_symbol[iid] = sym

    ids = list(dict.fromkeys(ids))
    if not ids:
        return {"rates": [], "count": 0}

    all_rates = []
    try:
        for i in range(0, len(ids), 100):
            chunk = ids[i:i + 100]
            resp = requests.get(
                f"{ETORO_PROXY}/etoro/market-data/instruments/rates?instrumentIds={','.join(map(str, chunk))}&account={account}",
                timeout=8
            )
            if not resp.ok:
                raise HTTPException(502, f"eToro rates: {resp.status_code}")
            data = resp.json()
            rate_items = data if isinstance(data, list) else (
                data.get("rates") or data.get("Rates") or data.get("instrumentRates") or []
            )
            for rate in rate_items:
                iid = rate.get("instrumentID") or rate.get("instrumentId") or rate.get("InstrumentID")
                all_rates.append({
                    "instrumentId": iid,
                    "symbol": id_to_symbol.get(iid, str(iid)),
                    "bid": rate.get("bid") or rate.get("Bid"),
                    "ask": rate.get("ask") or rate.get("Ask"),
                    "last": rate.get("lastExecution") or rate.get("LastExecution"),
                    "date": rate.get("date") or rate.get("Date"),
                    "conversionRateAsk": rate.get("conversionRateAsk") or rate.get("ConversionRateAsk"),
                    "conversionRateBid": rate.get("conversionRateBid") or rate.get("ConversionRateBid"),
                })
        return {"rates": all_rates, "count": len(all_rates)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(502, str(e))


@app.get("/api/etoro/instrument-id")
def resolve_instrument_id(symbol: str = Query(...), account: str = Query("1")):
    iid = get_instrument_id(symbol, account)
    if iid is None: raise HTTPException(404, f"Symbol nenajdeny")
    return {"symbol": symbol.upper(), "instrumentId": iid}


@app.get("/api/etoro/ws-keys")
def get_ws_keys(account: str = Query("1")):
    """Vráti API kľúče pre WebSocket autentifikáciu."""
    try:
        resp = requests.get(f"{ETORO_PROXY}/keys?account={account}", timeout=3)
        if resp.ok:
            return resp.json()
        # /keys endpoint nie je dostupný — vráť prázdne (WS nebude autentifikovaný)
        return {"api_key": None, "user_key": None}
    except Exception:
        return {"api_key": None, "user_key": None}


PORTFOLIO_DISK_CACHE_DIR = CACHE_DIR / "portfolio"


def _portfolio_disk_path(account: str):
    return PORTFOLIO_DISK_CACHE_DIR / f"processed_{account}"


@app.get("/api/etoro/portfolio")
def get_portfolio(account: str = Query("1"), refresh: int = Query(0)):
    """Vracia kompletné portfolio dáta pre portfolio panel — pozície + mirrors.
    RAM cache (POSITIONS_CACHE_TTL) je zálohovaná na disk, aby cold start
    (Render free tier po nečinnosti) nemusel vždy čakať na live eToro round-trip —
    poslednou známy stav sa načíta z disku hneď pri prvom requeste po reštarte."""
    disk_path = _portfolio_disk_path(account)
    cached = _positions_cache.get(account)
    if not cached:
        disk = cache_read(disk_path)
        if disk:
            _positions_cache[account] = disk
            cached = disk
    if cached and not refresh and (time.time() - cached["ts"]) < POSITIONS_CACHE_TTL:
        return {"positions": cached["data"], "summary": cached["summary"],
                "mirrors": cached.get("mirrors", []),
                "orders": cached.get("orders", []), "cached": True}

    instruments = load_instruments()

    try:
        resp = requests.get(f"{ETORO_PROXY}/pnl/real?account={account}", timeout=ETORO_PROXY_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        # Stale-while-erroring: eToro proxy výpadok nesmie zhodiť graf/portfólio,
        # keď máme (aj staršie) dáta z predošlého úspešného fetchu.
        if cached:
            return {"positions": cached["data"], "summary": cached["summary"],
                    "mirrors": cached.get("mirrors", []),
                    "orders": cached.get("orders", []), "cached": True, "stale": True}
        raise HTTPException(502, f"eToro proxy nedostupný: {e}")

    port = data.get("clientPortfolio", {})
    positions_raw = port.get("positions", [])
    mirrors_raw   = port.get("mirrors", [])

    # Spracuj pozície
    result = []
    for pos in positions_raw:
        iid  = pos.get("instrumentID") or pos.get("instrumentId")
        info = instruments.get(iid, {})
        sym  = info.get("symbol") or str(iid)
        name = info.get("name") or sym
        asset_class = info.get("typeID") or info.get("InstrumentTypeID") or 0
        # Typ: 1=Forex, 5=Stocks, 6=ETF, 10=Crypto, 4=Commodities, 2=Indices
        type_map = {1:"Forex", 2:"Index", 4:"Commodity", 5:"Stock", 6:"ETF", 10:"Crypto"}
        asset_type = type_map.get(asset_class, "Other")
        unrealized = pos.get("unrealizedPnL") if isinstance(pos.get("unrealizedPnL"), dict) else {}
        pnl = unrealized.get("pnL") if unrealized else pos.get("pnL") or 0
        pnl = pnl or 0
        amount = pos.get("amount") or 0
        raw_current_rate = (
            pos.get("currentRate")
            or pos.get("closeRate")
            or unrealized.get("closeRate")
            or unrealized.get("currentRate")
        )
        current_rate = raw_current_rate if (isinstance(raw_current_rate, (int, float)) and raw_current_rate > 0) else None
        units_val = pos.get("units") or 0
        is_buy = pos.get("isBuy", True)
        if current_rate is None:
            # Fallback: odhad z openRate + PnL + units (BUY/SELL)
            try:
                open_rate = float(pos.get("openRate") or 0)
                u = float(units_val)
                if open_rate > 0 and u > 0 and pnl is not None:
                    delta = float(pnl) / u
                    current_rate = open_rate + delta if is_buy else open_rate - delta
                    if current_rate <= 0:
                        current_rate = None
            except Exception:
                current_rate = None
        prev_close = _get_market_prev_close(sym, asset_type) if sym and not sym.isdigit() else None
        if prev_close and current_rate and units_val:
            daily_pnl = round((current_rate - prev_close) * float(units_val) * (1 if is_buy else -1), 2)
        else:
            daily_pnl = 0.0

        result.append({
            "positionId":   pos.get("positionID"),
            "instrumentId": iid,
            "symbol":       sym,
            "name":         name,
            "type":         asset_type,
            "isBuy":        is_buy,
            "openDateTime": pos.get("openDateTime", ""),
            "amount":       round(amount, 2),
            "units":        units_val,
            "openRate":     pos.get("openRate"),
            "currentRate":  current_rate,
            "previousClose": prev_close,
            "dailyPnl":     daily_pnl,
            "pnl":          round(pnl, 2),
            "pnlPct":       round(pnl / amount * 100, 2) if amount else 0,
            "fees":         round(pos.get("totalFees") or pos.get("fees") or 0, 2),
            "leverage":     pos.get("leverage", 1),
            "stopLoss":     pos.get("stopLossRate"),
            "takeProfit":   pos.get("takeProfitRate"),
        })
    result.sort(key=lambda x: (x["symbol"], x["openDateTime"]))

    # Spracuj mirrors
    mirrors = []
    for m in mirrors_raw:
        mir_pnl  = sum((pp.get("unrealizedPnL") or {}).get("pnL", 0) or 0 for pp in m.get("positions", []))
        mir_pnl += m.get("closedPositionsNetProfit") or 0
        mir_amt  = m.get("amount") or 0
        mirrors.append({
            "mirrorId":   m.get("mirrorID") or m.get("mirrorId"),
            "name":       m.get("fullName") or m.get("name") or f"Mirror #{m.get('mirrorID')}",
            "amount":     round(mir_amt, 2),
            "pnl":        round(mir_pnl, 2),
            "pnlPct":     round(mir_pnl / mir_amt * 100, 2) if mir_amt else 0,
            "closedPnl":  round(m.get("closedPositionsNetProfit") or 0, 2),
            "posCount":   len(m.get("positions", [])),
        })

    # Čakajúce objednávky — orders (limitky) + ordersForOpen (market objednávky
    # čakajúce na exekúciu, napr. mimo obchodných hodín); mirror orders vynechané
    type_map_orders = {1: "Forex", 2: "Index", 4: "Commodity", 5: "Stock", 6: "ETF", 10: "Crypto"}
    orders = []
    for kind, raw_list in (("limit", port.get("orders") or []),
                           ("market", port.get("ordersForOpen") or [])):
        for o in raw_list:
            if kind == "market" and (o.get("mirrorID") or o.get("mirrorId") or 0) != 0:
                continue
            iid = o.get("instrumentID") or o.get("instrumentId")
            info = instruments.get(iid, {})
            sym = info.get("symbol") or str(iid or "?")
            orders.append({
                "orderId":      o.get("orderID") or o.get("orderId"),
                "kind":         kind,
                "instrumentId": iid,
                "symbol":       sym,
                "name":         info.get("name") or sym,
                "type":         type_map_orders.get(info.get("typeID") or 0, "Other"),
                "isBuy":        o.get("isBuy", True),
                "rate":         o.get("rate"),
                "amount":       round(o.get("amount") or 0, 2),
                "units":        o.get("units"),
                "leverage":     o.get("leverage", 1),
                "stopLoss":     o.get("stopLossRate"),
                "takeProfit":   o.get("takeProfitRate"),
                "isTsl":        bool(o.get("isTslEnabled")),
                "openDateTime": o.get("openDateTime", ""),
            })
    orders.sort(key=lambda x: (x["symbol"], x["openDateTime"]))

    # Summary
    credit      = port.get("credits", 0) or port.get("credit", 0) or 0
    pend_open   = sum(o.get("amount", 0) or 0 for o in port.get("ordersForOpen", []) if (o.get("mirrorID") or 0) == 0)
    pend_orders = sum(o.get("amount", 0) or 0 for o in port.get("orders", []))
    cash        = credit - pend_open - pend_orders
    pos_inv     = sum(p.get("amount", 0) or 0 for p in positions_raw)
    mir_pos_inv = sum(p.get("amount", 0) or 0 for m in mirrors_raw for p in m.get("positions", []))
    mir_adj     = sum((m.get("availableAmount") or 0) - (m.get("closedPositionsNetProfit") or 0) for m in mirrors_raw)
    invested    = pos_inv + mir_pos_inv + mir_adj + pend_open + pend_orders
    pos_pnl     = sum((p.get("unrealizedPnL") or {}).get("pnL", 0) or 0 for p in positions_raw)
    mir_pnl_t   = sum((pp.get("unrealizedPnL") or {}).get("pnL", 0) or 0 for m in mirrors_raw for pp in m.get("positions", []))
    mir_closed  = sum(m.get("closedPositionsNetProfit") or 0 for m in mirrors_raw)
    total_pnl   = pos_pnl + mir_pnl_t + mir_closed
    equity      = cash + invested + total_pnl
    daily_pnl   = sum(p.get("dailyPnl") or 0 for p in result)
    summary = {
        "cash": round(cash, 2), "invested": round(invested, 2),
        "total_pnl": round(total_pnl, 2), "equity": round(equity, 2),
        "daily_pnl": round(daily_pnl, 2),
        "positions_count": len(result), "mirrors_count": len(mirrors),
        "orders_count": len(orders),
    }

    _positions_cache[account] = {"data": result, "summary": summary, "mirrors": mirrors,
                                 "orders": orders, "ts": time.time()}
    cache_write(disk_path, _positions_cache[account])
    return {"positions": result, "summary": summary, "mirrors": mirrors,
            "orders": orders, "cached": False}


@app.get("/api/etoro/trade-history")
def get_trade_history(
    account: str = Query("1"),
    minDate: str = Query(""),
    maxDate: str = Query(""),
    page: int = Query(0),
    pageSize: int = Query(100),
):
    """Uzavrete obchody z eToro trade history, obohatene o symbol/name.
    Interval je podla DATUMU UZAVRETIA (closeTimestamp) v [minDate, maxDate].

    Defenzivny prefetch: posielame eToru minDate posunuty CLOSE_LOOKBACK_DAYS
    (10 rokov) spat, lokalny close-filter potom oreze presne. Pokryva aj velmi
    stare pozicie zatvorene v zvolenom intervale. Cez stranky idem dokym eToro
    vracia plnu stranku (hard cap MAX_PAGES proti runaway loopu)."""
    if not minDate:
        minDate = (datetime.now(timezone.utc) - timedelta(days=365)).date().isoformat()
    if not maxDate:
        maxDate = datetime.now(timezone.utc).date().isoformat()
    # Posun eToro minDate o 10 rokov spat — pokryje aj velmi stare swing/long
    # pozicie (uzivatel ma napr. 5-rocne krypto obchody, takze eToro endpoint
    # vie velmi stare obchody vratit). Lokalny close-filter potom oreze presne.
    CLOSE_LOOKBACK_DAYS = 3650
    try:
        etoro_min_date = (datetime.fromisoformat(minDate)
                          - timedelta(days=CLOSE_LOOKBACK_DAYS)).date().isoformat()
    except Exception:
        etoro_min_date = (datetime.now(timezone.utc)
                          - timedelta(days=365 + CLOSE_LOOKBACK_DAYS)).date().isoformat()
    instruments = load_instruments()

    # Veľký pageSize — eToro `page` parameter je cez náš proxy nespoľahlivý
    # (vracia stále prvú stránku; možno je 1-indexovaný takže page 0 == page 1).
    # Veľký fetch znižuje počet requestov; paging je len poistka. Cap 1000 je
    # bezpečný (eToro by mal honorovať aspoň do nízkych tisícov).
    FETCH_PAGE_SIZE = 1000

    # Dedup na per-riadok kľúč (orderId; fallback positionId+close+open). Jedna
    # pozícia má pri čiastočných uzavretiach viac history riadkov so zhodným
    # positionId, preto positionId NIE je vhodný dedup kľúč.
    def _row_key(it):
        oid = it.get("orderId") or it.get("orderID") or it.get("OrderID")
        if oid is not None:
            return ("o", oid)
        return ("c",
                it.get("positionId") or it.get("positionID") or it.get("PositionID"),
                it.get("closeTimestamp") or it.get("closeDateTime") or it.get("CloseTimestamp"),
                it.get("openTimestamp") or it.get("openDateTime") or it.get("OpenTimestamp"))

    items: list = []
    seen_keys: set = set()
    MAX_PAGES = 50    # × 1000 = 50 000 obchodov, hard cap
    # Tolerujeme 2 po sebe duplicitné/prázdne stránky než ukončíme — ak je eToro
    # `page` 1-indexovaný, page 0 a page 1 vrátia tú istú prvú stránku (0 nových),
    # ale page 2 už nesie nové dáta. Break až po MAX_CONSEC_EMPTY za sebou.
    MAX_CONSEC_EMPTY = 2
    truncated = False
    cur_page = page
    last_page_size = 0
    consec_empty = 0
    for _ in range(MAX_PAGES):
        url = (
            f"{ETORO_PROXY}/etoro/trading/info/trade/history"
            f"?minDate={requests.utils.quote(etoro_min_date)}&page={cur_page}&pageSize={FETCH_PAGE_SIZE}&account={account}"
        )
        try:
            resp = fetch_with_retry(url, timeout=15, retries=2)
            raw = resp.json()
        except Exception as e:
            raise HTTPException(502, f"eToro trade history zlyhalo: {e}")
        page_items = raw if isinstance(raw, list) else raw.get("items", raw.get("trades", []))
        last_page_size = len(page_items)
        new_count = 0
        for it in page_items:
            k = _row_key(it)
            if k in seen_keys:
                continue
            seen_keys.add(k)
            items.append(it)
            new_count += 1
        if new_count == 0:
            consec_empty += 1
            if consec_empty >= MAX_CONSEC_EMPTY:
                break
        else:
            consec_empty = 0
        cur_page += 1
    else:
        truncated = True

    def pick(row, *keys, default=None):
        for key in keys:
            if key in row and row.get(key) is not None:
                return row.get(key)
        return default

    result = []
    for t in items:
        iid = pick(t, "instrumentId", "instrumentID", "InstrumentID")
        inst = instruments.get(iid, {})
        open_ts = pick(t, "openTimestamp", "openDateTime", "OpenTimestamp", "OpenDateTime")
        close_ts = pick(t, "closeTimestamp", "closeDateTime", "CloseTimestamp", "CloseDateTime")
        days_held = None
        try:
            if open_ts and close_ts:
                o = datetime.fromisoformat(open_ts.replace("Z", "+00:00"))
                c = datetime.fromisoformat(close_ts.replace("Z", "+00:00"))
                days_held = round((c - o).total_seconds() / 86400, 2)
        except Exception:
            days_held = None
        investment = pick(t, "investment", "initialInvestment", "Investment", "InitialInvestment", default=0) or 0
        net_profit = pick(t, "netProfit", "NetProfit", default=0) or 0
        result.append({
            "positionId": pick(t, "positionId", "positionID", "PositionID"),
            "orderId": pick(t, "orderId", "orderID", "OrderID"),
            "instrumentId": iid,
            "symbol": etoro_symbol_to_yf(inst.get("symbol") or str(iid)),
            "name": inst.get("name") or str(iid),
            "isBuy": pick(t, "isBuy", "IsBuy", default=True),
            "leverage": pick(t, "leverage", "Leverage"),
            "openRate": pick(t, "openRate", "OpenRate"),
            "closeRate": pick(t, "closeRate", "CloseRate"),
            "openTimestamp": open_ts,
            "closeTimestamp": close_ts,
            "daysHeld": days_held,
            "investment": investment,
            "initialInvestment": pick(t, "initialInvestment", "InitialInvestment"),
            "netProfit": net_profit,
            "profitPct": round(net_profit / investment * 100, 2) if investment else None,
            "fees": pick(t, "fees", "Fees"),
            "units": pick(t, "units", "Units"),
            "stopLossRate": pick(t, "stopLossRate", "StopLossRate"),
            "takeProfitRate": pick(t, "takeProfitRate", "TakeProfitRate"),
            "trailingStopLoss": pick(t, "trailingStopLoss", "TrailingStopLoss"),
        })

    # Filter na interval podla close dátumu (minDate poslaný aj eToru, maxDate len tu)
    def _close_in_range(x):
        ts = x.get("closeTimestamp") or ""
        if not ts:
            return True
        d = ts[:10]
        return minDate <= d <= maxDate
    result = [x for x in result if _close_in_range(x)]

    wins = [x for x in result if (x.get("netProfit") or 0) > 0]
    losses = [x for x in result if (x.get("netProfit") or 0) < 0]
    total_profit = sum(x.get("netProfit") or 0 for x in result)
    total_investment = sum(x.get("investment") or 0 for x in result)
    summary = {
        "count": len(result),
        "wins": len(wins),
        "losses": len(losses),
        "winRate": round(len(wins) / len(result) * 100, 2) if result else 0,
        "netProfit": round(total_profit, 2),
        "profitPct": round(total_profit / total_investment * 100, 2) if total_investment else None,
        "fees": round(sum(x.get("fees") or 0 for x in result), 2),
        "avgDaysHeld": round(sum(x.get("daysHeld") or 0 for x in result if x.get("daysHeld") is not None) / max(1, len([x for x in result if x.get("daysHeld") is not None])), 2) if result else 0,
    }
    return {"trades": result, "summary": summary,
            "fetchPageSize": FETCH_PAGE_SIZE, "pagesFetched": cur_page - page,
            "rawFetched": len(seen_keys), "lastPageSize": last_page_size,
            "truncated": truncated,
            "etoroMinDate": etoro_min_date,
            "minDate": minDate, "maxDate": maxDate}


# ── Nastavenia prahov (jeden zdroj pre backend aj frontend) ──────────────────
# DCA prahy konzumuje aj Investor Inbox server-side, preto žijú na disku
# a nie v localStorage — jedna hodnota pre obe strany, meniteľná bez redeployu.
DASH_SETTINGS_FILE = DATA_ROOT / "dashboard_settings.json"
_dash_settings_lock = threading.Lock()
DASH_SETTINGS_DEFAULTS = {
    "dca_loss_pct": 15.0,        # strata pozície ≥ x % → DCA úvaha
    "dca_dip_min": 90.0,         # DIP skóre ≥ x → kvalitný dip
    "dca_max_weight": 10.0,      # váha pozície < x % equity → bez koncentračného rizika
    "attention_daily_pct": 2.0,  # denný pohyb ≥ x % → dôvod Pohyb v Pozornosti
    "earnings_warn_days": 7,     # earnings ≤ x dní → ⚠ badge v scanneri
    "risk_per_trade_pct": 1.0,   # riziko na obchod ako % equity — Verdikt kalkulátor
    "atr_stop_mult": 1.5,        # stop = x × ATR14 pod vstupnou cenou
}
_DASH_SETTINGS_LIMITS = {
    "dca_loss_pct": (1, 90),
    "dca_dip_min": (0, 200),
    "dca_max_weight": (1, 100),
    "attention_daily_pct": (0.1, 50),
    "earnings_warn_days": (1, 60),
    "risk_per_trade_pct": (0.1, 10),
    "atr_stop_mult": (0.5, 5),
}


def _dash_settings() -> dict:
    out = dict(DASH_SETTINGS_DEFAULTS)
    with _dash_settings_lock:
        if DASH_SETTINGS_FILE.exists():
            try:
                saved = json.loads(DASH_SETTINGS_FILE.read_text(encoding="utf-8"))
                for k in DASH_SETTINGS_DEFAULTS:
                    if k in saved and isinstance(saved[k], (int, float)):
                        out[k] = saved[k]
            except Exception:
                pass
    return out


@app.get("/api/settings")
def get_dash_settings():
    return {"settings": _dash_settings(), "defaults": DASH_SETTINGS_DEFAULTS}


@app.post("/api/settings")
async def save_dash_settings(request: Request):
    body = await request.json()
    current = _dash_settings()
    for k, (lo, hi) in _DASH_SETTINGS_LIMITS.items():
        if k in body:
            try:
                v = float(body[k])
            except (TypeError, ValueError):
                raise HTTPException(400, f"{k}: musí byť číslo")
            if not (lo <= v <= hi):
                raise HTTPException(400, f"{k}: mimo rozsahu {lo}–{hi}")
            current[k] = int(v) if k == "earnings_warn_days" else v
    with _dash_settings_lock:
        DASH_SETTINGS_FILE.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "settings": current}


@app.get("/api/portfolio/dca")
def get_dca_candidates(
    account: str = Query("1"),
    loss_pct: float | None = Query(None, ge=1, le=90),
    dip_min: float | None = Query(None, ge=0, le=200),
    max_weight: float | None = Query(None, ge=1, le=100),
    refresh: int = Query(0),
):
    """DCA kandidáti — pozície v strate ≥ loss_pct, klasifikované podľa DIP skóre
    a váhy. Spája agregovaný per-ticker P/L (eToro) s DIP rankingom (Finviz import).

    Prahy default z /api/settings (dashboard_settings.json) — meniteľné cez ⚙
    bez redeployu; explicitné query parametre ich prebijú.

    Flagy:
      dca          → strata ≥ loss_pct, DIP ≥ dip_min, váha < max_weight (kvalitný dip)
      value_trap   → strata ≥ loss_pct, DIP < dip_min (trigger OK, slabé skóre)
      no_data      → strata ≥ loss_pct, ticker mimo DIP datasetu
      concentrated → dca podmienky OK, ALE váha ≥ max_weight (koncentračné riziko)

    Interpretačná vrstva — NEOVPLYVŇUJE C1–C4, scanner, bota ani portfolio účtovníctvo.
    DCA skóre je z manuálneho importu → vraciame aj vek dát (dip_updated_at)."""
    _s = _dash_settings()
    if loss_pct is None: loss_pct = _s["dca_loss_pct"]
    if dip_min is None: dip_min = _s["dca_dip_min"]
    if max_weight is None: max_weight = _s["dca_max_weight"]
    portfolio = get_portfolio(account=account, refresh=refresh)
    positions = portfolio.get("positions", [])
    equity = (portfolio.get("summary") or {}).get("equity") or 0

    # Agreguj per ticker (viac tranží jedného titulu = jedna pozícia)
    by_sym: dict[str, dict] = {}
    for pos in positions:
        typ = str(pos.get("type") or "").lower()
        if typ not in ("stock", "etf"):
            continue
        sym = (pos.get("symbol") or str(pos.get("instrumentId"))).upper()
        entry = by_sym.setdefault(sym, {
            "symbol": sym, "name": pos.get("name"), "amount": 0.0, "pnl": 0.0, "count": 0,
        })
        entry["amount"] += pos.get("amount") or 0
        entry["pnl"] += pos.get("pnl") or 0
        entry["count"] += 1

    dip_raw = load_dip_scores()
    dip_scores = {k.upper(): v for k, v in dip_raw.items() if not k.startswith("_")}
    dip_meta = dip_raw.get("_meta") if isinstance(dip_raw.get("_meta"), dict) else {}
    dip_updated_at = (dip_meta or {}).get("updated_at")

    candidates = []
    for sym, e in by_sym.items():
        amount = e["amount"]
        pnl_pct = (e["pnl"] / amount * 100) if amount else 0.0
        if pnl_pct > -loss_pct:
            continue  # nie je dosť v strate na DCA úvahu
        weight = (amount / equity * 100) if equity else 0.0
        dip = dip_scores.get(sym)
        dip_total = _num_or_none(dip.get("total")) if dip else None

        if dip_total is None:
            flag = "no_data"
        elif dip_total < dip_min:
            flag = "value_trap"
        elif weight >= max_weight:
            flag = "concentrated"
        else:
            flag = "dca"

        candidates.append({
            "symbol": sym,
            "name": e["name"],
            "pnl": round(e["pnl"], 2),
            "pnl_pct": round(pnl_pct, 2),
            "amount": round(amount, 2),
            "weight_pct": round(weight, 2),
            "trades": e["count"],
            "dip_total": dip_total,
            "dip_label": _dip_label(dip_total),
            "flag": flag,
        })

    # Najhoršie straty hore; v rámci toho DCA pred value_trap
    flag_order = {"dca": 0, "concentrated": 1, "value_trap": 2, "no_data": 3}
    candidates.sort(key=lambda c: (flag_order.get(c["flag"], 9), c["pnl_pct"]))

    return {
        "account": account,
        "thresholds": {"loss_pct": loss_pct, "dip_min": dip_min, "max_weight": max_weight},
        "dip_updated_at": dip_updated_at,
        "candidates": candidates,
        "counts": {
            "dca": sum(1 for c in candidates if c["flag"] == "dca"),
            "concentrated": sum(1 for c in candidates if c["flag"] == "concentrated"),
            "value_trap": sum(1 for c in candidates if c["flag"] == "value_trap"),
            "no_data": sum(1 for c in candidates if c["flag"] == "no_data"),
        },
    }


@app.get("/api/movers")
def get_movers(
    account: str = Query("1"),
    n: int = Query(6, ge=1, le=20),
    direction: str = Query("down"),
    min_change: float = Query(0.0, ge=0.0, le=50.0),
):
    """Top N titulov podľa denného % pohybu naprieč watchlistom + portfóliom.
    Len stock/ETF (crypto vynechané). Denný % z OHLCV cache (žiadne nové API
    volania — cold tituly bez cache sa preskočia, zahrejú sa prefetchom).

    direction=down → najväčšie poklesy (default), up → najväčšie rasty.
    Pre 'dynamický preset' v Grafoch — frontend otvorí grafy vrátených tickerov."""
    down = str(direction).lower() != "up"

    # Watchlist beriem celý (sú to akcie); portfólio filtrujem na stock/ETF cez type
    universe: dict[str, str] = {}   # symbol -> source
    live_rates: dict[str, float] = {}
    for item in _read_watchlist_file():
        sym = str(item.get("symbol") or "").upper()
        if sym:
            universe[sym] = "watchlist"
    try:
        portfolio = get_portfolio(account=account)
        for pos in portfolio.get("positions", []):
            typ = str(pos.get("type") or "").lower()
            if typ not in ("stock", "etf"):
                continue
            sym = str(pos.get("symbol") or "").upper()
            if not sym:
                continue
            universe[sym] = "both" if sym in universe else "portfolio"
            current_rate = pos.get("currentRate")
            if isinstance(current_rate, (int, float)) and current_rate > 0:
                live_rates[sym] = float(current_rate)
    except Exception as e:
        print(f"[movers] portfolio error: {_scrub_token(e)}")

    rows = []
    skipped = 0
    for sym, source in universe.items():
        dc = None
        price_source = "ohlcv_cache"
        live_rate = live_rates.get(sym)
        prev_close = _get_market_prev_close(sym, "Stock") if live_rate else None
        if live_rate and prev_close:
            dc = ((live_rate - prev_close) / prev_close * 100, live_rate)
            price_source = "etoro_live"
        if dc is None:
            dc = _daily_change_from_cache(sym)
        if dc is None:
            skipped += 1
            continue
        change_pct, last_close = dc
        if abs(change_pct) < min_change:
            continue
        rows.append({
            "symbol": sym,
            "change_pct": round(change_pct, 2),
            "last_close": round(last_close, 4),
            "source": source,
            "price_source": price_source,
        })

    rows.sort(key=lambda r: r["change_pct"], reverse=not down)
    top = rows[:n]

    return {
        "account": account,
        "direction": "down" if down else "up",
        "n": n,
        "min_change": min_change,
        "movers": top,
        "universe_size": len(universe),
        "evaluated": len(rows),
        "skipped": skipped,
    }


@app.get("/api/summary")
def get_summary(symbol: str = Query(...)):
    """Yahoo Finance quote — 52w High/Low, analyst recommendation, názov."""
    try:
        # Použi v8 quote endpoint — stabilnejší ako v10 quoteSummary
        url = f"https://query2.finance.yahoo.com/v8/finance/chart/{requests.utils.quote(symbol)}?range=1y&interval=1d&includePrePost=false"
        resp = fetch_with_retry(url, headers=YF_HEADERS, timeout=8, retries=2)
        data = resp.json()
        meta = data.get("chart", {}).get("result", [{}])[0].get("meta", {})

        # 52w High/Low z meta
        w52h = meta.get("fiftyTwoWeekHigh") or meta.get("regularMarketDayHigh")
        w52l = meta.get("fiftyTwoWeekLow")  or meta.get("regularMarketDayLow")
        name = meta.get("longName") or meta.get("shortName") or ""

        # Ak nemáme 52w z meta, vypočítaj z OHLCV
        if not w52h or not w52l:
            quotes = data.get("chart", {}).get("result", [{}])[0]
            highs = quotes.get("indicators", {}).get("quote", [{}])[0].get("high", [])
            lows  = quotes.get("indicators", {}).get("quote", [{}])[0].get("low",  [])
            highs = [h for h in highs if h is not None]
            lows  = [l for l in lows  if l is not None]
            if highs: w52h = max(highs)
            if lows:  w52l = min(lows)

        # Technický sentiment z OHLCV — RSI14 + trend (SMA20 vs SMA50)
        sentiment, score = None, None
        try:
            quotes = data.get("chart", {}).get("result", [{}])[0]
            closes = quotes.get("indicators", {}).get("quote", [{}])[0].get("close", [])
            closes = [c for c in closes if c is not None]
            if len(closes) >= 20:
                # RSI 14
                import pandas as pd
                s = pd.Series(closes)
                delta = s.diff()
                gain = delta.clip(lower=0).ewm(com=13, adjust=False).mean()
                loss = (-delta.clip(upper=0)).ewm(com=13, adjust=False).mean()
                rsi = 100 - (100 / (1 + gain / loss.replace(0, 1e-10)))
                rsi_val = float(rsi.iloc[-1])

                # Trend: SMA20 vs SMA50
                sma20 = float(s.tail(20).mean())
                sma50 = float(s.tail(50).mean()) if len(closes) >= 50 else sma20
                trend_up = sma20 > sma50

                # Kombinácia RSI + trend
                if rsi_val >= 60 and trend_up:     sentiment, score = "Bullish", 5
                elif rsi_val >= 50 and trend_up:    sentiment, score = "Bullish", 4
                elif rsi_val >= 45 and not trend_up: sentiment, score = "Neutral", 3
                elif rsi_val < 45 and trend_up:     sentiment, score = "Neutral", 3
                elif rsi_val < 40 and not trend_up: sentiment, score = "Bearish", 2
                else:                               sentiment, score = "Bearish", 1
        except Exception:
            pass

        return {
            "symbol":         symbol.upper(),
            "name":           name,
            "w52h":           round(w52h, 4) if w52h else None,
            "w52l":           round(w52l, 4) if w52l else None,
            "sentiment":      sentiment,
            "sentimentScore": score,
        }
    except HTTPException: raise
    except Exception as e:
        raise HTTPException(502, str(e))


FUND_ANALYSIS_DIR = DATA_ROOT / "fund_analysis"
FUND_ANALYSIS_TTL_H = 24
FUND_ANALYSIS_SCHEMA_VERSION = 1

def _fund_num(value, default=None):
    try:
        if value in (None, "", "None", "N/A", "-"):
            return default
        value = float(value)
        return value if math.isfinite(value) else default
    except Exception:
        return default

def _fund_ratio(num, den):
    if num is None or den in (None, 0):
        return None
    return num / den

def _fund_clamp(value):
    return max(0, min(100, int(round(value))))

def _fund_score_growth(pct, good=0.10, bad=-0.05):
    if pct is None:
        return 50
    if pct >= good:
        return 85
    if pct <= bad:
        return 25
    return 25 + (pct - bad) / (good - bad) * 60

def _fund_score_low(value, good, bad):
    if value is None:
        return 50
    if value <= good:
        return 85
    if value >= bad:
        return 25
    return 85 - (value - good) / (bad - good) * 60

def _fund_label(score, good="Strong", mid="Mixed", bad="Weak"):
    if score >= 70:
        return good
    if score <= 40:
        return bad
    return mid

def _fund_latest(raw, key):
    rows = raw.get(key) or []
    return rows[:4] if isinstance(rows, list) else []

def _fund_get(row, *keys):
    for key in keys:
        value = _fund_num((row or {}).get(key))
        if value is not None:
            return value
    return None

def _alpha_vantage(function: str, symbol: str) -> dict:
    key = os.getenv("ALPHA_VANTAGE_API_KEY", "").strip()
    if not key:
        raise HTTPException(status_code=503, detail="ALPHA_VANTAGE_API_KEY nie je nastaveny")
    resp = requests.get(
        "https://www.alphavantage.co/query",
        params={"function": function, "symbol": symbol, "apikey": key},
        timeout=14,
    )
    resp.raise_for_status()
    data = resp.json() or {}
    if data.get("Note") or data.get("Information"):
        raise HTTPException(status_code=429, detail=data.get("Note") or data.get("Information"))
    if data.get("Error Message"):
        raise HTTPException(status_code=404, detail=data.get("Error Message"))
    return data

def _build_fund_analysis(sym: str, raw: dict) -> dict:
    overview = raw.get("overview") or {}
    income = _fund_latest(raw.get("income") or {}, "annualReports")
    balance = _fund_latest(raw.get("balance") or {}, "annualReports")
    cashflow = _fund_latest(raw.get("cashflow") or {}, "annualReports")
    li, pi = (income[0] if income else {}), (income[1] if len(income) > 1 else {})
    lb = balance[0] if balance else {}
    lc, pc = (cashflow[0] if cashflow else {}), (cashflow[1] if len(cashflow) > 1 else {})

    revenue = _fund_get(li, "totalRevenue")
    prev_revenue = _fund_get(pi, "totalRevenue")
    net_income = _fund_get(li, "netIncome")
    op_income = _fund_get(li, "operatingIncome")
    cash = _fund_get(lb, "cashAndCashEquivalentsAtCarryingValue", "cashAndShortTermInvestments")
    debt = _fund_get(lb, "shortLongTermDebtTotal", "longTermDebt", "shortTermDebt")
    equity = _fund_get(lb, "totalShareholderEquity")
    ocf = _fund_get(lc, "operatingCashflow")
    capex = _fund_get(lc, "capitalExpenditures")
    prev_ocf = _fund_get(pc, "operatingCashflow")
    prev_capex = _fund_get(pc, "capitalExpenditures")

    fcf = None if ocf is None or capex is None else ocf - abs(capex)
    prev_fcf = None if prev_ocf is None or prev_capex is None else prev_ocf - abs(prev_capex)
    revenue_growth = _fund_ratio(revenue - prev_revenue, abs(prev_revenue)) if revenue is not None and prev_revenue else None
    fcf_growth = _fund_ratio(fcf - prev_fcf, abs(prev_fcf)) if fcf is not None and prev_fcf else None
    profit_margin = _fund_num(overview.get("ProfitMargin"))
    net_margin = _fund_ratio(net_income, revenue)
    debt_to_equity = _fund_ratio(debt, equity)
    net_debt = None if debt is None else debt - (cash or 0)
    pe = _fund_num(overview.get("PERatio"))
    fpe = _fund_num(overview.get("ForwardPE"))
    ps = _fund_num(overview.get("PriceToSalesRatioTTM"))
    ev_ebitda = _fund_num(overview.get("EVToEBITDA"))

    valuation = _fund_clamp(
        _fund_score_low(pe, 18, 45) * 0.35 +
        _fund_score_low(fpe, 16, 40) * 0.25 +
        _fund_score_low(ps, 2.5, 9) * 0.20 +
        _fund_score_low(ev_ebitda, 10, 25) * 0.20
    )
    fundamentals = _fund_clamp(
        _fund_score_growth(revenue_growth, 0.08, -0.03) * 0.30 +
        _fund_score_growth(fcf_growth, 0.10, -0.10) * 0.25 +
        (85 if fcf and fcf > 0 else 35) * 0.20 +
        (85 if (profit_margin or net_margin or 0) > 0.12 else 60 if (profit_margin or net_margin or 0) > 0.04 else 30) * 0.25
    )
    risk = _fund_clamp(
        _fund_score_low(debt_to_equity, 0.8, 2.5) * 0.55 +
        (85 if net_debt is not None and net_debt <= 0 else 55 if debt_to_equity is None or debt_to_equity < 1.5 else 30) * 0.25 +
        (80 if fcf and fcf > 0 else 35) * 0.20
    )
    analyst = 55 + (5 if _fund_num(overview.get("AnalystTargetPrice")) else 0)
    overall = _fund_clamp(fundamentals * 0.38 + valuation * 0.27 + risk * 0.25 + analyst * 0.10)
    flags = []
    if valuation < 42:
        flags.append("valuation demanding")
    if risk < 45:
        flags.append("balance-sheet / FCF risk")
    if fcf is not None and fcf < 0:
        flags.append("negative recent FCF")
    if revenue_growth is not None and revenue_growth < -0.03:
        flags.append("revenue deterioration")
    if not flags:
        flags.append("no major quantitative red flag")
    memo = (
        f"{sym}: fundamentals {_fund_label(fundamentals).lower()}, "
        f"valuation {_fund_label(valuation, 'reasonable', 'fair', 'stretched').lower()}, "
        f"risk {_fund_label(risk, 'low', 'medium', 'high').lower()}."
    )
    return {
        "schema_version": FUND_ANALYSIS_SCHEMA_VERSION,
        "symbol": sym,
        "company": overview.get("Name") or sym,
        "source": "Alpha Vantage",
        "scores": {"overall": overall, "fundamentals": fundamentals, "valuation": valuation, "risk": risk, "analyst": analyst},
        "labels": {
            "overall": _fund_label(overall, "Good", "Mixed", "Weak"),
            "fundamentals": _fund_label(fundamentals),
            "valuation": _fund_label(valuation, "Reasonable", "Fair", "Stretched"),
            "risk": _fund_label(risk, "Low", "Medium", "High"),
        },
        "flags": flags[:3],
        "memo": memo,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
    }

@app.get("/api/ticker/fund-analysis/{symbol}")
def get_ticker_fund_analysis(symbol: str, refresh: int = Query(0)):
    sym = _validate_ticker_symbol(symbol)
    path = FUND_ANALYSIS_DIR / f"{sym}.json"
    if not refresh:
        cached = _read_ticker_json_cache(path, FUND_ANALYSIS_SCHEMA_VERSION, FUND_ANALYSIS_TTL_H)
        if cached is not None:
            cached["cached"] = True
            return cached
    raw = {
        "overview": _alpha_vantage("OVERVIEW", sym),
        "income": _alpha_vantage("INCOME_STATEMENT", sym),
        "balance": _alpha_vantage("BALANCE_SHEET", sym),
        "cashflow": _alpha_vantage("CASH_FLOW", sym),
    }
    payload = _build_fund_analysis(sym, raw)
    payload["cached"] = False
    try:
        _write_ticker_json_cache(path, payload)
    except Exception:
        pass
    return payload

# ── BACKGROUND PREFETCH ──────────────────────────────────────────────────────
_prefetch_running = False
_prefetch_log: list = []

# Všetky intervaly ktoré dashboard používa
PREFETCH_INTERVALS = ["OneDay", "OneWeek", "OneHour", "FourHours"]

def _get_portfolio_holdings() -> dict:
    """Agreguje pozície oboch účtov na {YF_SYMBOL: {pnl, amount, pnl_pct}}.
    Číta z RAM cache get_portfolio (fallback disk processed_{acct}) — žiadne
    extra eToro volania. Predtým čítala legacy raw-payload disk cache
    (portfolio_{acct}), ktorú od zavedenia processed cache (de2fe81) už nikto
    nezapisoval — Scanner PORT badge/DCA/Inbox/Verdikt tak stáli na navždy
    zamrznutých dátach. get_portfolio.result už je stocks/ETF aj ostatné typy,
    preto sa tu type filter aplikuje explicitne (rovnaký zámer ako predtým)."""
    holdings: dict = {}
    for acct in ["1", "2"]:
        try:
            cached = _positions_cache.get(acct) or cache_read(_portfolio_disk_path(acct))
            if not cached:
                continue
            for pos in cached.get("data", []):
                if pos.get("type") not in ("Stock", "ETF"):
                    continue
                sym = pos.get("symbol")
                if not sym:
                    continue
                pnl = pos.get("pnl", 0) or 0
                amt = pos.get("amount", 0) or 0
                h = holdings.setdefault(sym, {"pnl": 0.0, "amount": 0.0})
                h["pnl"] += pnl
                h["amount"] += amt
        except Exception as e:
            print(f"  [holdings] account {acct} error: {e}")
    for sym, h in holdings.items():
        h["pnl"] = round(h["pnl"], 2)
        h["amount"] = round(h["amount"], 2)
        h["pnl_pct"] = round(h["pnl"] / h["amount"] * 100, 2) if h["amount"] else None
    return holdings


@app.get("/api/portfolio/holdings")
def get_portfolio_holdings():
    """Ľahký snapshot držaných titulov a čakajúcich objednávok z cache."""
    order_symbols = set()
    for acct in ("1", "2"):
        try:
            cached = _positions_cache.get(acct) or cache_read(_portfolio_disk_path(acct))
            for order in (cached or {}).get("orders", []):
                symbol = str(order.get("symbol") or "").strip().upper()
                if symbol:
                    order_symbols.add(symbol)
        except Exception:
            pass
    return {"holdings": _get_portfolio_holdings(), "order_symbols": sorted(order_symbols)}


# ── Korelačná mapa portfólia ──────────────────────────────────────────────────
# Interpretačná vrstva: odhalí skrytú koncentráciu (tituly čo sa hýbu spolu).
# Čisto z OHLCV disk cache — žiadne nové API volania, žiadny vplyv na účtovanie.
_CORR_CACHE = {"key": None, "data": None, "ts": 0.0}
CORR_CACHE_TTL = 900          # 15 min RAM cache
CORR_MAX_SYMBOLS = 30         # najväčšie pozície podľa amount (30x30 matica stačí)
CORR_MIN_OVERLAP = 30         # min. spoločných dní pre validnú koreláciu


@app.get("/api/portfolio/correlation")
def get_portfolio_correlation(days: int = Query(90, ge=30, le=365)):
    """Pearsonova korelácia denných výnosov držaných titulov (oba účty).
    Tituly bez dostatočnej OHLCV cache sa preskočia (fail-soft, `skipped`)."""
    holdings = _get_portfolio_holdings()
    symbols = sorted(holdings.keys(), key=lambda s: -abs(float(holdings[s].get("amount") or 0)))
    symbols = symbols[:CORR_MAX_SYMBOLS]
    cache_key = (tuple(sorted(symbols)), days)
    now = _time_module.time()
    if _CORR_CACHE["key"] == cache_key and now - _CORR_CACHE["ts"] < CORR_CACHE_TTL:
        return _CORR_CACHE["data"]

    series: dict = {}
    skipped: list = []
    for sym in symbols:
        closes: dict = {}
        try:
            raw = cache_read(CACHE_DIR / "ohlcv" / f"{sym}_OneDay")
            candles = _flatten_ohlcv(raw) if raw else []
        except Exception:
            candles = []
        for c in candles[-(days + 5):]:
            d = str(c.get("fromDate", ""))[:10]
            try:
                px = float(c.get("close") or 0)
            except (TypeError, ValueError):
                px = 0.0
            if d and px > 0:
                closes[d] = px
        if len(closes) > CORR_MIN_OVERLAP:
            series[sym] = closes
        else:
            skipped.append(sym)

    syms = sorted(series.keys())
    if len(syms) < 2:
        return {"symbols": syms, "matrix": [], "skipped": skipped, "days": days,
                "pairs_high": [], "note": "V OHLCV cache nie je dosť dát (pomôže prefetch/otvorenie grafov)."}

    df = pd.DataFrame(series).sort_index()
    rets = df.pct_change()
    corr = rets.corr(min_periods=CORR_MIN_OVERLAP)

    matrix = []
    for a in syms:
        row = []
        for b in syms:
            try:
                v = corr.at[a, b]
                row.append(round(float(v), 3) if pd.notna(v) else None)
            except Exception:
                row.append(None)
        matrix.append(row)

    pairs_high = []
    for i in range(len(syms)):
        for j in range(i + 1, len(syms)):
            v = matrix[i][j]
            if v is not None and v >= 0.8:
                pairs_high.append({"a": syms[i], "b": syms[j], "corr": v})
    pairs_high.sort(key=lambda p: -p["corr"])

    data = {
        "symbols": syms,
        "matrix": matrix,
        "skipped": skipped,
        "days": days,
        "pairs_high": pairs_high[:12],
        "computed_at": datetime.now().isoformat(timespec="seconds"),
    }
    _CORR_CACHE.update(key=cache_key, data=data, ts=now)
    return data


def _get_portfolio_symbols() -> set:
    """Načíta symboly zo všetkých portfólií (oba účty)."""
    syms = set()
    for acct in ["1", "2"]:
        try:
            port_cache_path = CACHE_DIR / "portfolio" / f"portfolio_{acct}"
            cached = cache_read(port_cache_path)
            if cached:
                # BUG (fixed 2026-07-09): cache_read vracia surový /pnl/real
                # payload — pozície sú pod clientPortfolio.positions, nie na
                # top-level "positions". Predtým to vždy vrátilo [] a prefetch
                # tak ticho vynechával všetky portfolio symboly.
                for pos in cached.get("clientPortfolio", {}).get("positions", []):
                    iid = pos.get("instrumentID") or pos.get("instrumentId")
                    inst = load_instruments().get(iid, {})
                    sym = inst.get("symbol")
                    if sym:
                        syms.add(sym)
            else:
                # Načítaj live
                resp = fetch_with_retry(f"{ETORO_PROXY}/pnl/real?account={acct}", timeout=15, retries=2)
                data = resp.json()
                cache_write(port_cache_path, data)
                port = data.get("clientPortfolio", {})
                for pos in port.get("positions", []):
                    iid = pos.get("instrumentID") or pos.get("instrumentId")
                    inst = load_instruments().get(iid, {})
                    sym = inst.get("symbol")
                    if sym:
                        syms.add(sym)
        except Exception as e:
            print(f"  [prefetch] Portfolio {acct} error: {e}")
    return syms


def _prefetch_ohlcv_symbol(sym: str, iid: int, intervals: list):
    """Prefetchne OHLCV pre jeden symbol — všetky intervaly."""
    for interval in intervals:
        cache_path = CACHE_DIR / "ohlcv" / f"{sym}_{interval}"
        age = cache_age_seconds(cache_path)

        # TTL per interval
        ttl = {"OneDay": 86400, "OneWeek": 604800,
               "OneHour": 3600, "FourHours": 14400}.get(interval, 3600)

        if age < ttl:
            continue  # Cache je čerstvá

        cached = cache_read(cache_path)
        if cached:
            # Inkrementálny update — zisti koľko nových sviečok
            cached_candles = []
            for g in cached.get("candles", []):
                cached_candles.extend(g.get("candles", []))
            if cached_candles:
                last_date = _last_candle_key(cached_candles, interval)
                today = _latest_expected_candle_date(interval)
                n_new = _days_missing(last_date, today, interval)
                if n_new <= 0 and interval not in ("OneDay", "OneWeek") and _intraday_refresh_window():
                    n_new = 10
                if n_new <= 0:
                    # Nič nové — len obnov timestamp
                    cache_write(cache_path, cached)
                    continue
                fetch_count = min(n_new + 3, 50)
                try:
                    url = f"{ETORO_PROXY}/etoro/market-data/instruments/{iid}/history/candles/asc/{interval}/{fetch_count}?account=1"
                    resp = fetch_with_retry(url, timeout=15, retries=2)
                    new_raw = resp.json()
                    merged = _merge_ohlcv(cached, new_raw, last_date, interval)
                    cache_write(cache_path, merged)
                    n = sum(len(g.get("candles",[])) for g in new_raw.get("candles",[]))
                    print(f"  [prefetch] {sym} {interval}: +{n} nových sviečok")
                except Exception as e:
                    print(f"  [prefetch] {sym} {interval} update error: {e}")
        else:
            # Prvé načítanie — fetch 1000
            try:
                url = f"{ETORO_PROXY}/etoro/market-data/instruments/{iid}/history/candles/asc/{interval}/1000?account=1"
                resp = fetch_with_retry(url, timeout=20, retries=2)
                raw = resp.json()
                cache_write(cache_path, raw)
                total = sum(len(g.get("candles",[])) for g in raw.get("candles",[]))
                print(f"  [prefetch] {sym} {interval}: {total} sviečok (full)")
            except Exception as e:
                print(f"  [prefetch] {sym} {interval} fetch error: {e}")


def _prefetch_worker(symbols: list, account: str):
    """Bežiaci v pozadí — prefetchne všetky dáta dashboardu."""
    global _prefetch_running
    _prefetch_log.clear()

    def log(msg):
        print(f"  [prefetch] {msg}")
        _prefetch_log.append(msg)

    try:
        log("Štart...")

        # 1. Instruments
        inst = load_instruments()
        log(f"Instruments: {len(inst)}")

        # 2. Portfolio pre oba účty + zozbieraj symboly z portfólia
        port_symbols = _get_portfolio_symbols()
        log(f"Portfolio symboly: {len(port_symbols)}")

        # 3. Zlúč všetky symboly: watchlist + portfólio
        all_symbols = set(s.upper() for s in symbols) | port_symbols
        log(f"Celkom symbolov: {len(all_symbols)}")

        # 4. OHLCV pre každý symbol
        instruments = load_instruments()
        # Vytvor reverznú mapu symbol -> instrumentId
        sym_to_iid = {v["symbol"]: k for k, v in instruments.items() if v.get("symbol")}
        # Doplň z _instrument_id_cache
        for sym, iid in _instrument_id_cache.items():
            if sym not in sym_to_iid:
                sym_to_iid[sym] = iid

        ok, skip, err = 0, 0, 0
        for sym in sorted(all_symbols):
            iid = sym_to_iid.get(sym) or get_instrument_id(sym, "1")
            if not iid:
                log(f"{sym}: nenájdený, skip")
                err += 1
                continue
            _prefetch_ohlcv_symbol(sym, iid, PREFETCH_INTERVALS)
            ok += 1

        log(f"OHLCV: {ok} OK, {err} chýb")
        log("Hotovo ✓")
    except Exception as e:
        log(f"Chyba: {e}")
    finally:
        _prefetch_running = False


@app.post("/api/prefetch")
async def start_prefetch(request: Request):
    """Spustí background prefetch OHLCV + portfolio + instruments."""
    global _prefetch_running
    if _prefetch_running:
        return {"status": "already_running", "log": _prefetch_log}
    body = await request.json()
    symbols = body.get("symbols", [])
    account = body.get("account", "1")
    if not symbols:
        return {"status": "no_symbols"}
    _prefetch_running = True
    t = threading.Thread(target=_prefetch_worker, args=(symbols, account), daemon=True)
    t.start()
    return {"status": "started", "symbols": len(symbols)}


@app.get("/api/prefetch/status")
def prefetch_status():
    """Vráti stav prefetch."""
    return {"running": _prefetch_running, "log": _prefetch_log[-20:]}


@app.get("/api/etoro/status")
def etoro_status():
    """Rýchla kontrola či eToro proxy beží."""
    try:
        resp = requests.get(f"{ETORO_PROXY}/pnl/real", timeout=3)
        return {"ok": resp.ok, "status": resp.status_code}
    except Exception as e:
        return {"ok": False, "error": str(e)}

# ── ETORO INSTRUMENT ID CACHE ───────────────────────────────────────────────
# symbol -> instrumentId (immutable, cachujeme natrvalo)
_instrument_id_cache: dict = {}

def get_instrument_id(symbol: str, account: str = "1") -> int | None:
    """Zistí instrumentId pre symbol cez eToro search. Cachuje natrvalo."""
    sym = symbol.upper().strip()
    if sym in _instrument_id_cache:
        return _instrument_id_cache[sym]
    try:
        # Skús presný match cez internalSymbolFull parameter
        url = f"{ETORO_PROXY}/etoro/market-data/search?internalSymbolFull={sym}&fields=instrumentId,symbolName,internalSymbolFull&pageSize=20&account={account}"
        resp = requests.get(url, timeout=6)
        items = resp.json().get("items", []) if resp.ok else []

        # Ak nenájdeme cez internalSymbolFull, skúsime searchText
        if not items:
            url2 = f"{ETORO_PROXY}/etoro/market-data/search?searchText={sym}&fields=instrumentId,symbolName,internalSymbolFull&pageSize=20&account={account}"
            resp2 = requests.get(url2, timeout=6)
            items = resp2.json().get("items", []) if resp2.ok else []

        if items:
            # Hľadaj presný match
            match = next((i for i in items
                if i.get("internalSymbolFull","").upper() == sym
                or i.get("symbolName","").upper() == sym), None)
            # Ak nenájdeme presný, vezmi prvý
            if not match and items:
                match = items[0]
            if match and match.get("instrumentId"):
                _instrument_id_cache[sym] = match["instrumentId"]
                print(f"  Resolved {sym} -> instrumentId {match['instrumentId']} (symbolName: {match.get('symbolName')})")
                return match["instrumentId"]
    except Exception as e:
        print(f"  get_instrument_id({sym}) error: {e}")
    return None

# ── ETORO OHLCV ──────────────────────────────────────────────────────────────

def _days_missing(last_date_str: str, today_str: str, interval: str) -> int:
    """Odhadni počet chýbajúcich sviečok medzi poslednou cached a dneškom."""
    try:
        from datetime import datetime, timezone
        last = datetime.strptime(last_date_str[:10], "%Y-%m-%d")
        today = datetime.strptime(today_str[:10], "%Y-%m-%d")
        delta = (today - last).days
        if delta <= 0:
            return 0
        # Odfiltruj víkendy pre denné/hodinové intervaly
        if interval in ("OneDay",):
            # ~5/7 obchodných dní
            return max(1, int(delta * 5 / 7) + 1)
        elif interval in ("OneHour", "FourHours", "ThirtyMinutes"):
            return max(1, int(delta * 8))   # ~8 sviečok/deň pre hodinové
        elif interval in ("OneWeek",):
            return max(1, delta // 7 + 1)
        else:
            return max(1, delta * 2)
    except Exception:
        return 10  # fallback — fetch posledných 10


def _latest_expected_candle_date(interval: str, now: datetime | None = None) -> str:
    """Najnovsi den, pre ktory ma zmysel cakat uzavretu sviecku."""
    now = now or datetime.now(timezone.utc)
    d = now.date()

    if interval == "OneWeek":
        monday = d - timedelta(days=d.weekday())
        if d.weekday() == 0 and now.hour < 22:
            monday -= timedelta(days=7)
        return monday.isoformat()

    if interval == "OneDay":
        if d.weekday() >= 5:
            d -= timedelta(days=d.weekday() - 4)
        elif now.hour < 22:
            d -= timedelta(days=3 if d.weekday() == 0 else 1)
        return d.isoformat()

    return now.strftime("%Y-%m-%d")


def _intraday_refresh_window(now: datetime | None = None) -> bool:
    """US market/pravidelne intraday okno v UTC, s malou rezervou."""
    now = now or datetime.now(timezone.utc)
    return now.weekday() < 5 and 13 <= now.hour <= 22


def _candle_key(candle: dict, interval: str) -> str:
    from_date = candle.get("fromDate", "")
    return from_date[:10] if interval in ("OneDay", "OneWeek") else from_date


def _last_candle_key(candles: list, interval: str) -> str:
    if not candles:
        return ""
    return _candle_key(candles[-1], interval)


def _merge_ohlcv(cached: dict, new_data: dict, last_cached_date: str, interval: str = "OneDay") -> dict:
    """Mergni nové sviečky do cached — odstráň duplikáty podľa fromDate."""
    import copy
    result = copy.deepcopy(cached)

    # Zbier existujúce dátumy
    existing_dates = set()
    for group in result.get("candles", []):
        for c in group.get("candles", []):
            d = _candle_key(c, interval)
            if d:
                existing_dates.add(d)

    # Pridaj nové sviečky ktoré ešte nemáme
    new_added = 0
    for group in new_data.get("candles", []):
        new_candles = [
            c for c in group.get("candles", [])
            if _candle_key(c, interval) not in existing_dates
               and _candle_key(c, interval) > last_cached_date
        ]
        if new_candles:
            if result.get("candles"):
                result["candles"][-1]["candles"].extend(new_candles)
            else:
                result["candles"] = [{"candles": new_candles}]
            new_added += len(new_candles)

    return result


def _flatten_ohlcv(raw: dict) -> list:
    candles = []
    for group in raw.get("candles", []):
        candles.extend(group.get("candles", []))
    return candles


def _get_prev_close(sym: str) -> float | None:
    """Vráti close predchádzajúcej dennej sviečky z OHLCV cache (nie dnešnej)."""
    try:
        cache_path = CACHE_DIR / "ohlcv" / f"{sym}_OneDay"
        raw = cache_read(cache_path)
        if not raw:
            return None
        candles = _flatten_ohlcv(raw)
        if len(candles) < 2:
            return None
        # Posledná sviečka môže byť dnešná (neuzavretá) — vezmi predposlednú
        today_str = datetime.now(timezone.utc).date().isoformat()
        closed = [c for c in candles if (c.get("fromDate") or "")[:10] < today_str]
        if not closed:
            return None
        return float(closed[-1].get("close") or closed[-1].get("c") or 0) or None
    except Exception:
        return None


MARKET_PREV_CLOSE_TTL = 6 * 3600


def _market_prev_close_cache_key(sym: str) -> _Path:
    safe = re.sub(r"[^A-Z0-9_.-]+", "_", str(sym or "").upper())
    return CACHE_DIR / "market_close" / safe


def _market_close_cutoff_allows_today(now_local: datetime | None = None) -> bool:
    """Po lokalnom eToro cut-offe moze byt baseline dnesna market close."""
    now_local = now_local or datetime.now(ZoneInfo("Europe/Bratislava"))
    return now_local.hour >= 22


def _select_market_prev_close(df: pd.DataFrame) -> tuple[float | None, str | None]:
    if df is None or len(df) < 2 or "Close" not in df.columns:
        return None, None
    work = df.dropna(subset=["Close"]).copy()
    if work.empty:
        return None, None
    now_local = datetime.now(ZoneInfo("Europe/Bratislava"))
    local_today = now_local.date()
    allow_today = _market_close_cutoff_allows_today(now_local)
    selected = []
    for idx, row in work.iterrows():
        try:
            day = pd.Timestamp(idx).date()
        except Exception:
            continue
        if day < local_today or (allow_today and day == local_today):
            selected.append((day, row))
    if not selected:
        return None, None
    day, row = selected[-1]
    close = _num_or_none(row.get("Close"))
    if close is None or close <= 0:
        return None, None
    return float(close), day.isoformat()


def _get_market_prev_close(sym: str, asset_type: str | None = None) -> float | None:
    """Market previous close pre portfolio daily P/L.

    Stock/ETF pouziva daily data z Massive/yfinance. Ostatne typy ostavaju na
    povodnom eToro OHLCV cache fallbacku.
    """
    sym = str(sym or "").upper().strip()
    if not sym:
        return None
    typ = str(asset_type or "").lower()
    if typ not in ("stock", "etf"):
        return _get_prev_close(sym)

    cache_path = _market_prev_close_cache_key(sym)
    try:
        cached = cache_read(cache_path)
        if cached and cache_age_seconds(cache_path) < MARKET_PREV_CLOSE_TTL:
            close = _num_or_none(cached.get("close"))
            if close and close > 0:
                return float(close)
    except Exception:
        pass

    try:
        raw = _yf_download_cached(sym, "1mo", "1d", prefer_massive=True)
        close, close_date = _select_market_prev_close(raw)
        if close and close > 0:
            (CACHE_DIR / "market_close").mkdir(exist_ok=True)
            cache_write(cache_path, {
                "symbol": sym,
                "close": round(float(close), 8),
                "date": close_date,
                "source": "massive_or_yfinance",
                "updated_at": datetime.now(timezone.utc).isoformat(),
            })
            return float(close)
    except Exception as exc:
        print(f"[market_close] {sym}: {type(exc).__name__}")

    return _get_prev_close(sym)


def _daily_change_from_cache(sym: str) -> tuple[float, float] | None:
    """Posledný denný % pohyb z OHLCV cache (cache-only, žiadne API volanie).
    Vracia (change_pct, last_close) alebo None keď nie je dosť dát."""
    try:
        raw = cache_read(CACHE_DIR / "ohlcv" / f"{sym}_OneDay")
        if not raw:
            return None
        candles = _flatten_ohlcv(raw)
        if len(candles) < 2:
            return None
        last = float(candles[-1].get("close") or candles[-1].get("c") or 0)
        prev = float(candles[-2].get("close") or candles[-2].get("c") or 0)
        if not prev or not last:
            return None
        return (last - prev) / prev * 100, last
    except Exception:
        return None


def _merge_ohlcv_tail(cached: dict, new_data: dict, interval: str, max_candles: int = 1000) -> dict:
    """Mergni malý live tail tak, aby prekrývajúce sa sviečky prepísali cache."""
    import copy
    result = copy.deepcopy(cached)
    by_key = {}

    for candle in _flatten_ohlcv(cached):
        key = _candle_key(candle, interval)
        if key:
            by_key[key] = candle

    for candle in _flatten_ohlcv(new_data):
        key = _candle_key(candle, interval)
        if key:
            by_key[key] = candle

    merged = sorted(by_key.values(), key=lambda c: c.get("fromDate", ""))
    if max_candles and len(merged) > max_candles:
        merged = merged[-max_candles:]

    result["candles"] = [{"candles": merged}]
    return result


def _tail_adds_new_candle(cached: dict, new_data: dict, interval: str) -> bool:
    cached_keys = {_candle_key(c, interval) for c in _flatten_ohlcv(cached)}
    for candle in _flatten_ohlcv(new_data):
        key = _candle_key(candle, interval)
        if key and key not in cached_keys:
            return True
    return False


def _tail_refresh_count(interval: str) -> int:
    if interval in ("OneMinute", "FiveMinutes", "ThirtyMinutes"):
        return 3
    if interval in ("OneHour", "FourHours"):
        return 3
    return 3


def _tail_refresh_ttl(interval: str) -> int:
    if interval == "OneMinute":
        return 30
    if interval in ("FiveMinutes", "ThirtyMinutes"):
        return 90
    if interval in ("OneHour", "FourHours"):
        return 180
    return 300


def _full_fetch_count(interval: str) -> int:
    if interval == "OneMinute":
        return 390
    if interval in ("FiveMinutes", "ThirtyMinutes"):
        return 500
    if interval in ("OneHour", "FourHours"):
        return 750
    return 1000


# ── CANDLESTICK PATTERN DETECTION ────────────────────────────────────────────
PATTERN_META = {
    "hammer":         {"name":"Hammer",             "dir":"bull","rel":"Low"},
    "inv_hammer":     {"name":"Inverted Hammer",    "dir":"bull","rel":"High"},
    "shooting_star":  {"name":"Shooting Star",      "dir":"bear","rel":"High"},
    "hanging_man":    {"name":"Hanging Man",        "dir":"bear","rel":"Low"},
    "doji":           {"name":"Doji",               "dir":"neut","rel":"High"},
    "dragonfly_doji": {"name":"Dragonfly Doji",    "dir":"bull","rel":"Low"},
    "gravestone_doji":{"name":"Gravestone Doji",   "dir":"bear","rel":"High"},
    "bullish_engulf": {"name":"Bullish Engulfing", "dir":"bull","rel":"Low"},
    "bearish_engulf": {"name":"Bearish Engulfing", "dir":"bear","rel":"High"},
    "morning_star":   {"name":"Morning Star",       "dir":"bull","rel":"Low"},
    "evening_star":   {"name":"Evening Star",       "dir":"bear","rel":"High"},
    "bullish_harami": {"name":"Bullish Harami",    "dir":"bull","rel":"Low"},
    "bearish_harami": {"name":"Bearish Harami",    "dir":"bear","rel":"High"},
    "piercing":       {"name":"Piercing Line",      "dir":"bull","rel":"Low"},
    "dark_cloud":     {"name":"Dark Cloud Cover",  "dir":"bear","rel":"High"},
}

def detect_patterns(df: pd.DataFrame, is_intraday_flag: bool = False) -> list:
    if is_intraday_flag or len(df) < 3:
        return []
    O = df["Open"].values.astype(float)
    H = df["High"].values.astype(float)
    L = df["Low"].values.astype(float)
    C = df["Close"].values.astype(float)

    body      = np.abs(C - O)
    total     = np.where(H - L > 0, H - L, 1e-8)
    upper_wick = H - np.maximum(O, C)
    lower_wick = np.minimum(O, C) - L
    bull = C > O
    bear = C < O
    avg_body = pd.Series(body).rolling(14, min_periods=3).mean().fillna(body.mean()).values

    results = []
    def add(i, key):
        meta = PATTERN_META[key]
        t = df.index[i]
        try:
            ts = int(pd.Timestamp(t).timestamp()) if pd.Timestamp(t).hour != 0 else str(t)[:10]
        except:
            ts = str(t)[:10]
        results.append({"time": ts, "key": key, "name": meta["name"],
                         "dir": meta["dir"], "rel": meta["rel"]})

    n = len(df)
    for i in range(2, n):
        b = body[i]; ab = avg_body[i] or 1e-8; t = total[i]
        uw = upper_wick[i]; lw = lower_wick[i]
        sm = b < ab * 0.35; lg = b > ab * 0.6
        trend_up = C[max(0,i-5):i].mean() < C[i] if i >= 5 else False

        # Doji
        if b / t < 0.08:
            if lw > t*0.6 and uw < t*0.15: add(i,"dragonfly_doji")
            elif uw > t*0.6 and lw < t*0.15: add(i,"gravestone_doji")
            else: add(i,"doji")
        # Hammer / Hanging Man
        elif lw >= b*2 and uw <= b*0.5 and b > 0:
            add(i, "hanging_man" if trend_up else "hammer")
        # Inv Hammer / Shooting Star
        elif uw >= b*2 and lw <= b*0.5 and b > 0:
            add(i, "shooting_star" if trend_up else "inv_hammer")

        if i >= 1:
            pb = body[i-1]
            # Engulfing
            if pb > 0 and lg:
                if bear[i-1] and bull[i] and O[i]<=C[i-1] and C[i]>=O[i-1]:
                    add(i,"bullish_engulf")
                elif bull[i-1] and bear[i] and O[i]>=C[i-1] and C[i]<=O[i-1]:
                    add(i,"bearish_engulf")
            # Harami
            if pb > 0 and sm:
                if bear[i-1] and bull[i] and O[i]>C[i-1] and C[i]<O[i-1]:
                    add(i,"bullish_harami")
                elif bull[i-1] and bear[i] and O[i]<C[i-1] and C[i]>O[i-1]:
                    add(i,"bearish_harami")
            # Piercing / Dark Cloud
            if pb > ab*0.6:
                mid = (O[i-1]+C[i-1])/2
                if bear[i-1] and bull[i] and O[i]<C[i-1] and mid<C[i]<O[i-1]:
                    add(i,"piercing")
                if bull[i-1] and bear[i] and O[i]>C[i-1] and O[i-1]<C[i]<mid:
                    add(i,"dark_cloud")

        if i >= 2:
            p2b = body[i-2]
            if p2b > ab*0.7:
                if bear[i-2] and sm and bull[i] and body[i]>ab*0.5 and C[i]>(O[i-2]+C[i-2])/2:
                    add(i,"morning_star")
                if bull[i-2] and sm and bear[i] and body[i]>ab*0.5 and C[i]<(O[i-2]+C[i-2])/2:
                    add(i,"evening_star")

    return results


@app.get("/api/ohlcv")
def get_ohlcv(
    symbol:     str  = Query(...),
    period:     str  = Query("1y"),
    interval:   str  = Query("1d"),
    indicators: str  = Query(""),
    ha:         int  = Query(0),
    account:    str  = Query("1"),
    refresh:    int  = Query(1),
    limit:      int  = Query(0, ge=0, le=1000),
    before:     str  = Query(""),
):
    """Načíta OHLCV z eToro API. Indikátory počítame lokálne."""
    sym = symbol.upper().strip()

    # Mapuj interval
    etoro_interval = YAHOO_TO_ETORO_INTERVAL.get(interval, "OneDay")

    # 4h a 12h resampling
    use_resample = interval in ("15m", "4h", "12h")
    fetch_interval = "FiveMinutes" if interval == "15m" else ("OneHour" if use_resample else etoro_interval)

    # Počet sviečok
    candles_count = AUTO_INTERVAL_TO_COUNT.get(interval, 252) if period == "auto" else YAHOO_PERIOD_TO_COUNT.get((period, interval), 252)
    if use_resample:
        mult = {"15m": 3, "4h": 4, "12h": 12}.get(interval, 1)
        candles_count = min(candles_count * mult, 1000)

    # ── INKREMENTÁLNY OHLCV CACHE ────────────────────────────────────────────
    # Kľúč bez candles_count — cache je per symbol+interval, nie per count
    ohlcv_cache_key  = f"{sym}_{fetch_interval}"
    ohlcv_cache_path = CACHE_DIR / "ohlcv" / ohlcv_cache_key.replace("/", "_")

    cached_raw = cache_read(ohlcv_cache_path)
    raw = None
    iid = _instrument_id_cache.get(sym)

    if cached_raw:
        # Zisti počet a dátum sviečok v cache
        cached_candles = _flatten_ohlcv(cached_raw)

        if cached_candles:
            cached_count = len(cached_candles)
            if not refresh:
                raw = cached_raw
                print(f"  OHLCV cache-only: {sym} {fetch_interval} ({cached_count} cached)")
            else:
                iid = iid or get_instrument_id(sym, account)
                if iid is None:
                    raise HTTPException(404, f"Instrument '{sym}' nenájdený na eToro")
                last_cached_date = _last_candle_key(cached_candles, fetch_interval)
                today = _latest_expected_candle_date(fetch_interval)
                cache_age = cache_age_seconds(ohlcv_cache_path)
                if cache_age < _tail_refresh_ttl(fetch_interval):
                    raw = cached_raw
                    print(f"  OHLCV fresh cache: {sym} {fetch_interval} ({cached_count} cached, age={int(cache_age)}s)")
                else:
                    fetch_new_count = _days_missing(last_cached_date, today, fetch_interval)
                    tail_count = max(_tail_refresh_count(fetch_interval), min(fetch_new_count + 3, 50) if fetch_new_count > 0 else 0)
                    url_tail = f"{ETORO_PROXY}/etoro/market-data/instruments/{iid}/history/candles/asc/{fetch_interval}/{tail_count}?account={account}"
                    try:
                        resp_tail = fetch_with_retry(url_tail, timeout=6, retries=1)
                        tail_raw = resp_tail.json()
                        raw = _merge_ohlcv_tail(cached_raw, tail_raw, fetch_interval)
                        if _tail_adds_new_candle(cached_raw, tail_raw, fetch_interval):
                            cache_write(ohlcv_cache_path, raw)
                        tail_n = len(_flatten_ohlcv(tail_raw))
                        print(f"  OHLCV cache+tail: {sym} {fetch_interval} ({cached_count} cached, tail {tail_n}, age={int(cache_age)}s)")
                    except Exception as e:
                        raw = cached_raw
                        print(f"  OHLCV tail fallback: {sym} {fetch_interval}: {e}")

    if raw is None:
        iid = iid or get_instrument_id(sym, account)
        if iid is None:
            raise HTTPException(404, f"Instrument '{sym}' nenájdený na eToro")
        # Prvé načítanie — fetchni vždy maximum (1000) pre daný interval
        # Takto máme plnú históriu lokálne pre akúkoľvek periódu
        FULL_FETCH_COUNT = _full_fetch_count(fetch_interval)
        url = f"{ETORO_PROXY}/etoro/market-data/instruments/{iid}/history/candles/asc/{fetch_interval}/{FULL_FETCH_COUNT}?account={account}"
        try:
            resp = fetch_with_retry(url, timeout=20, retries=2)
            raw  = resp.json()
            cache_write(ohlcv_cache_path, raw)
            total = sum(len(g.get("candles",[])) for g in raw.get("candles",[]))
            print(f"  OHLCV full fetch: {sym} {fetch_interval} ({total} sviečok, max=1000)")
        except Exception as e:
            stale = cache_read(ohlcv_cache_path)
            if stale:
                print(f"  OHLCV stale fallback: {sym}")
                raw = stale
            else:
                raise HTTPException(502, f"eToro candles zlyhalo: {e}")

    # Spracuj response — z cache môžeme mať 1000 sviečok, vráť len potrebný počet
    candle_list = []
    for group in raw.get("candles", []):
        for c in group.get("candles", []):
            candle_list.append({
                "Open":   c.get("open"),
                "High":   c.get("high"),
                "Low":    c.get("low"),
                "Close":  c.get("close"),
                "Volume": c.get("volume", 0),
                "Date":   c.get("fromDate", ""),
            })

    if not candle_list:
        raise HTTPException(404, f"Žiadne sviečky pre '{sym}'")

    # Orezaj na požadovaný počet sviečok (najnovšie)
    if not use_resample and len(candle_list) > candles_count:
        candle_list = candle_list[-candles_count:]

    # Vytvor DataFrame
    df = pd.DataFrame(candle_list)
    df["Date"] = pd.to_datetime(df["Date"], utc=True)
    df.set_index("Date", inplace=True)
    df.index = df.index.tz_localize(None)
    df.sort_index(inplace=True)
    df = df[~df.index.duplicated(keep="last")]
    df = df.dropna(subset=["Open", "High", "Low", "Close"])
    df.columns = [c.capitalize() for c in df.columns]
    df.rename(columns={"Open": "Open", "High": "High", "Low": "Low",
                        "Close": "Close", "Volume": "Volume"}, inplace=True)

    # Resample 4h/12h
    if use_resample:
        rule = {"15m": "15min", "4h": "4h", "12h": "12h"}.get(interval, interval)
        df = resample_ohlcv(df, rule)
        if df.empty:
            raise HTTPException(404, "Resample vrátil prázdny DataFrame")

    # Heikin Ashi
    if ha:
        df = heikin_ashi(df)

    # Indikátory
    ind_set = set(i.strip() for i in indicators.split(",") if i.strip())

    # (rovnaká logika indikátorov ako predtým)
    ema20 = ema50 = ema200 = None
    if "ema" in ind_set:
        ema20  = df["Close"].ewm(span=20,  adjust=False).mean()
        ema50  = df["Close"].ewm(span=50,  adjust=False).mean()
        ema200 = df["Close"].ewm(span=200, adjust=False).mean()

    rsi_s = None
    if "rsi" in ind_set:
        rsi_s = calc_rsi(df["Close"], 14)

    adx_df = None
    if "adx" in ind_set:
        adx_df = calc_adx(df, 14)

    macd_l = macd_s = macd_h = None
    if "macd" in ind_set:
        macd_l, macd_s, macd_h = calc_macd(df["Close"])

    bb_upper = bb_mid = bb_lower = None
    if "bb" in ind_set:
        bb_upper, bb_mid, bb_lower = calc_bollinger(df["Close"])

    obv_s = None
    if "obv" in ind_set:
        obv_s = calc_obv(df)

    stoch_k = stoch_d = None
    if "stochrsi" in ind_set:
        stoch_k, stoch_d = calc_stoch_rsi(df["Close"])

    ichi = None
    if "ichimoku" in ind_set:
        _ichi = calc_ichimoku(df)
        # calc_ichimoku vracia tuple (tenkan, kijun, span_a, span_b, chikou)
        ichi = {"tenkan": _ichi[0], "kijun": _ichi[1], "span_a": _ichi[2], "span_b": _ichi[3], "chikou": _ichi[4]}

    # Zostav výstup
    def safe(v):
        if v is None: return None
        try:
            f = float(v)
            return None if (f != f) else round(f, 6)
        except: return None

    # Zostav výstup
    is_intraday = interval in ("15m", "1h", "4h", "12h")

    def to_date_str(ts):
        try:
            d = pd.Timestamp(ts)
            if is_intraday:
                return int(d.timestamp())
            return d.strftime("%Y-%m-%d")
        except: return str(ts)

    result = []
    for i in range(len(df)):
        row = df.iloc[i]
        point = {
            "time":   to_date_str(df.index[i]),
            "open":   safe(row.get("Open")),
            "high":   safe(row.get("High")),
            "low":    safe(row.get("Low")),
            "close":  safe(row.get("Close")),
            "volume": safe(row.get("Volume")),
        }
        if ema20  is not None: point["ema20"]  = safe(ema20.iloc[i])
        if ema50  is not None: point["ema50"]  = safe(ema50.iloc[i])
        if ema200 is not None: point["ema200"] = safe(ema200.iloc[i])
        if rsi_s  is not None: point["rsi"]    = safe(rsi_s.iloc[i])
        if adx_df is not None:
            point["adx"]      = safe(adx_df["adx"].iloc[i])
            point["di_plus"]  = safe(adx_df["di_plus"].iloc[i])
            point["di_minus"] = safe(adx_df["di_minus"].iloc[i])
        if macd_l is not None:
            point["macd"]        = safe(macd_l.iloc[i])
            point["macd_signal"] = safe(macd_s.iloc[i])
            point["macd_hist"]   = safe(macd_h.iloc[i])
        if bb_upper is not None:
            point["bb_upper"] = safe(bb_upper.iloc[i])
            point["bb_mid"]   = safe(bb_mid.iloc[i])
            point["bb_lower"] = safe(bb_lower.iloc[i])
        if obv_s   is not None: point["obv"]     = safe(obv_s.iloc[i])
        if stoch_k is not None:
            point["stoch_k"] = safe(stoch_k.iloc[i])
            point["stoch_d"] = safe(stoch_d.iloc[i])
        if ichi is not None:
            for k in ["tenkan","kijun","chikou","span_a","span_b"]:
                point[k] = safe(ichi[k].iloc[i]) if k in ichi else None
        result.append(point)

    patterns = []
    if not is_intraday and len(df) >= 3:
        try: patterns = detect_patterns(df)
        except Exception as e: print(f"  Pattern err: {e}")

    # Frontend môže požiadať iba o posledný blok a staršiu históriu dopĺňať
    # pri posune doľava. Výpočty indikátorov stále bežia nad celým cached DF,
    # takže prvý bod každého bloku má korektné warm-up hodnoty.
    if before:
        try:
            boundary = int(before) if is_intraday else before[:10]
            result = [point for point in result if point["time"] < boundary]
        except (TypeError, ValueError):
            raise HTTPException(400, "Neplatný parameter before")
    available_count = len(result)
    if limit and available_count > limit:
        result = result[-limit:]
    has_more = available_count > len(result)

    return {"symbol": sym, "name": sym, "interval": interval,
            "count": len(result), "data": result, "instrumentId": iid,
            "hasMore": has_more,
            "patterns": patterns}


@app.post("/api/ohlcv/batch")
async def get_ohlcv_batch(request: Request):
    """
    Batch verzia /api/ohlcv — načíta viac symbolov paralelne v jednom requeste.
    Body: {"requests": [{"symbol":"AAPL","period":"6mo","interval":"1d","indicators":"ema,rsi","ha":0,"account":"1"}, ...]}
    Response: {"AAPL|6mo|1d|0": {data...}, "MSFT|6mo|1d|0": {data...}, ...}
    """
    body = await request.json()
    reqs = body.get("requests", [])
    if not reqs:
        return {}

    def fetch_one(req):
        sym      = req.get("symbol", "").upper().strip()
        period   = req.get("period", "1y")
        interval = req.get("interval", "1d")
        inds     = req.get("indicators", "")
        ha       = int(req.get("ha", 0))
        account  = str(req.get("account", "1"))
        refresh  = int(req.get("refresh", 1))
        limit    = int(req.get("limit", 0))
        before   = str(req.get("before", ""))
        key      = f"{sym}|{period}|{interval}|{ha}"
        try:
            result = get_ohlcv(
                symbol=sym, period=period, interval=interval,
                indicators=inds, ha=ha, account=account, refresh=refresh,
                limit=limit, before=before,
            )
            return key, result
        except HTTPException as e:
            return key, {"error": e.detail}
        except Exception as e:
            return key, {"error": str(e)}

    def run_batch():
        max_workers = min(len(reqs), 4)
        results = {}
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            from concurrent.futures import as_completed
            futures = {pool.submit(fetch_one, r): r for r in reqs}
            for future in as_completed(futures):
                key, data = future.result()
                results[key] = data
        return results

    # POZOR: čakanie na futures MUSÍ bežať mimo event loopu — synchrónne
    # as_completed/result() v async funkcii zmrazí celú appku na dobu batchu
    # (eToro fetch má 6-20 s timeouty a batch beží pri každom multi-chart loade).
    from fastapi.concurrency import run_in_threadpool
    return await run_in_threadpool(run_batch)


# ── INDIKÁTORY ────────────────────────────────────────────────────────────────

def safe(v):
    """None/NaN → None, inak round na 6 desatinnych miest."""
    if v is None:
        return None
    try:
        f = float(v)
        return None if math.isnan(f) or math.isinf(f) else round(f, 6)
    except Exception:
        return None

def calc_ema(series: pd.Series, period: int) -> pd.Series:
    return series.ewm(span=period, adjust=False).mean()

def calc_rsi(series: pd.Series, period: int = 14) -> pd.Series:
    delta = series.diff()
    gain  = delta.clip(lower=0)
    loss  = (-delta).clip(lower=0)
    avg_gain = gain.ewm(com=period - 1, min_periods=period).mean()
    avg_loss = loss.ewm(com=period - 1, min_periods=period).mean()
    rs = avg_gain / avg_loss.replace(0, float('nan'))
    return 100 - (100 / (1 + rs))

def calc_adx(df: pd.DataFrame, period: int = 14) -> pd.DataFrame:
    high, low, close = df["High"], df["Low"], df["Close"]
    prev_close = close.shift(1)
    prev_high  = high.shift(1)
    prev_low   = low.shift(1)

    tr = pd.concat([
        high - low,
        (high - prev_close).abs(),
        (low  - prev_close).abs(),
    ], axis=1).max(axis=1)

    dm_plus  = ((high - prev_high).clip(lower=0)).where(
        (high - prev_high) > (prev_low - low), 0)
    dm_minus = ((prev_low - low).clip(lower=0)).where(
        (prev_low - low) > (high - prev_high), 0)

    atr     = tr.ewm(com=period - 1, min_periods=period).mean()
    di_plus  = 100 * dm_plus.ewm(com=period - 1,  min_periods=period).mean() / atr.replace(0, float('nan'))
    di_minus = 100 * dm_minus.ewm(com=period - 1, min_periods=period).mean() / atr.replace(0, float('nan'))

    dx  = (100 * (di_plus - di_minus).abs() / (di_plus + di_minus).replace(0, float('nan')))
    adx = dx.ewm(com=period - 1, min_periods=period).mean()

    return pd.DataFrame({"adx": adx, "di_plus": di_plus, "di_minus": di_minus})

def calc_macd(series: pd.Series, fast=12, slow=26, signal=9):
    ema_fast   = series.ewm(span=fast,   adjust=False).mean()
    ema_slow   = series.ewm(span=slow,   adjust=False).mean()
    macd_line  = ema_fast - ema_slow
    signal_line= macd_line.ewm(span=signal, adjust=False).mean()
    histogram  = macd_line - signal_line
    return macd_line, signal_line, histogram

def calc_obv(df: pd.DataFrame) -> pd.Series:
    """On Balance Volume — kumulatívny objem podľa smeru sviečky."""
    direction = df["Close"].diff().apply(lambda x: 1 if x > 0 else (-1 if x < 0 else 0))
    return (direction * df["Volume"]).cumsum()

def calc_stoch_rsi(series: pd.Series, rsi_period=14, stoch_period=14, k=3, d=3) -> tuple:
    """Stochastic RSI — %K a %D."""
    rsi = calc_rsi(series, rsi_period)
    rsi_min = rsi.rolling(stoch_period).min()
    rsi_max = rsi.rolling(stoch_period).max()
    stoch = 100 * (rsi - rsi_min) / (rsi_max - rsi_min).replace(0, float('nan'))
    k_line = stoch.rolling(k).mean()
    d_line = k_line.rolling(d).mean()
    return k_line, d_line

def calc_bollinger(series: pd.Series, period: int = 20, std: float = 2.0):
    sma   = series.rolling(period).mean()
    sigma = series.rolling(period).std()
    upper = sma + std * sigma
    lower = sma - std * sigma
    return upper, sma, lower

def calc_ichimoku(df: pd.DataFrame):
    high, low = df["High"], df["Low"]

    tenkan  = (high.rolling(9).max()  + low.rolling(9).min())  / 2
    kijun   = (high.rolling(26).max() + low.rolling(26).min()) / 2
    span_a  = ((tenkan + kijun) / 2).shift(26)
    span_b  = ((high.rolling(52).max() + low.rolling(52).min()) / 2).shift(26)
    chikou  = df["Close"].shift(-26)

    return tenkan, kijun, span_a, span_b, chikou

# ── OHLCV + INDIKÁTORY ────────────────────────────────────────────────────────

def heikin_ashi(df: pd.DataFrame) -> pd.DataFrame:
    """Vypočíta Heikin Ashi sviečky z OHLC dát."""
    ha = pd.DataFrame(index=df.index)
    ha["Close"] = (df["Open"] + df["High"] + df["Low"] + df["Close"]) / 4
    ha["Open"]  = 0.0
    ha.iloc[0, ha.columns.get_loc("Open")] = (df["Open"].iloc[0] + df["Close"].iloc[0]) / 2
    for i in range(1, len(ha)):
        ha.iloc[i, ha.columns.get_loc("Open")] = (ha["Open"].iloc[i-1] + ha["Close"].iloc[i-1]) / 2
    ha["High"]   = pd.concat([ha["Open"], ha["Close"], df["High"]], axis=1).max(axis=1)
    ha["Low"]    = pd.concat([ha["Open"], ha["Close"], df["Low"]],  axis=1).min(axis=1)
    ha["Volume"] = df["Volume"]
    return ha

def resample_ohlcv(df, rule):
    return df.resample(rule, closed="left", label="left").agg(
        {"Open":"first","High":"max","Low":"min","Close":"last","Volume":"sum"}
    ).dropna(subset=["Open","Close"])



# ══ PREDICTIVE CHART — routes ══════════════════════════════

@app.get("/api/search/predictive")
def search_ticker_predictive(q: str = ""):
    if not q or len(q) < 1:
        return {"results": []}
    try:
        results = yf.Search(q, max_results=8)
        quotes = results.quotes
        out = []
        for item in quotes:
            symbol = item.get("symbol", "")
            name   = item.get("shortname") or item.get("longname") or ""
            qtype  = item.get("quoteType", "")
            exch   = item.get("exchDisp") or item.get("exchange") or ""
            if symbol and qtype in ("EQUITY", "ETF", "INDEX", "MUTUALFUND"):
                out.append({"symbol": symbol, "name": name, "type": qtype, "exchange": exch})
        return {"results": out}
    except Exception as e:
        return {"results": [], "error": str(e)}



_CHART_WEEKLY_PERIOD_TO_COUNT = {"1y": 52, "2y": 104, "3y": 156, "5y": 260}


def _etoro_display_candles(sym: str, interval: str, count: int, account: str = "1") -> list | None:
    """Rovnaké OHLCV dáta ako Grafy panel (eToro broker feed z /api/ohlcv) — LEN
    na vykreslenie sviečok v Analytike. Signály/backtest/predikcia zostávajú
    počítané z yfinance/Massive (nemenené). Fail-soft: None → volajúci použije
    pôvodné yfinance sviečky (ticker mimo eToro univerza, proxy nedostupný...)."""
    try:
        fetch_period = "5y" if interval == "1wk" else "6mo"
        payload = get_ohlcv(symbol=sym, period=fetch_period, interval=interval,
                            indicators="", ha=0, account=account, refresh=1,
                            limit=count, before="")
    except Exception:
        return None
    rows = payload.get("data") or []
    out = []
    for r in rows:
        o, h, l, c = r.get("open"), r.get("high"), r.get("low"), r.get("close")
        if None in (o, h, l, c):
            continue
        try:
            ts = int(pd.Timestamp(r["time"]).timestamp())
        except Exception:
            continue
        out.append({"time": ts, "open": o, "high": h, "low": l, "close": c,
                    "volume": r.get("volume")})
    return out if len(out) >= 2 else None


@app.get("/api/chart")
def get_chart(
    ticker: str = "AAPL",
    period: str = "2y",
    reoptimize: bool = False,
    detail: str = "advanced",
):
    import logging
    detail_mode = "advanced" if str(detail).lower() == "advanced" else "basic"
    logging.warning(f"[CHART] Request received: ticker={ticker} period={period} detail={detail_mode}")
    print(f"[CHART] Request: {ticker} {period} {detail_mode}", flush=True)
    try:
        # Zdieľaná 30-min yf cache (rovnaká ako scanner/RS) — opakované otvorenie
        # tickera neťahá yfinance znova. prefer_massive=False + munge_dots=False
        # zachovávajú pôvodné správanie (yfinance primárny, bodkové LSE tickery).
        raw = _yf_download_cached(ticker, period, "1wk",
                                  prefer_massive=False, munge_dots=False)
        if raw is not None and isinstance(raw.columns, pd.MultiIndex):
            raw.columns = raw.columns.get_level_values(0)
        if raw is None or raw.empty:
            # yfinance zlyhal (rate-limit/timeout) → skús Massive ako záchranu
            mb = _massive_daily_bars(ticker, period, "1wk")
            if mb is not None and len(mb) >= 2:
                print(f"[CHART] yfinance prázdne → Massive fallback ({len(mb)} rows)", flush=True)
                raw = mb
            else:
                raise HTTPException(404, f"No data for {ticker}")

        print(f"[CHART] Step 2: downloaded {len(raw)} rows", flush=True)

        raw = raw.dropna(subset=["Open", "High", "Low", "Close"])
        print(f"[CHART] Step 3: cleaned {len(raw)} rows", flush=True)

        # Drop current incomplete week — last candle is only closed if its
        # Monday-open week has fully passed (i.e. its date < this Monday)
        # Detect if last candle is current incomplete week
        today = pd.Timestamp.now("UTC").tz_localize(None).normalize().tz_localize(None)
        this_monday = today - pd.Timedelta(days=today.weekday())
        last_ts = pd.Timestamp(raw.index[-1])
        if last_ts.tz is not None:
            last_ts = last_ts.tz_localize(None)
        current_week_open = last_ts >= this_monday

        df      = add_indicators(raw)
        candles = df_to_candles(df)

        # Bežné otvorenie používa posledné uložené váhy a iba jeden backtest.
        # Dvojitý default→optimalizovaný beh sa spustí výhradne po explicitnom
        # kliknutí na „Prepočítať váhy“ (reoptimize=1).
        keys     = ["ema", "rsi", "macd", "vol", "ichi", "stoch"]

        def hit_rate_weights(hit_rate, keys, floor=0.05, cap_below=0.10, cap_above=0.50):
            """Convert hit rates to weights.
            - Above 50%: proportional to excess hit rate
            - Below 50%: fixed floor (5%), capped at cap_below (10%)
            - All indicators always get at least floor weight
            """
            MIN_GUARANTEED = floor  # guaranteed minimum for every indicator

            raw = {}
            for k in keys:
                hr = hit_rate.get(k)
                try:
                    hr = float(hr) if hr is not None else None
                except (TypeError, ValueError):
                    hr = None
                if hr is None or hr < 50:
                    raw[k] = MIN_GUARANTEED
                else:
                    raw[k] = MIN_GUARANTEED + (hr - 50.0)

            # Normalize to sum=1
            total = sum(raw.values())
            w = {k: raw[k] / total for k in keys}

            # Clamp: below-50% indicators capped at cap_below after normalization
            for _ in range(10):  # max 10 iterations to prevent infinite loop
                clamped = False
                for k in keys:
                    hr = hit_rate.get(k)
                    try:
                        hr = float(hr) if hr is not None else None
                    except (TypeError, ValueError):
                        hr = None
                    if (hr is None or hr < 50) and w[k] > cap_below:
                        w[k] = cap_below
                        clamped = True
                if not clamped:
                    break
                total2 = sum(w.values())
                if total2 > 0:
                    w = {k: v / total2 for k, v in w.items()}

            # Round and fix sum
            w = {k: round(v, 4) for k, v in w.items()}
            diff = round(1.0 - sum(w.values()), 4)
            if diff != 0:
                best = max((k for k in keys if hit_rate.get(k, 0) >= 50), key=lambda k: w[k], default=keys[0])
                w[best] = round(w[best] + diff, 4)
            return w

        log = load_weights_log()
        ticker_key = ticker.upper()
        saved_entry = log.get(ticker_key, {}) if isinstance(log.get(ticker_key), dict) else {}
        saved_weights = saved_entry.get("weights")
        if isinstance(saved_weights, dict) and all(k in saved_weights for k in keys):
            final_weights = {k: float(saved_weights[k]) for k in keys}
            weights_source = "stored-hit-rate"
        else:
            final_weights = DEFAULT_WEIGHTS.copy()
            weights_source = "default"

        bt_default = None
        if reoptimize:
            print("[CHART] Step 6: recalculating indicator weights...", flush=True)
            bt_default = run_backtest_cached(df, weights=DEFAULT_WEIGHTS)
            try:
                hit_rate = bt_default.get("indicator_hit_rate", {})
                final_weights = hit_rate_weights(hit_rate, keys)
                weights_source = "hit-rate"
            except Exception as e_w:
                print(f"[CHART] weights error: {e_w}, using defaults", flush=True)
                final_weights = DEFAULT_WEIGHTS.copy()
                weights_source = "default"

        print(f"[CHART] Step 7: backtest ({weights_source})...", flush=True)
        backtest = run_backtest_cached(df, weights=final_weights)
        print("[CHART] Step 8: backtest done", flush=True)

        if reoptimize:
            log[ticker_key] = {
                "weights": final_weights,
                "optimized_at": pd.Timestamp.now("UTC").tz_localize(None).strftime("%Y-%m-%d"),
                "candles": len(df),
                "accuracy_default": bt_default.get("direction_accuracy") if bt_default else None,
                "accuracy_optimized": backtest.get("direction_accuracy"),
            }
            save_weights_log(log)
            saved_entry = log[ticker_key]

        # Cache key pre ML + HMM: kým je posledná uzavretá sviečka rovnaká,
        # výsledok fitu je identický → cacheujeme do najbližšej novej sviečky.
        try:
            _model_key = f"{ticker}:{period}:1wk:{pd.Timestamp(df.index[-1]).date()}:{len(df)}"
        except Exception:
            _model_key = None

        print("[CHART] Step 11: ML training...", flush=True)
        if detail_mode == "advanced" and ENABLE_PREDICTIVE_ML:
            try:
                _ml_model, ml_acc, ml_bull_prob = train_ml_model(df, cache_key=_model_key)
                print("[CHART] Step 12: ML done", flush=True)
            except Exception:
                ml_acc, ml_bull_prob = None, 0.5
        else:
            ml_acc, ml_bull_prob = None, 0.5
            print("[CHART] Step 12: ML skipped (ENABLE_PREDICTIVE_ML=0)", flush=True)

        print("[CHART] Step 12b: HMM regime...", flush=True)
        if detail_mode == "advanced" and ENABLE_PREDICTIVE_HMM:
            try:
                regime_info = detect_market_regime(df, cache_key=_model_key)
            except Exception as e:
                regime_info = {"error": str(e)[:80]}
        else:
            regime_info = {"label": "disabled", "disabled": True,
                           "reason": "ENABLE_PREDICTIVE_HMM=0"}

        pred = predict_next_candle(df, weights=final_weights, ml_bull_prob=ml_bull_prob)
        pred_default = predict_next_candle(df, weights=DEFAULT_WEIGHTS) if detail_mode == "advanced" else None
        opt_weights = final_weights

        last_ts = candles[-1]["time"]
        next_ts = last_ts + 7 * 24 * 3600

        pred_candle = {
            "time":  next_ts,
            "open":  pred["open"],
            "high":  pred["high"],
            "low":   pred["low"],
            "close": pred["close"],
        }

        # full backtest overlay: timestamp + pred/actual close
        overlay = []
        for r in backtest.get("detail", []):
            try:
                ts = int(pd.Timestamp(r["date"]).timestamp())
                overlay.append({
                    "time":         ts,
                    "pred_open":    r["pred_open"],
                    "pred_high":    r["pred_high"],
                    "pred_low":     r["pred_low"],
                    "pred_close":   r["pred_close"],
                    "actual_close": r["actual_close"],
                    "correct":      r["correct"],
                    "pred_change_pct":   r.get("pred_change_pct"),
                    "actual_change_pct": r.get("actual_change_pct"),
                })
            except Exception:
                pass

        # If current week is open, append its prediction to overlay
        cur_pred = None
        if current_week_open and len(candles) >= 2:
            cur = candles[-1]
            try:
                cur_pred = predict_next_candle(add_indicators(raw.iloc[:-1]), weights=opt_weights)
                overlay.append({
                    "time":         cur["time"],
                    "pred_open":    cur_pred["open"],
                    "pred_high":    cur_pred["high"],
                    "pred_low":     cur_pred["low"],
                    "pred_close":   cur_pred["close"],
                    "actual_close": cur["close"],
                    "correct":      None,   # not yet decided
                })
            except Exception:
                pass

        # Candle for current open week slot
        if cur_pred is not None:
            pred_current_candle = {
                "time":  candles[-1]["time"],
                "open":  cur_pred["open"],
                "high":  cur_pred["high"],
                "low":   cur_pred["low"],
                "close": cur_pred["close"],
            }
        else:
            pred_current_candle = None

        # Serialize indicator series for overlay/subpanel
        def ind_series(col):
            out = []
            for ts, row in df.iterrows():
                v = safe_float(row.get(col))
                if v is None: continue
                out.append({"time": int(pd.Timestamp(ts).timestamp()), "value": v})
            return out

        indicators = {
            "ema10":      ind_series("ema10"),
            "ema20":      ind_series("ema20"),
            "ichi_tenkan":ind_series("ichi_tenkan"),
            "ichi_kijun": ind_series("ichi_kijun"),
            "ichi_sa":    ind_series("ichi_sa"),
            "ichi_sb":    ind_series("ichi_sb"),
            "rsi":        ind_series("rsi"),
            "macd":       ind_series("macd"),
            "macd_sig":   ind_series("macd_sig"),
            "macd_hist":  ind_series("macd_hist"),
            "stoch_k":    ind_series("stoch_k"),
            "stoch_d":    ind_series("stoch_d"),
            "adx":        ind_series("adx"),
            "di_plus":    ind_series("di_plus"),
            "di_minus":   ind_series("di_minus"),
        }

        # Daily timeframe — 3 months of daily candles + buy signal logic
        daily_signal      = 0.0
        daily_candles     = []
        daily_indicators  = {}
        daily_buy_signals = []
        signal_outcome_summary = {}
        signal_outcome_segments = {}
        weekly_bias       = {}
        today_score       = 0
        today_raw_score   = 0
        today_details     = {}

        try:
            # Dva roky dávajú priestor na vyhodnotenie 30/60/90 obchodných
            # sviečok aj pre staršie signály. Do grafu sa naďalej posiela len tail.
            raw_d = _yf_download_cached(ticker, "2y", "1d",
                                        prefer_massive=False, munge_dots=False)
            if isinstance(raw_d.columns, pd.MultiIndex):
                raw_d.columns = raw_d.columns.get_level_values(0)
            raw_d = raw_d.dropna(subset=["Open", "High", "Low", "Close"])

            if len(raw_d) >= 20:
                df_d = add_indicators(raw_d)

                # --- Weekly bias check (Donchian 20w + SMA50w + EMA10/20) ---
                weekly_trend = _weekly_trend(df)
                weekly_bullish = _weekly_trend_to_bullish(weekly_trend) or False
                w_composite = pred["composite"]
                weekly_bias = {
                    "bullish":     weekly_bullish,
                    "composite":   round(w_composite * 100, 1),
                    "trend":       weekly_trend,  # nový 5-stupňový label
                    "above_kumo":  None,
                    "ema_bull":    weekly_trend.get("ema_bull") if weekly_trend else False,
                }

                # --- Daily signal computation ---
                last_d  = df_d.iloc[-1]
                lc_d    = float(last_d["Close"])
                ema_d   = float(np.tanh(((last_d["ema10"]-last_d["ema20"])/lc_d)*20))
                rsi_d   = float(last_d["rsi"])
                rsi_ds  = -((rsi_d-65)/35) if rsi_d>65 else (35-rsi_d)/35 if rsi_d<35 else 0.0
                macd_ds = float(np.tanh(last_d["macd_hist"]/(lc_d*0.001+1e-9)))
                daily_signal = float(np.clip((ema_d*0.4 + rsi_ds*0.3 + macd_ds*0.3), -1.0, 1.0))

                # --- Buy signal scoring (long only) — zdieľaná score_signal_day (c1..c4) ---
                # Rolling z-score na celom DF, potom zarovnané na posledných 90 sviečok
                _zscore_series_d = rolling_zscore(df_d["Close"])
                zscore_slice = _zscore_series_d.iloc[-90:].values

                # Load existing signals for this ticker
                slog = load_signals_log()
                ticker_slog = slog.get(ticker.upper(), {})

                # Score all closed candles (exclude last if today is open)
                today_date = pd.Timestamp.now("UTC").tz_localize(None).normalize().tz_localize(None)
                latest_closed_date = latest_closed_daily_date(df_d)
                df_score   = df_d.iloc[-90:].reset_index()
                # Auto-kontext budget: max HMM fitov navyše na jedno načítanie grafu
                # (okrem poslednej sviečky, ktorá ho dostane vždy). Drží /api/chart
                # svižné a zároveň postupne dopĺňa kontext do starých signálov pri
                # bežnom prezeraní — manuálny backfill tak nie je nutný.
                _ctx_budget = 4 if ENABLE_SIGNAL_CONTEXT_BACKFILL else 0

                for i in range(5, len(df_score)):
                    row      = df_score.iloc[i]
                    row_date = pd.Timestamp(row.iloc[0]).tz_localize(None) if pd.Timestamp(row.iloc[0]).tzinfo else pd.Timestamp(row.iloc[0])
                    # Skip today's candle — not yet closed
                    if row_date.date() >= today_date.date():
                        continue
                    date_key = str(row_date.date())
                    zscore   = float(zscore_slice[i]) if i < len(zscore_slice) else 0.0
                    sc, details = score_signal_day(row, zscore)
                    if sc >= 2:
                        ts = int(pd.Timestamp(row.iloc[0]).timestamp())
                        tier = signal_tier(sc, details["trend"])
                        # Save to log if not already there
                        if date_key not in ticker_slog:
                            ticker_slog[date_key] = {
                                "score":   sc,
                                "tier":    tier,
                                "close":   round(float(row["Close"]), 2),
                                "details": details,
                                "rules_version": SIGNAL_RULES_VERSION,
                            }
                        # Auto-kontext: dopočítaj chýbajúci kontext (nový aj starý
                        # záznam). Posledná uzavretá sviečka vždy; staršie do vyčerpania
                        # rozpočtu. Žiadny look-ahead — build_signal_context si dáta oreže.
                        entry = ticker_slog[date_key]
                        if ENABLE_SIGNAL_CONTEXT_BACKFILL and not entry.get("context"):
                            is_latest = latest_closed_date is not None and row_date.normalize() == latest_closed_date
                            if is_latest or _ctx_budget > 0:
                                ctx = build_signal_context(df_d, row_date, details, zscore, weekly_bullish)
                                if ctx:
                                    entry["context"] = ctx
                                    if not is_latest:
                                        _ctx_budget -= 1
                        # Always include in chart signals (from log if exists)
                        saved = ticker_slog.get(date_key, {})
                        daily_buy_signals.append({
                            "time":  ts,
                            "score": saved.get("score", sc),
                            "tier":  saved.get("tier", tier),
                            "close": saved.get("close", round(float(row["Close"]), 2)),
                        })

                # Persist any new signals
                if ticker_slog:
                    slog[ticker.upper()] = ticker_slog
                    save_signals_log(slog)

                # Today's score (informational only — not saved, candle not closed)
                # Weekly bias must still confirm for today's live marker
                try:
                    today_z = float(_zscore_series_d.iloc[-1]) if len(_zscore_series_d) else 0.0
                    today_sc, today_details = score_signal_day(df_d.iloc[-1], today_z)
                    today_raw_score = int(today_sc)
                    today_score = int(today_sc) if weekly_bullish else 0
                except Exception:
                    today_score = 0
                    today_raw_score = 0
                    today_details = {}

                # Also show older saved signals (beyond 90 days) if any
                for date_key, sig in ticker_slog.items():
                    ts = int(pd.Timestamp(date_key).timestamp())
                    if not any(s["time"] == ts for s in daily_buy_signals):
                        daily_buy_signals.append({
                            "time":  ts,
                            "score": sig["score"],
                            "tier":  sig.get("tier", signal_tier(sig["score"])),
                            "close": sig["close"],
                        })
                daily_buy_signals.sort(key=lambda x: x["time"])
                # Pripoj regime label z uloženého kontextu (pre per-regime analytiku)
                for _s in daily_buy_signals:
                    _dk = str(pd.Timestamp(int(_s["time"]), unit="s").date())
                    _ctx = (ticker_slog.get(_dk) or {}).get("context") or {}
                    _reg = (_ctx.get("regime") or {}).get("label")
                    if _reg:
                        _s["regime"] = _reg
                if detail_mode == "advanced" and ENABLE_SIGNAL_ANALYTICS:
                    daily_buy_signals, signal_outcome_summary, signal_outcome_segments = build_signal_outcome_analytics(
                        df_d, daily_buy_signals
                    )
                else:
                    signal_outcome_summary = {"disabled": True, "reason": "ENABLE_SIGNAL_ANALYTICS=0"}
                    signal_outcome_segments = {}

                # All candles for chart (last 90 days)
                for ts, row in df_d.iloc[-90:].iterrows():
                    o = safe_float(row["Open"]); h = safe_float(row["High"])
                    l = safe_float(row["Low"]);  c = safe_float(row["Close"])
                    if None not in (o, h, l, c):
                        daily_candles.append({
                            "time":  int(pd.Timestamp(ts).timestamp()),
                            "open": o, "high": h, "low": l, "close": c,
                            # volume: pattern overlay potrebuje objemové potvrdenie breakoutu;
                            # LWC candlestick series extra pole ignoruje (aditívne, fail-soft)
                            "volume": safe_float(row.get("Volume")),
                        })

                # Daily indicator series for overlay
                def daily_ind(col):
                    out = []
                    for ts, row in df_d.iloc[-90:].iterrows():
                        v = safe_float(row.get(col))
                        if v is not None:
                            out.append({"time": int(pd.Timestamp(ts).timestamp()), "value": v})
                    return out

                daily_indicators = {
                    "ema20":       daily_ind("ema20"),
                    "ema50":       daily_ind("ema10"),   # use ema10 as proxy for ema50-like
                    "ichi_kijun":  daily_ind("ichi_kijun"),
                    "ichi_sa":     daily_ind("ichi_sa"),
                    "ichi_sb":     daily_ind("ichi_sb"),
                }

                # Mini chart: last 10 candles (for left panel)
                # already in daily_candles, just take last 10
        except Exception:
            daily_signal = 0.0

        # Modulate prediction with daily signal (max ±15% adjustment)
        def apply_daily_filter(pred_dict, daily_sig):
            orig_comp = pred_dict["composite"]
            lc        = pred_dict["open"]
            atr       = pred_dict["signals"]["atr"]["value"]
            # If daily confirms weekly → amplify slightly, if contradicts → dampen
            if (orig_comp > 0 and daily_sig > 0) or (orig_comp < 0 and daily_sig < 0):
                mod = 1.0 + abs(daily_sig) * 0.15
            else:
                mod = 1.0 - abs(daily_sig) * 0.15
            new_comp  = float(np.clip(orig_comp * mod, -1.0, 1.0))
            new_close = lc * (1 + new_comp * 0.6) + new_comp * atr * 0.4
            mid       = (lc + new_close) / 2
            return {**pred_dict,
                    "close":     round(new_close, 4),
                    "high":      round(mid + atr * 0.75, 4),
                    "low":       round(mid - atr * 0.75, 4),
                    "composite": round(new_comp, 4),
                    "daily_signal": round(daily_sig, 4)}

        pred = apply_daily_filter(pred, daily_signal)

        # Earnings dates — bulk kalendár → Finnhub per-symbol → yfinance
        # (jednotná logika v _earnings_next_date, zdieľaná s GET /api/earnings/{symbol})
        earnings_dates = []
        try:
            d = _earnings_next_date(ticker)
            if d:
                earnings_dates.append(int(pd.Timestamp(d).timestamp()))
        except Exception:
            pass

        # Vykresľovacie sviečky = rovnaký eToro broker feed ako Grafy panel
        # (/api/ohlcv), aby graf v Analytike vyzeral rovnako ako v Grafoch.
        # Signály/backtest/predikcia vyššie zostávajú počítané z yfinance/Massive
        # (candles/daily_candles interné premenné) — mení sa len to, čo sa vracia
        # na vykreslenie. Fail-soft: bez eToro dát (ticker mimo univerza, proxy
        # nedostupný) ostáva pôvodná yfinance sada.
        display_candles = _etoro_display_candles(
            ticker, "1wk", _CHART_WEEKLY_PERIOD_TO_COUNT.get(period, 104)) or candles
        display_daily_candles = _etoro_display_candles(ticker, "1d", 90) or daily_candles

        response = {
            "detail":              detail_mode,
            "ticker":              ticker.upper(),
            "current_week_open":   current_week_open,
            "daily_candles":       display_daily_candles,
            "daily_signal":        round(daily_signal, 3),
            "daily_indicators":    daily_indicators,
            "daily_buy_signals":   daily_buy_signals,
            "weekly_bias":         weekly_bias,
            "today_score":         today_score,
            "today_raw_score":     today_raw_score,
            "today_details":       today_details,
            "earnings_dates":      sorted(earnings_dates),
            "candles":             display_candles,
            "prediction":          pred,
            "pred_candle":         pred_candle,
            "pred_current_candle": pred_current_candle,
            "indicators":          indicators,
            "backtest": {
                "total":              backtest.get("total_predictions"),
                "direction_accuracy": backtest.get("direction_accuracy"),
                "test_accuracy":      backtest.get("test_accuracy"),
                "test_total":         backtest.get("test_total"),
                "base_rate_up":       backtest.get("base_rate_up"),
                "test_base_rate_up":  backtest.get("test_base_rate_up"),
                "avg_error_pct":      backtest.get("avg_error_pct"),
                "indicator_hit_rate": backtest.get("indicator_hit_rate") if detail_mode == "advanced" else {},
                "overlay":            overlay,
            },
        }
        if detail_mode == "advanced":
            response.update({
                "weights": final_weights,
                "weights_source": weights_source,
                "weights_default": DEFAULT_WEIGHTS,
                "optimized_at": saved_entry.get("optimized_at"),
                "accuracy_opt": backtest.get("direction_accuracy"),
                "accuracy_def": (
                    bt_default.get("direction_accuracy") if bt_default
                    else saved_entry.get("accuracy_default")
                ),
                "pred_default_close": round(pred_default["close"], 4) if pred_default else None,
                "ml_accuracy": ml_acc,
                "ml_bull_prob": round(ml_bull_prob * 100, 1) if ml_bull_prob else None,
                "regime": regime_info,
                "signal_outcome_summary": signal_outcome_summary,
                "signal_outcome_segments": signal_outcome_segments,
            })
        return response
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"[CHART] ERROR: {e}", flush=True)
        print(traceback.format_exc(), flush=True)
        raise HTTPException(500, str(e))



def _finite_float(value, default=None):
    try:
        v = float(value)
        if math.isnan(v) or math.isinf(v):
            return default
        return v
    except Exception:
        return default


def build_setup_assessment(df_d, weekly_bullish: bool, recent_signal: dict | None, signal_count: int, latest_zscore: float | None = None) -> dict:
    """Lightweight ranking layer for Opportunities. It is a decision aid, not a trade instruction."""
    last = df_d.iloc[-1]
    close = _finite_float(last.get("Close"), 0) or 0
    ema20 = _finite_float(last.get("ema20"), close) or close
    kijun = _finite_float(last.get("ichi_kijun"), close) or close
    rsi = _finite_float(last.get("rsi"), 50) or 50
    atr = _finite_float(last.get("atr"), 0) or 0
    vol_ratio = _finite_float(last.get("vol_ratio"), 1) or 1
    macd_hist = _finite_float(last.get("macd_hist"), 0) or 0
    adx = _finite_float(last.get("adx"), 0) or 0

    score = 0
    positive = []
    risk = []

    if weekly_bullish:
        score += 24
        positive.append("weekly trend podporuje long")
    else:
        risk.append("weekly trend este nepotvrdzuje")

    if recent_signal:
        sig_score = int(recent_signal.get("score") or 0)
        sig_tier = recent_signal.get("tier") or signal_tier(sig_score)
        if sig_tier == "buy":
            score += 28
            positive.append(f"novy buy signal {sig_score}/4")
        elif sig_tier == "counter":
            score += 6
            risk.append(f"proti-trendovy dip {sig_score}/4 (downtrend)")
        else:
            score += 16
            positive.append(f"watch signal {sig_score}/4")
    else:
        risk.append("bez cerstveho daily signalu")

    if signal_count >= 3:
        score += 10
        positive.append("opakovane historicke signaly")
    elif signal_count >= 1:
        score += 5

    if close > ema20:
        score += 10
        positive.append("cena nad EMA20")
    else:
        risk.append("cena pod EMA20")

    dist_ema = abs(close - ema20) / close * 100 if close else 0
    if dist_ema <= 6:
        score += 8
        positive.append("nie je daleko od EMA20")
    elif dist_ema > 12:
        score -= 8
        risk.append("cena je uz natiahnuta od EMA20")

    support_dist = min(abs(close - ema20), abs(close - kijun)) / close * 100 if close else 0
    if support_dist <= 3:
        score += 8
        positive.append("blizko EMA/Kijun zony")

    if 38 <= rsi <= 62:
        score += 8
        positive.append("RSI v rozumnom pasme")
    elif rsi > 72:
        score -= 10
        risk.append("RSI prehriate")
    elif rsi < 32:
        score -= 4
        risk.append("RSI velmi slabe")

    atr_pct = atr / close * 100 if close else 0
    if atr_pct <= 4:
        score += 6
    elif atr_pct > 8:
        score -= 8
        risk.append("vysoka volatilita")

    if latest_zscore is not None:
        if latest_zscore <= SIGNAL_ZSCORE_THRESHOLD:
            score += 8
            positive.append(f"statisticky dip (z-score {latest_zscore:.1f})")
        elif latest_zscore >= 1.5:
            score -= 6
            risk.append(f"cena natiahnuta (z-score {latest_zscore:.1f})")

    if vol_ratio >= 1.2:
        score += 5
        positive.append("objem nad priemerom")

    if macd_hist > 0:
        score += 5
        positive.append("MACD momentum pozitivne")

    if adx >= 20:
        score += 4
        positive.append("trend ma silu")

    score = int(max(0, min(100, round(score))))
    grade = "A" if score >= 78 else "B" if score >= 62 else "Watch" if score >= 45 else "Risky"
    return {
        "setup_score": score,
        "setup_grade": grade,
        "positive_factors": positive[:5],
        "risk_flags": risk[:5],
        "metrics": {
            "rsi": round(rsi, 1),
            "atr_pct": round(atr_pct, 2),
            "dist_ema20_pct": round(dist_ema, 2),
            "vol_ratio": round(vol_ratio, 2),
            "zscore": round(latest_zscore, 2) if latest_zscore is not None else None,
        },
    }


def _last_swing_points(values, kind: str, lookback: int = 90, radius: int = 2) -> list[tuple[int, float]]:
    vals = [float(x) for x in values if pd.notna(x)]
    if len(vals) < radius * 2 + 3:
        return []
    start = max(radius, len(vals) - lookback)
    pivots: list[tuple[int, float]] = []
    for i in range(start, len(vals) - radius):
        win = vals[i - radius:i + radius + 1]
        v = vals[i]
        if kind == "high" and v == max(win) and win.count(v) == 1:
            pivots.append((i, v))
        elif kind == "low" and v == min(win) and win.count(v) == 1:
            pivots.append((i, v))
    return pivots[-4:]


def chart_health(df: pd.DataFrame, timeframe: str = "daily") -> dict:
    """Vizuálna kvalita grafu: OK/Risk/Bad ako ľudský filter nad signálmi.

    Neovplyvňuje C1-C4 ani DIP skóre. Má len povedať, či graf vyzerá ako
    zdravý pullback, rizikový bounce alebo poškodený downtrend.
    """
    if df is None or len(df) < 30:
        return {"status": "unknown", "label": "N/A", "score": 0, "reasons": ["málo dát"]}

    close = df["Close"].astype(float)
    high = df["High"].astype(float)
    low = df["Low"].astype(float)
    vol = df["Volume"].astype(float) if "Volume" in df.columns else pd.Series([0] * len(df), index=df.index)
    last_close = float(close.iloc[-1])

    ema20 = df["ema20"].astype(float) if "ema20" in df.columns else close.ewm(span=20, adjust=False).mean()
    ema10 = df["ema10"].astype(float) if "ema10" in df.columns else close.ewm(span=10, adjust=False).mean()
    ema50 = close.ewm(span=50, adjust=False).mean() if len(close) >= 50 else None

    score = 0
    reasons: list[str] = []
    e20 = float(ema20.iloc[-1])
    e10 = float(ema10.iloc[-1])
    e50 = float(ema50.iloc[-1]) if ema50 is not None else None

    if e50 is not None:
        if last_close > e20 > e50:
            score += 2
            reasons.append("cena nad EMA20/EMA50")
        elif last_close < e20 < e50:
            score -= 3
            reasons.append("cena pod EMA20/EMA50")
        elif last_close < e20:
            score -= 1
            reasons.append("cena pod EMA20")
        elif last_close > e20:
            score += 1
            reasons.append("cena nad EMA20")
    else:
        if e10 > e20 and last_close > e20:
            score += 1
            reasons.append("krátky trend nad EMA20")
        elif e10 < e20 and last_close < e20:
            score -= 2
            reasons.append("krátky trend pod EMA20")

    window_high = float(high.tail(min(len(high), 126)).max())
    drawdown = (last_close / window_high - 1.0) * 100 if window_high > 0 else 0.0
    bad_dd = -35 if timeframe == "daily" else -30
    warn_dd = -20 if timeframe == "daily" else -18
    if drawdown <= bad_dd:
        score -= 3
        reasons.append(f"hlboko pod maximom ({drawdown:.0f}%)")
    elif drawdown <= warn_dd:
        score -= 1
        reasons.append(f"ďaleko od maxima ({drawdown:.0f}%)")

    highs = _last_swing_points(high.tail(120).values, "high", lookback=120, radius=2)
    lows = _last_swing_points(low.tail(120).values, "low", lookback=120, radius=2)
    if len(highs) >= 2 and len(lows) >= 2:
        lower_high = highs[-1][1] < highs[-2][1]
        lower_low = lows[-1][1] < lows[-2][1]
        higher_low = lows[-1][1] > lows[-2][1]
        if lower_high and lower_low:
            score -= 3
            reasons.append("nižšie high aj low")
        elif lower_high:
            score -= 1
            reasons.append("nižšie swing high")
        elif higher_low:
            score += 1
            reasons.append("vyššie swing low")

    daily_ret = close.pct_change()
    worst = float(daily_ret.tail(30).min() * 100) if len(daily_ret.dropna()) else 0.0
    red_vol = 0
    if len(df) >= 20 and "Volume" in df.columns:
        recent = df.tail(20).copy()
        vavg = float(vol.tail(60).mean()) if len(vol) >= 20 and float(vol.tail(60).mean()) else 0.0
        if vavg > 0:
            red_vol = int(((recent["Close"] < recent["Open"]) & (recent["Volume"] > vavg * 1.5)).sum())
    if timeframe == "daily" and worst <= -8:
        score -= 2
        reasons.append(f"prudký denný pád {worst:.0f}%")
    if red_vol >= 2:
        score -= 1
        reasons.append("červené volume spiky")
    if last_close > e20 and daily_ret.tail(5).sum() > 0:
        score += 1
        reasons.append("odraz nad EMA20")

    if score >= 2:
        status, label = "ok", "OK"
    elif score <= -3:
        status, label = "bad", "Bad"
    else:
        status, label = "risk", "Risk"

    return {
        "status": status,
        "label": label,
        "score": int(score),
        "drawdown_pct": round(drawdown, 1),
        "reasons": reasons[:4],
    }


NASDAQ100_TICKERS = [
    "AAPL", "ABNB", "ADBE", "ADI", "ADP", "ADSK", "AEP", "AMAT", "AMD", "AMGN",
    "AMZN", "ANSS", "APP", "ARM", "ASML", "AVGO", "AXON", "AZN", "BIIB", "BKNG",
    "BKR", "CCEP", "CDNS", "CDW", "CEG", "CHTR", "CMCSA", "COST", "CPRT", "CRWD",
    "CSCO", "CSGP", "CSX", "CTAS", "CTSH", "DASH", "DDOG", "DXCM", "EA", "EXC",
    "FANG", "FAST", "FTNT", "GEHC", "GFS", "GILD", "GOOG", "GOOGL", "HON", "IDXX",
    "INTC", "INTU", "ISRG", "KDP", "KHC", "KLAC", "LIN", "LRCX", "LULU", "MAR",
    "MCHP", "MDLZ", "MELI", "META", "MNST", "MRVL", "MSFT", "MSTR", "MU", "NFLX",
    "NVDA", "NXPI", "ODFL", "ON", "ORLY", "PANW", "PAYX", "PCAR", "PDD", "PEP",
    "PLTR", "PYPL", "QCOM", "REGN", "ROP", "ROST", "SBUX", "SNPS", "TEAM", "TMUS",
    "TSLA", "TTD", "TTWO", "TXN", "VRSK", "VRTX", "WBD", "WDAY", "XEL", "ZS",
]

SCANNER_MAX_WORKERS = int(os.getenv("SCANNER_MAX_WORKERS", "2" if _LOW_MEMORY else "3"))
SCANNER_YF_TIMEOUT = int(os.getenv("SCANNER_YF_TIMEOUT", "8"))
SCANNER_DIP_UNIVERSE_MAX = int(os.getenv("SCANNER_DIP_UNIVERSE_MAX", "300"))
SCANNER_GC_INTERVAL = int(os.getenv("SCANNER_GC_INTERVAL", "20"))
SCANNER_DEFAULT_DAYS = 3
SIGNAL_PROXIMITY_TOL = 0.005     # fallback keď ATR nie je k dispozícii
SIGNAL_RSI_PULLBACK = 45
SIGNAL_VOLUME_MULT = 1.2
SIGNAL_ZSCORE_THRESHOLD = -1.5   # 60-period rolling z-score ≤ this = cena je štatisticky lacná
DIP_STRONG_THRESHOLD = 90
DIP_VERY_STRONG_THRESHOLD = 100
DIP_SCANNER_VISIBILITY_THRESHOLD = 80  # show all imported WATCH+ DIP rows even without a fresh technical signal
SIGNAL_BUY_THRESHOLD = 3   # 3/4+ = plný buy signál, 2/4 = watch
SIGNAL_OUTCOME_HORIZONS = (30, 60, 90)
SIGNAL_OUTCOME_MOVE_THRESHOLD = 1.5  # fallback keď ATR nie je k dispozícii
# rules v2: prahy škálujú s volatilitou (ATR%) — fixné % znamenalo pre TSLA niečo
# iné než pre MSFT. Signály v logu nesú rules_version pre porovnateľnosť.
SIGNAL_RULES_VERSION = 2
SIGNAL_PROXIMITY_ATR_MULT = 0.35     # tolerancia C1 = 0.35×ATR%, clamp 0.3–1.2 %
SIGNAL_PROXIMITY_MIN = 0.003
SIGNAL_PROXIMITY_MAX = 0.012
SIGNAL_OUTCOME_ATR_MULT = 1.0        # win/loss prah = 1×ATR%, clamp 1.0–3.0 %
SIGNAL_OUTCOME_MIN = 1.0
SIGNAL_OUTCOME_MAX = 3.0


def rolling_zscore(close_series):
    """60-period rolling z-score ceny. Jeden zdroj pravdy pre scanner aj predikt."""
    roll = close_series.rolling(60, min_periods=30)
    return ((close_series - roll.mean()) / roll.std().replace(0, float("nan"))).fillna(0)


def score_signal_day(row, zscore: float) -> tuple[int, dict]:
    """Jediný zdroj pravdy pre denné buy-signal skórovanie (c1..c4).
    Volá scanner aj prediktívna signal history — obe hovoria rovnakým jazykom (skóre /4)."""
    close = float(row["Close"])
    open_ = float(row["Open"])
    ema20 = float(row["ema20"])
    kijun = float(row["ichi_kijun"]) if not math.isnan(float(row["ichi_kijun"])) else close
    rsi = float(row["rsi"]) if not math.isnan(float(row["rsi"])) else 50
    vol = float(row["Volume"])
    vol_ma = float(row["vol_ma"]) if not math.isnan(float(row["vol_ma"])) else vol
    ema10 = float(row["ema10"]) if not math.isnan(float(row["ema10"])) else ema20
    # rules v2: C1 tolerancia škáluje s ATR% (volatilnejší titul = širšia zóna "dotyku")
    prox_tol = SIGNAL_PROXIMITY_TOL
    try:
        atr = float(row["atr"])
        if not math.isnan(atr) and atr > 0 and close > 0:
            prox_tol = min(SIGNAL_PROXIMITY_MAX, max(SIGNAL_PROXIMITY_MIN, SIGNAL_PROXIMITY_ATR_MULT * atr / close))
    except (KeyError, TypeError, ValueError):
        pass
    c1 = abs(close - ema20) / close < prox_tol or abs(close - kijun) / close < prox_tol
    c2 = rsi < SIGNAL_RSI_PULLBACK
    c3 = close > open_ and vol > vol_ma * SIGNAL_VOLUME_MULT
    c4 = zscore <= SIGNAL_ZSCORE_THRESHOLD
    sc = int(sum([c1, c2, c3, c4]))
    # Trend-primárna klasifikácia (per-bar) podľa štruktúry EMA:
    #   up   = ema10 > ema20            → dip v uptrende = buyovateľný (DIP stratégia)
    #   down = ema10 < ema20 a cena < ema20 → proti-trendový dip (falling-knife)
    #   side = prechod/sideways
    if ema10 > ema20:
        trend = "up"
    elif ema10 < ema20 and close < ema20:
        trend = "down"
    else:
        trend = "side"
    details = {
        "ema_kijun_touch": bool(c1),
        "rsi_pullback": bool(c2),
        "bull_volume": bool(c3),
        "zscore_dip": bool(c4),
        "trend": trend,
    }
    return sc, details


def signal_tier(score: int, trend: str = "side") -> str:
    """Trend-primárny tier. O farbe rozhoduje kontext trendu, nie hrubé skóre:
    uptrend = 'buy' (zelená, aj pri 2/4), downtrend = 'counter' (červená,
    proti-trendový dip), inak 'watch' (oranžová). Skóre ostáva ako sila/text."""
    if trend == "up":
        return "buy"
    if trend == "down":
        return "counter"
    return "watch"


SIGNAL_CONTEXT_VERSION = 1


def latest_closed_daily_date(df_daily: pd.DataFrame):
    """Najnovší uzavretý denný bar; dnešný potenciálne otvorený bar ignoruje."""
    if df_daily is None or df_daily.empty:
        return None
    today = pd.Timestamp.now("UTC").tz_localize(None).normalize()
    closed = []
    for value in df_daily.index:
        ts = pd.Timestamp(value)
        if ts.tzinfo is not None:
            ts = ts.tz_localize(None)
        if ts.normalize() < today:
            closed.append(ts.normalize())
    return max(closed) if closed else None


def build_signal_context(
    df_daily: pd.DataFrame,
    signal_date,
    details: dict,
    zscore: float,
    weekly_bullish: bool | None,
    context_source: str = "live",
) -> dict | None:
    """Snapshot kontextu pre nový signál bez použitia budúcich dát."""
    if df_daily is None or df_daily.empty:
        return None
    cutoff = pd.Timestamp(signal_date)
    if cutoff.tzinfo is not None:
        cutoff = cutoff.tz_localize(None)
    cutoff = cutoff.normalize()

    normalized_index = []
    for value in df_daily.index:
        ts = pd.Timestamp(value)
        if ts.tzinfo is not None:
            ts = ts.tz_localize(None)
        normalized_index.append(ts.normalize())
    history = df_daily.loc[[ts <= cutoff for ts in normalized_index]].copy()
    if history.empty:
        return None

    close = float(history["Close"].iloc[-1])
    returns = history["Close"].astype(float).pct_change()
    returns_5d = (close / float(history["Close"].iloc[-6]) - 1) * 100 if len(history) >= 6 else None
    returns_20d = (close / float(history["Close"].iloc[-21]) - 1) * 100 if len(history) >= 21 else None
    volatility_20d = (
        float(returns.iloc[-20:].std()) * math.sqrt(252) * 100
        if len(history) >= 21 else None
    )
    high_52w = float(history["High"].astype(float).iloc[-252:].max())
    price_vs_52w_high = (close / high_52w - 1) * 100 if high_52w > 0 else None
    atr = safe_float(history.iloc[-1].get("atr"))
    atr_pct = atr / close * 100 if atr is not None and close > 0 else None

    regime = detect_market_regime(history)
    if regime.get("error"):
        regime_snapshot = {
            "label": None,
            "confidence": None,
            "probabilities": None,
            "error": regime["error"],
            "model": regime.get("model"),
        }
    else:
        regime_snapshot = {
            "label": regime.get("regime"),
            "confidence": regime.get("confidence"),
            "probabilities": regime.get("regime_probabilities"),
            "state_means_pct": regime.get("state_means_pct"),
            "state_vols_pct": regime.get("state_vols_pct"),
            "model": regime.get("model"),
            "error": None,
        }

    return {
        "context_version": SIGNAL_CONTEXT_VERSION,
        "context_source": context_source,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "data_through": str(cutoff.date()),
        "daily_bars": int(len(history)),
        "regime": regime_snapshot,
        "returns_5d": round(returns_5d, 3) if returns_5d is not None else None,
        "returns_20d": round(returns_20d, 3) if returns_20d is not None else None,
        "volatility_20d": round(volatility_20d, 3) if volatility_20d is not None else None,
        "atr_pct": round(atr_pct, 3) if atr_pct is not None else None,
        "price_vs_52w_high": round(price_vs_52w_high, 3) if price_vs_52w_high is not None else None,
        "weekly_bias": (
            "bullish" if weekly_bullish is True
            else "not_bullish" if weekly_bullish is False
            else "unavailable"
        ),
        "trend": details.get("trend"),
        "zscore": round(float(zscore), 4),
        "conditions": {
            "c1_ema_kijun_touch": bool(details.get("ema_kijun_touch")),
            "c2_rsi_pullback": bool(details.get("rsi_pullback")),
            "c3_bull_volume": bool(details.get("bull_volume")),
            "c4_zscore_dip": bool(details.get("zscore_dip")),
        },
    }


def _normalize_ts(value):
    ts = pd.Timestamp(value)
    if ts.tzinfo is not None:
        ts = ts.tz_localize(None)
    return ts.normalize()


def _weekly_bullish_asof(raw_w: pd.DataFrame, signal_date) -> bool | None:
    """Reprodukuje weekly bias k dátumu signálu — rovnaká logika ako live
    (Donchian 20w + SMA50w + EMA10/20), ale na weekly dátach orezaných po
    signal_date (žiadny look-ahead)."""
    if raw_w is None or raw_w.empty:
        return None
    cutoff = _normalize_ts(signal_date)
    hist = raw_w.loc[[_normalize_ts(ix) <= cutoff for ix in raw_w.index]]
    if len(hist) < 20:
        return None
    df_w = add_indicators(hist)
    trend = _weekly_trend(df_w)
    return _weekly_trend_to_bullish(trend)


def _backfill_ticker_context(ticker: str, entries: dict, force: bool = False) -> dict:
    """Doplní `context` do historických signálov tickera, ktorým chýba. Reuse
    build_signal_context (single source of truth), zscore a weekly_bullish sa
    dopočítajú z dát orezaných po dátume signálu. Mutuje `entries` in-place."""
    todo = [d for d, e in entries.items()
            if isinstance(e, dict) and (force or not e.get("context"))]
    if not todo:
        return {"backfilled": 0, "skipped": len(entries), "errors": 0}
    try:
        raw_d = _yf_download_cached(ticker, "2y", "1d")
        if len(raw_d) < 30:
            return {"backfilled": 0, "skipped": 0, "errors": len(todo), "reason": "málo daily dát"}
        df_d = add_indicators(raw_d)
        zs = rolling_zscore(df_d["Close"])
        raw_w = None
        try:
            raw_w = _yf_download_cached(ticker, "5y", "1wk")
        except Exception:
            pass
    except Exception as e:
        return {"backfilled": 0, "skipped": 0, "errors": len(todo), "reason": str(e)}
    done = err = 0
    for date_key in todo:
        try:
            sig_date = _normalize_ts(date_key)
            details = entries[date_key].get("details") or {}
            zsub = zs.loc[[_normalize_ts(ix) <= sig_date for ix in zs.index]]
            zval = float(zsub.iloc[-1]) if len(zsub) else 0.0
            wb = _weekly_bullish_asof(raw_w, sig_date)
            ctx = build_signal_context(df_d, sig_date, details, zval, wb, context_source="backfill")
            if ctx:
                entries[date_key]["context"] = ctx
                done += 1
            else:
                err += 1
        except Exception:
            err += 1
    del raw_d, df_d, zs, raw_w
    gc.collect()
    return {"backfilled": done, "skipped": len(entries) - len(todo), "errors": err}


def build_signal_outcome_analytics(df_daily: pd.DataFrame, signals: list[dict]) -> tuple[list[dict], dict, dict]:
    """Pridá k signálom forward výsledky po 30/60/90 obchodných sviečkach.

    Výpočet je čisto analytický a nemení signal score ani tier. MFE/MAE sa merajú
    od close signálnej sviečky po high/low nasledujúcich sviečok.
    """
    if df_daily is None or df_daily.empty:
        return signals, {}, {}

    frame = df_daily.copy()
    today = pd.Timestamp.now("UTC").tz_localize(None).normalize()
    frame = frame.loc[
        [
            (pd.Timestamp(value).tz_localize(None) if pd.Timestamp(value).tzinfo else pd.Timestamp(value)).normalize() < today
            for value in frame.index
        ]
    ]
    if frame.empty:
        return signals, {}, {}
    normalized_dates = []
    for value in frame.index:
        ts = pd.Timestamp(value)
        if ts.tzinfo is not None:
            ts = ts.tz_localize(None)
        normalized_dates.append(ts.normalize())
    date_to_index = {str(ts.date()): idx for idx, ts in enumerate(normalized_dates)}

    enriched = []
    for signal in signals:
        item = dict(signal)
        signal_date = str(pd.Timestamp(int(signal["time"]), unit="s").date())
        idx = date_to_index.get(signal_date)
        entry = safe_float(signal.get("close"))
        outcomes = {}

        for horizon in SIGNAL_OUTCOME_HORIZONS:
            result = {
                "horizon": horizon,
                "status": "unavailable",
                "days_available": 0,
            }
            if idx is None or entry is None or entry <= 0:
                outcomes[str(horizon)] = result
                continue

            available = max(0, len(frame) - idx - 1)
            result["days_available"] = min(available, horizon)
            if available < horizon:
                result["status"] = "pending"
                outcomes[str(horizon)] = result
                continue

            future = frame.iloc[idx + 1:idx + horizon + 1]
            end_close = safe_float(future.iloc[-1].get("Close"))
            max_high = safe_float(future["High"].max())
            min_low = safe_float(future["Low"].min())
            if None in (end_close, max_high, min_low):
                outcomes[str(horizon)] = result
                continue

            return_pct = (end_close - entry) / entry * 100
            mfe_pct = (max_high - entry) / entry * 100
            mae_pct = (min_low - entry) / entry * 100
            high_pos = int(np.argmax(future["High"].to_numpy(dtype=float))) + 1
            low_pos = int(np.argmin(future["Low"].to_numpy(dtype=float))) + 1
            # rules v2: win/loss prah škáluje s ATR% pri dátume signálu (clamp 1–3 %)
            move_threshold = SIGNAL_OUTCOME_MOVE_THRESHOLD
            if "atr" in frame.columns:
                atr_at_signal = safe_float(frame.iloc[idx].get("atr"))
                if atr_at_signal and atr_at_signal > 0:
                    move_threshold = min(SIGNAL_OUTCOME_MAX, max(SIGNAL_OUTCOME_MIN, SIGNAL_OUTCOME_ATR_MULT * atr_at_signal / entry * 100))
            outcome = (
                "win" if return_pct >= move_threshold
                else "loss" if return_pct <= -move_threshold
                else "flat"
            )
            outcomes[str(horizon)] = {
                "horizon": horizon,
                "status": "complete",
                "outcome": outcome,
                "move_threshold_pct": round(move_threshold, 2),
                "return_pct": round(return_pct, 2),
                "mfe_pct": round(mfe_pct, 2),
                "mae_pct": round(mae_pct, 2),
                "days_to_mfe": high_pos,
                "days_to_mae": low_pos,
                "days_available": horizon,
            }

        item["outcomes"] = outcomes
        enriched.append(item)

    def summarize(rows: list[dict], horizon: int) -> dict:
        key = str(horizon)
        completed = [
            signal["outcomes"][key]
            for signal in rows
            if signal.get("outcomes", {}).get(key, {}).get("status") == "complete"
        ]
        pending = sum(
            1 for signal in rows
            if signal.get("outcomes", {}).get(key, {}).get("status") == "pending"
        )
        returns = [row["return_pct"] for row in completed]
        mfes = [row["mfe_pct"] for row in completed]
        maes = [row["mae_pct"] for row in completed]
        wins = sum(row.get("outcome") == "win" for row in completed)
        losses = sum(row.get("outcome") == "loss" for row in completed)
        flats = sum(row.get("outcome") == "flat" for row in completed)
        return {
            "horizon": horizon,
            "completed": len(completed),
            "pending": pending,
            "unavailable": len(rows) - len(completed) - pending,
            "wins": wins,
            "losses": losses,
            "flats": flats,
            "win_rate": round(wins / len(completed) * 100, 1) if completed else None,
            "avg_return_pct": round(float(np.mean(returns)), 2) if returns else None,
            "median_return_pct": round(float(np.median(returns)), 2) if returns else None,
            "avg_mfe_pct": round(float(np.mean(mfes)), 2) if mfes else None,
            "avg_mae_pct": round(float(np.mean(maes)), 2) if maes else None,
        }

    summary = {}
    for horizon in SIGNAL_OUTCOME_HORIZONS:
        summary[str(horizon)] = summarize(enriched, horizon)

    segment_groups = {
        "tier": [
            ("buy", "Buy", [signal for signal in enriched if signal.get("tier") == "buy"]),
            ("watch", "Watch", [signal for signal in enriched if signal.get("tier") == "watch"]),
            ("counter", "Counter", [signal for signal in enriched if signal.get("tier") == "counter"]),
        ],
        "score": [
            (str(score), f"{score}/4", [signal for signal in enriched if int(signal.get("score", 0)) == score])
            for score in (2, 3, 4)
        ],
        # Per-regime — vyžaduje backfillnutý kontext (context.regime.label) na signáloch.
        # Bez kontextu zostane skupina prázdna (degraduje pôvabne).
        "regime": [
            ("bull", "Bull", [s for s in enriched if s.get("regime") == "bull"]),
            ("sideways", "Sideways", [s for s in enriched if s.get("regime") == "sideways"]),
            ("bear", "Bear", [s for s in enriched if s.get("regime") == "bear"]),
            ("high_volatility", "Vysoká vol.", [s for s in enriched if s.get("regime") == "high_volatility"]),
        ],
    }
    segments = {}
    for group_name, groups in segment_groups.items():
        segments[group_name] = {}
        for horizon in SIGNAL_OUTCOME_HORIZONS:
            rows = []
            for key, label, group_signals in groups:
                row = summarize(group_signals, horizon)
                row.update({"key": key, "label": label, "total": len(group_signals)})
                rows.append(row)
            segments[group_name][str(horizon)] = rows

    return enriched, summary, segments


SCANNER_CACHE_FILE = DATA_ROOT / "market_scanner_cache.json"
DIP_SCORES_FILE = DATA_ROOT / "dip_scores.json"
_scanner_lock = threading.Lock()
_scanner_state = {
    "running": False,
    "started_at": None,
    "finished_at": None,
    "progress": 0,
    "total": 0,
    "current": None,
    "error": None,
    "memory_start_mb": None,
    "memory_peak_mb": None,
}


def load_scanner_cache():
    if SCANNER_CACHE_FILE.exists():
        try:
            return json.loads(SCANNER_CACHE_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_scanner_cache(data):
    SCANNER_CACHE_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_dip_scores():
    if DIP_SCORES_FILE.exists():
        try:
            return json.loads(DIP_SCORES_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_dip_scores(data):
    DIP_SCORES_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _num_or_none(value):
    try:
        if value is None or value == "":
            return None
        v = float(value)
        if math.isnan(v) or math.isinf(v):
            return None
        return int(v) if abs(v - int(v)) < 1e-9 else v
    except Exception:
        return None


def _dip_label(total):
    total = _num_or_none(total)
    if total is None:
        return "TECH ONLY"
    if total >= DIP_VERY_STRONG_THRESHOLD:
        return "VERY STRONG"
    if total >= DIP_STRONG_THRESHOLD:
        return "STRONG"
    if total >= 80:
        return "WATCH"
    return "WEAK DIP"


def enrich_with_dip(row: dict, dip_scores: dict) -> dict:
    out = dict(row)
    ticker = str(out.get("ticker") or "").upper()
    dip = dip_scores.get(ticker)
    if dip:
        out["dip"] = dip
        out["dip_total"] = dip.get("total")
        out["dip_rank"] = dip.get("rank")
        out["dip_label"] = _dip_label(dip.get("total"))
    else:
        out["dip"] = None
        out["dip_total"] = None
        out["dip_rank"] = None
        out["dip_label"] = "TECH ONLY"
    return out


def enrich_scanner_payload(payload: dict) -> dict:
    if not isinstance(payload, dict):
        return {}
    dip_scores = load_dip_scores()
    out = dict(payload)
    rows = out.get("results") or []
    out["dip_import"] = {
        "count": len(dip_scores),
        "updated_at": dip_scores.get("_meta", {}).get("updated_at") if isinstance(dip_scores.get("_meta"), dict) else None,
    }
    clean_scores = {k: v for k, v in dip_scores.items() if not k.startswith("_")}
    out["dip_import"]["count"] = len(clean_scores)
    out["results"] = [enrich_with_dip(r, clean_scores) for r in rows]
    massive = _massive_market_context(refresh=False)
    massive_tickers = {}
    massive_tickers.update((((massive or {}).get("sp500") or {}).get("tickers") or {}))
    massive_tickers.update((((massive or {}).get("nasdaq100") or {}).get("tickers") or {}))
    for row in out["results"]:
        row["market_day"] = massive_tickers.get(str(row.get("ticker") or "").upper())
    out["massive_market"] = massive
    out["crossover_matches"] = sum(1 for r in out["results"] if _num_or_none(r.get("dip_total")) is not None and r.get("dip_total") >= DIP_STRONG_THRESHOLD)
    return out


def _include_scanner_result(row: dict, dip_scores: dict) -> bool:
    """Keep live signals plus imported DIP titles worth manual chart review."""
    if row.get("recent_signal"):
        return True
    ticker = str(row.get("ticker") or "").upper()
    dip_total = _num_or_none((dip_scores.get(ticker) or {}).get("total"))
    return dip_total is not None and dip_total >= DIP_SCANNER_VISIBILITY_THRESHOLD


def _scanner_result_sort_key(row: dict) -> tuple:
    """Sort signal and high-DIP-only rows safely; recent_signal can be None."""
    return (
        _num_or_none(row.get("dip_total")) or -1,
        row.get("setup_score") or 0,
        (row.get("recent_signal") or {}).get("date", ""),
    )


def scanner_universe_from_dip() -> tuple[list[str], str, str]:
    """Scanner universe = imported DIP ranking, with Nasdaq-100 fallback.

    When a DIP Excel is imported, that file is the user's curated universe and
    can include NYSE/non-Nasdaq stocks. Scanning Nasdaq-100 on top of it doubles
    request volume, hides the imported names in timeout noise, and is wasteful on
    the Render free/low-memory tier. Without an import we keep the old Nasdaq-100
    behavior as a sensible fallback.
    """
    dip_scores = load_dip_scores()
    imported = [
        str(sym or "").strip().upper()
        for sym in dip_scores.keys()
        if sym and not str(sym).startswith("_")
    ]
    imported = [sym for sym in imported if re.fullmatch(r"[A-Z0-9.\-]{1,12}", sym)]

    if imported:
        out: list[str] = []
        seen = set()
        # Preserve ranking order from the imported file. Dict order is insertion
        # order; parse_dip_ranking_xlsx reads rows in Ranking order.
        for sym in imported[:max(0, SCANNER_DIP_UNIVERSE_MAX)]:
            if sym not in seen:
                out.append(sym)
                seen.add(sym)
        return out, "DIP import", "dip_import"

    return NASDAQ100_TICKERS[:], "Nasdaq-100 fallback", "nasdaq100"


def parse_dip_ranking_xlsx(raw: bytes, filename: str | None = None) -> dict:
    try:
        import openpyxl
    except Exception as e:
        raise HTTPException(500, f"openpyxl chyba/import dependency: {e}")

    wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
    sheet_name = "Ranking" if "Ranking" in wb.sheetnames else ("ranking" if "ranking" in wb.sheetnames else wb.sheetnames[0])
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Excel neobsahuje data")
    headers = [str(h or "").strip().lower() for h in rows[0]]

    def col(*names):
        for name in names:
            if name.lower() in headers:
                return headers.index(name.lower())
        return None

    idx_rank = col("rank")
    idx_ticker = col("ticker")
    idx_company = col("company")
    idx_price = col("price")
    idx_fa = col("fa")
    idx_ta = col("ta")
    idx_total = col("total")
    if idx_ticker is None or idx_total is None:
        raise HTTPException(400, "Zalozka Ranking musi obsahovat aspon stlpce Ticker a TOTAL")

    scores = {}
    for row in rows[1:]:
        ticker = str(row[idx_ticker] or "").strip().upper() if idx_ticker < len(row) else ""
        if not ticker:
            continue
        total = _num_or_none(row[idx_total] if idx_total < len(row) else None)
        scores[ticker] = {
            "rank": _num_or_none(row[idx_rank] if idx_rank is not None and idx_rank < len(row) else None),
            "ticker": ticker,
            "company": str(row[idx_company] or "").strip() if idx_company is not None and idx_company < len(row) and row[idx_company] is not None else "",
            "price": _num_or_none(row[idx_price] if idx_price is not None and idx_price < len(row) else None),
            "fa": _num_or_none(row[idx_fa] if idx_fa is not None and idx_fa < len(row) else None),
            "ta": _num_or_none(row[idx_ta] if idx_ta is not None and idx_ta < len(row) else None),
            "total": total,
            "label": _dip_label(total),
        }
    scores["_meta"] = {
        "sheet": sheet_name,
        "filename": filename or "",
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "count": len([k for k in scores.keys() if not k.startswith("_")]),
    }
    return scores


@app.post("/api/scanner/dip/import")
async def import_dip_scores(request: Request, filename: str | None = Query(None)):
    raw = await request.body()
    if not raw:
        raise HTTPException(400, "Chyba importny subor")
    scores = parse_dip_ranking_xlsx(raw, filename=filename)
    save_dip_scores(scores)
    meta = scores.get("_meta", {})
    return {"ok": True, "count": meta.get("count", 0), "sheet": meta.get("sheet"), "filename": meta.get("filename"), "updated_at": meta.get("updated_at")}


@app.post("/api/scanner/dip/import-html")
async def import_dip_html(request: Request):
    raise HTTPException(status_code=410, detail="Finviz HTML import je vypnuty; pouzi XLSX import DIP rankingu.")


@app.post("/api/public/finviz/import-html")
async def public_import_finviz_html(request: Request):
    """Legacy endpoint. HTML/Finviz import je vypnutý; DIP ranking sa importuje cez XLSX."""
    raise HTTPException(status_code=410, detail="Finviz public HTML import je vypnuty; pouzi XLSX import DIP rankingu.")


@app.get("/api/scanner/finviz/screeners")
def get_finviz_screeners():
    raise HTTPException(status_code=410, detail="Finviz screener URL zoznam je vypnuty.")


@app.post("/api/scanner/finviz/screeners")
async def save_finviz_screeners(request: Request):
    raise HTTPException(status_code=410, detail="Finviz screener URL zoznam je vypnuty.")


@app.get("/api/public/finviz/screeners")
def public_finviz_screeners(request: Request):
    """Legacy endpoint. Finviz screener URL list je vypnutý spolu s HTML importom."""
    raise HTTPException(status_code=410, detail="Finviz public screener URL zoznam je vypnuty.")


@app.get("/api/scanner/dip/html-preview")
def get_dip_html_preview():
    raise HTTPException(status_code=410, detail="Finviz HTML preview je vypnuty; pouzi XLSX import DIP rankingu.")


@app.get("/api/scanner/dip/status")
def get_dip_status():
    scores = load_dip_scores()
    meta = scores.get("_meta", {}) if isinstance(scores.get("_meta"), dict) else {}
    return {"count": meta.get("count", len([k for k in scores if not k.startswith("_")])), "updated_at": meta.get("updated_at"), "sheet": meta.get("sheet"), "filename": meta.get("filename")}


# In-process yfinance download cache — scanner aj /api/checklist ťahajú tie isté
# daily/weekly dáta opakovane. TTL 30 min, capped (Render free tier RAM).
_YF_CACHE: dict = {}
_YF_CACHE_LOCK = threading.Lock()
_YF_CACHE_TTL = 1800   # s
_YF_CACHE_MAX = 30 if _LOW_MEMORY else 60     # entries; low profile keeps RAM headroom

# ── Massive denné/týždenné bary (Polygon-style) — yfinance alternatíva ─────────
# yfinance je na free tieri najkrehkejší zdroj (rate-limit, timeouty → prázdne
# grafy a "chyby" v scanneri). Ak je MASSIVE_API_KEY nastavený a free plán
# podporuje per-ticker agregáty, ťaháme denné/týždenné bary odtiaľ a yfinance
# necháme len ako fallback. Intraday (1h/4h) ostáva na yfinance — free Massive
# plán ho nemá. Chart endpoint (get_ohlcv) má vlastnú cestu, tu sa nemení.
_MASSIVE_BARS_TIMESPAN = {"1d": "day", "1wk": "week"}
_MASSIVE_PERIOD_DAYS = {
    "1mo": 31, "3mo": 95, "6mo": 190, "1y": 370, "2y": 735, "5y": 1830, "max": 3650,
}
_massive_bars_disabled = False  # po prvom NOT_AUTHORIZED prestaneme skúšať (free plán)


def _massive_daily_bars(ticker: str, period: str, interval: str):
    """OHLCV DataFrame v yfinance tvare (Open/High/Low/Close/Volume, DatetimeIndex)
    z Massive per-ticker agregátov, alebo None (fail-soft)."""
    global _massive_bars_disabled
    if _massive_bars_disabled:
        return None
    timespan = _MASSIVE_BARS_TIMESPAN.get(interval)
    if not timespan:
        return None
    api_key = os.getenv("MASSIVE_API_KEY", "").strip()
    if not api_key:
        return None
    days = _MASSIVE_PERIOD_DAYS.get(period, 400)
    end = datetime.now(timezone.utc).date()
    start = end - timedelta(days=days)
    try:
        r = requests.get(
            f"https://api.massive.com/v2/aggs/ticker/{urllib.parse.quote(ticker)}"
            f"/range/1/{timespan}/{start.isoformat()}/{end.isoformat()}",
            params={"adjusted": "true", "sort": "asc", "limit": "5000", "apiKey": api_key},
            timeout=SCANNER_YF_TIMEOUT,
        )
        payload = r.json() if r.content else {}
        if payload.get("status") == "NOT_AUTHORIZED" or r.status_code in (401, 403):
            _massive_bars_disabled = True
            print("[massive] per-ticker bary nie sú na tomto pláne autorizované — "
                  "vypínam, používam yfinance")
            return None
        if not r.ok:
            return None
        rows = payload.get("results") or []
        if not rows:
            return None
        df = pd.DataFrame([
            {"Open": x.get("o"), "High": x.get("h"), "Low": x.get("l"),
             "Close": x.get("c"), "Volume": x.get("v"), "__t": x.get("t")}
            for x in rows if x.get("t") is not None
        ])
        if df.empty:
            return None
        df.index = pd.to_datetime(df.pop("__t"), unit="ms")
        df.index.name = "Date"
        df = df.dropna(subset=["Open", "High", "Low", "Close"])
        return df if len(df) >= 2 else None
    except Exception as exc:
        print(f"[massive] bary {ticker} {period}/{interval}: {type(exc).__name__}")
        return None


def _yf_download_cached(ticker: str, period: str, interval: str, prefer_massive: bool = True,
                        munge_dots: bool = True, retain_in_memory: bool = True):
    """munge_dots: BRK.B -> BRK-B pre yfinance (US class shares). get_chart volá
    s False, lebo jeho tickery môžu byť burzové suffixy (VWRD.L, CNDX.L), kde
    bodka MUSÍ ostať — pomlčková forma by vrátila prázdno."""
    key = (ticker, period, interval)
    now = time.time()
    with _YF_CACHE_LOCK:
        hit = _YF_CACHE.get(key)
        if hit and now - hit[0] < _YF_CACHE_TTL:
            return hit[1].copy()
    raw = _massive_daily_bars(ticker, period, interval) if prefer_massive else None
    if raw is None or len(raw) < 2:
        yf_symbol = ticker.replace(".", "-") if munge_dots else ticker
        raw = yf.download(yf_symbol, period=period, interval=interval, auto_adjust=True, progress=False, timeout=SCANNER_YF_TIMEOUT)
        if isinstance(raw.columns, pd.MultiIndex):
            raw.columns = raw.columns.get_level_values(0)
        raw = raw.dropna(subset=["Open", "High", "Low", "Close"])
    if retain_in_memory:
        with _YF_CACHE_LOCK:
            if len(_YF_CACHE) >= _YF_CACHE_MAX:
                oldest = sorted(_YF_CACHE.items(), key=lambda kv: kv[1][0])[: _YF_CACHE_MAX // 3]
                for k, _ in oldest:
                    _YF_CACHE.pop(k, None)
            _YF_CACHE[key] = (now, raw)
    return raw.copy()


def _scan_buy_signal_for_ticker(ticker: str, days: int, ticker_slog: dict | None = None) -> dict:
    raw_d = _yf_download_cached(ticker, "6mo", "1d", prefer_massive=False, retain_in_memory=False)
    if len(raw_d) < 20:
        return {"ticker": ticker, "error": "Nedostatok dat"}

    df_d = add_indicators(raw_d)
    daily_health = chart_health(df_d, "daily")

    raw_w = _yf_download_cached(ticker, "1y", "1wk", prefer_massive=False, retain_in_memory=False)
    weekly_bullish = None
    weekly_status = "insufficient_history"
    weekly_trend = None
    weekly_health = {"status": "unknown", "label": "N/A", "score": 0, "reasons": ["málo dát"]}
    if len(raw_w) >= 20:
        df_w = add_indicators(raw_w)
        weekly_trend = _weekly_trend(df_w)
        weekly_health = chart_health(df_w, "weekly")
        weekly_bullish = _weekly_trend_to_bullish(weekly_trend)
        if weekly_trend is not None:
            weekly_status = weekly_trend.get("key") or ("bullish" if weekly_bullish else "not_bullish")

    today_date = pd.Timestamp.now("UTC").tz_localize(None).normalize().tz_localize(None)
    latest_closed_date = latest_closed_daily_date(df_d)
    cutoff = today_date - pd.Timedelta(days=days)
    ticker_slog = dict(ticker_slog or {})
    recent_signal = None
    all_signals = []

    # Rolling 60-period z-score: cena vs vlastný 60p režim — jeden zdroj pravdy
    _zscore_series = rolling_zscore(df_d["Close"])
    zscore_values = _zscore_series.values  # zarovnané s df_d
    latest_zscore = float(_zscore_series.iloc[-1]) if len(_zscore_series) else 0.0

    df_score = df_d.reset_index()
    for i in range(5, len(df_score)):
        row = df_score.iloc[i]
        row_date = pd.Timestamp(row.iloc[0])
        if row_date.tzinfo:
            row_date = row_date.tz_localize(None)
        if row_date.date() >= today_date.date():
            continue

        date_key = str(row_date.date())
        close = float(row["Close"])
        zscore = float(zscore_values[i]) if i < len(zscore_values) else 0.0
        sc, details = score_signal_day(row, zscore)

        if sc >= 2:
            tier = signal_tier(sc, details["trend"])
            if date_key not in ticker_slog:
                entry = {
                    "score": sc,
                    "tier": tier,
                    "close": round(close, 2),
                    "details": details,
                    "rules_version": SIGNAL_RULES_VERSION,
                }
                if latest_closed_date is not None and row_date.normalize() == latest_closed_date:
                    entry["context"] = build_signal_context(
                        df_d, row_date, details, zscore, weekly_bullish
                    )
                ticker_slog[date_key] = entry
            sig = {"date": date_key, "score": sc, "tier": tier,
                   "close": round(close, 2), "zscore_dip": details["zscore_dip"]}
            all_signals.append(sig)
            if row_date.date() >= cutoff.date():
                if recent_signal is None or date_key > recent_signal["date"]:
                    recent_signal = sig

    setup = build_setup_assessment(df_d, bool(weekly_bullish), recent_signal, len(all_signals), latest_zscore)
    last_close = round(float(df_d.iloc[-1]["Close"]), 2)
    del raw_d, df_d, df_score, _zscore_series, zscore_values
    try:
        del raw_w, df_w
    except NameError:
        pass
    return {
        "ticker": ticker,
        "weekly_bullish": weekly_bullish,
        "weekly_status": weekly_status,
        "weekly_trend": weekly_trend,
        "chart_health": {
            "daily": daily_health,
            "weekly": weekly_health,
        },
        "recent_signal": recent_signal,
        "signal_count": len(all_signals),
        "last_close": last_close,
        "slog_update": ticker_slog,
        **setup,
        "error": None,
    }


def _run_nasdaq_scanner(days: int):
    started = datetime.now(timezone.utc).isoformat()
    results, errors = [], []
    slog_source = load_signals_log()
    slog_work = {k: dict(v) if isinstance(v, dict) else v for k, v in slog_source.items()}
    tickers, universe_label, universe_key = scanner_universe_from_dip()
    dip_scores = {k: v for k, v in load_dip_scores().items() if not k.startswith("_")}
    scanner_memory_start = _process_memory_mb().get("rss_mb")

    with _scanner_lock:
        _scanner_state.update({
            "running": True,
            "started_at": started,
            "finished_at": None,
            "progress": 0,
            "total": len(tickers),
            "current": None,
            "error": None,
            "memory_start_mb": scanner_memory_start,
            "memory_peak_mb": scanner_memory_start,
        })

    try:
        done_count = 0
        max_workers = max(1, min(SCANNER_MAX_WORKERS, len(tickers)))
        pool = ThreadPoolExecutor(max_workers=max_workers)
        futures = {}
        ticker_iter = iter(tickers)

        def submit_next():
            try:
                ticker = next(ticker_iter)
            except StopIteration:
                return None
            fut = pool.submit(_scan_buy_signal_for_ticker, ticker, days, dict(slog_work.get(ticker, {})))
            futures[fut] = ticker
            return fut

        try:
            for _ in range(max_workers):
                submit_next()

            pending = set(futures.keys())
            while pending:
                done, pending = wait(pending, timeout=1.0, return_when=FIRST_COMPLETED)

                for future in done:
                    ticker = futures[future]
                    with _scanner_lock:
                        _scanner_state.update({"progress": done_count, "current": ticker})
                    try:
                        row = future.result()
                        slog_update = row.pop("slog_update", None)
                        if isinstance(slog_update, dict) and slog_update:
                            slog_work[ticker] = slog_update
                        if row.get("error"):
                            errors.append(row)
                        elif _include_scanner_result(row, dip_scores):
                            results.append(row)
                    except Exception as e:
                        errors.append({"ticker": ticker, "error": str(e)[:80]})
                    done_count += 1
                    if SCANNER_GC_INTERVAL > 0 and done_count % SCANNER_GC_INTERVAL == 0:
                        gc.collect()
                    scanner_memory_peak = _process_memory_mb().get("rss_mb")
                    with _scanner_lock:
                        _scanner_state.update({
                            "progress": done_count,
                            "current": ticker,
                            "memory_peak_mb": max(_scanner_state.get("memory_peak_mb") or 0, scanner_memory_peak or 0) or None,
                        })
                    newest = submit_next()
                    if newest is not None:
                        pending.add(newest)
        finally:
            pool.shutdown(wait=False, cancel_futures=True)

        save_signals_log(slog_work)
        del slog_work, slog_source
        gc.collect()
        results = [enrich_with_dip(r, dip_scores) for r in results]
        results.sort(key=_scanner_result_sort_key, reverse=True)
        error_counts = {}
        for err in errors:
            reason = str(err.get("error") or "Unknown")
            error_counts[reason] = error_counts.get(reason, 0) + 1
        payload = {
            "universe": universe_key,
            "universe_label": universe_label,
            "days": days,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "total": len(tickers),
            "matches": sum(1 for row in results if row.get("recent_signal")),
            "high_dip_rows": sum(1 for row in results if _num_or_none(row.get("dip_total")) is not None and row.get("dip_total") >= DIP_SCANNER_VISIBILITY_THRESHOLD),
            "displayed_rows": len(results),
            "errors": len(errors),
            "error_counts": error_counts,
            "error_samples": errors[:25],
            "crossover_matches": sum(
                1 for r in results
                if r.get("recent_signal") and _num_or_none(r.get("dip_total")) is not None
                and r.get("dip_total") >= DIP_STRONG_THRESHOLD
            ),
            "results": results,
        }
        save_scanner_cache(payload)
        with _scanner_lock:
            _scanner_state.update({
                "running": False,
                "finished_at": payload["generated_at"],
                "progress": len(tickers),
                "current": None,
                "error": None,
            })
    except Exception as e:
        with _scanner_lock:
            _scanner_state.update({
                "running": False,
                "finished_at": datetime.now(timezone.utc).isoformat(),
                "current": None,
                "error": str(e)[:120],
            })


@app.post("/api/scanner/nasdaq/run")
def start_nasdaq_scanner(days: int = Query(SCANNER_DEFAULT_DAYS, ge=1, le=10)):
    with _scanner_lock:
        if _scanner_state.get("running"):
            return {"status": "running", "state": dict(_scanner_state), "cache": enrich_scanner_payload(load_scanner_cache())}
        tickers, _, _ = scanner_universe_from_dip()
        _scanner_state.update({"running": True, "progress": 0, "total": len(tickers), "current": None, "error": None})
    t = threading.Thread(target=_run_nasdaq_scanner, args=(days,), daemon=True)
    t.start()
    return {"status": "started", "state": dict(_scanner_state), "cache": enrich_scanner_payload(load_scanner_cache())}


@app.get("/api/scanner/nasdaq/results")
def get_nasdaq_scanner_results():
    with _scanner_lock:
        state = dict(_scanner_state)
    return {"state": state, "cache": enrich_scanner_payload(load_scanner_cache())}


def _scanner_candidate_rows(limit: int = 40) -> list[dict]:
    scanner = enrich_scanner_payload(load_scanner_cache())
    rows = list(scanner.get("results") or [])
    rows.sort(key=lambda r: (
        _num_or_none(r.get("dip_total")) or -1,
        r.get("setup_score") or 0,
        (r.get("recent_signal") or {}).get("date", ""),
    ), reverse=True)
    return rows[:limit]


def _watched_symbols_for_calendar() -> list[str]:
    symbols: set[str] = set()
    try:
        symbols.update(_get_portfolio_holdings().keys())
    except Exception:
        pass
    try:
        for item in _read_watchlist_file():
            sym = str(item.get("symbol") or "").upper()
            if sym:
                symbols.add(sym)
    except Exception:
        pass
    try:
        for row in _scanner_candidate_rows(25):
            sym = str(row.get("ticker") or "").upper()
            if sym:
                symbols.add(sym)
    except Exception:
        pass
    return sorted(symbols)


INVESTOR_INBOX_CACHE_TTL = 86400     # seconds; Scanner is a daily snapshot, manual refresh/scan bypasses cache
EARNINGS_CALENDAR_VIEW_TTL = 86400   # seconds; earnings dates are day-level, not live data
_investor_view_cache: dict[str, tuple[float, dict]] = {}
_investor_view_cache_lock = threading.Lock()


def _investor_cache_get(key: str, ttl: int, refresh: int = 0) -> dict | None:
    if refresh:
        return None
    with _investor_view_cache_lock:
        hit = _investor_view_cache.get(key)
        if not hit:
            return None
        ts, payload = hit
        if time.time() - ts <= ttl:
            out = dict(payload)
            out["cache"] = {"hit": True, "ttl_s": ttl, "age_s": round(time.time() - ts, 1)}
            return out
        _investor_view_cache.pop(key, None)
    return None


def _investor_cache_set(key: str, payload: dict) -> dict:
    with _investor_view_cache_lock:
        if len(_investor_view_cache) > 12:
            oldest = sorted(_investor_view_cache.items(), key=lambda kv: kv[1][0])[:4]
            for k, _ in oldest:
                _investor_view_cache.pop(k, None)
        _investor_view_cache[key] = (time.time(), payload)
    out = dict(payload)
    out["cache"] = {"hit": False}
    return out


@app.get("/api/earnings/calendar")
def get_earnings_calendar_view(days: int = Query(14, ge=1, le=31),
                               refresh: int = Query(0)):
    """Ľahký earnings kalendár pre relevantné tickery: portfólio + watchlist +
    poslední scanner kandidáti. Neprehľadáva celý trh a používa existujúcu cache
    / per-symbol fallback len pre malý zoznam symbolov."""
    cache_key = f"earnings_calendar:{days}:{datetime.now(timezone.utc).date().isoformat()}"
    cached = _investor_cache_get(cache_key, EARNINGS_CALENDAR_VIEW_TTL, refresh)
    if cached:
        return cached
    now = datetime.now(timezone.utc).date()
    end = now + timedelta(days=days)
    holdings = {}
    try:
        holdings = _get_portfolio_holdings()
    except Exception:
        pass
    rows = []
    for sym in _watched_symbols_for_calendar():
        try:
            raw_date = _earnings_next_date(sym)
        except Exception:
            raw_date = None
        if not raw_date:
            continue
        try:
            dt = datetime.fromisoformat(str(raw_date)[:10]).date()
        except Exception:
            continue
        if now <= dt <= end:
            h = holdings.get(sym) or {}
            rows.append({
                "ticker": sym,
                "date": dt.isoformat(),
                "days": (dt - now).days,
                "in_portfolio": sym in holdings,
                "pnl_pct": h.get("pnl_pct"),
                "pnl": h.get("pnl"),
                "amount": h.get("amount"),
            })
    rows.sort(key=lambda r: (r["date"], not r["in_portfolio"], r["ticker"]))
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "from": now.isoformat(),
        "to": end.isoformat(),
        "days": days,
        "items": rows,
        "count": len(rows),
    }
    return _investor_cache_set(cache_key, payload)


@app.get("/api/investor/inbox")
def get_investor_inbox(refresh: int = Query(0)):
    """Týždenný prehľad výnimiek pre človeka: čo si dnes/tento týždeň pozrieť.
    Používa už existujúce cache a interpretačné vrstvy; nespúšťa nový scan."""
    cache_key = f"investor_inbox:{datetime.now(timezone.utc).date().isoformat()}"
    cached = _investor_cache_get(cache_key, INVESTOR_INBOX_CACHE_TTL, refresh)
    if cached:
        return cached
    now = datetime.now(timezone.utc)
    holdings = {}
    try:
        holdings = _get_portfolio_holdings()
    except Exception:
        pass
    events: list[dict] = []

    def add(kind: str, ticker: str, title: str, detail: str, severity: str = "watch",
            priority: int = 50, summary: str | None = None, **extra):
        ticker = str(ticker or "").upper()
        if not ticker:
            return
        events.append({
            "id": f"{kind}:{ticker}:{now.date().isoformat()}",
            "kind": kind,
            "ticker": ticker,
            "title": title,
            "detail": detail,
            "summary": summary or detail,
            "severity": severity,
            "priority": priority,
            "time": now.isoformat(),
            **extra,
        })

    # DCA / thesis risk: stojace stavy nad existujúcim DCA endpointom.
    # Prahy z /api/settings — rovnaké ako DCA karta v Portfóliu (jeden zdroj pravdy).
    _dca_s = _dash_settings()
    for account in ("1", "2"):
        try:
            dca = get_dca_candidates(
                account=account, loss_pct=_dca_s["dca_loss_pct"],
                dip_min=_dca_s["dca_dip_min"], max_weight=_dca_s["dca_max_weight"], refresh=0
            )
            for row in dca.get("candidates", [])[:20]:
                sym = row.get("symbol")
                flag = row.get("flag")
                if flag == "dca":
                    add("dca", sym, "DCA kandidát",
                        f"Účet {account}: pozícia je {row.get('pnl_pct')} %, DIP {row.get('dip_total')} a váha {row.get('weight_pct')} % equity.",
                        "buy", 10,
                        summary=(f"{sym} je v strate {row.get('pnl_pct')} %, DIP ranking ostáva silný "
                                 f"({row.get('dip_total')}) a váha pozície je {row.get('weight_pct')} % equity. "
                                 "Stojí za kontrolu DCA."),
                        source="portfolio", account=account)
                elif flag in {"value_trap", "no_data"}:
                    label = "Slabý DIP" if flag == "value_trap" else "Mimo DIP dát"
                    add("broken", sym, label,
                        f"Účet {account}: pozícia je {row.get('pnl_pct')} %, ale DIP filter nepotvrdzuje dokup.",
                        "counter", 35,
                        summary=(f"{sym} je síce v strate {row.get('pnl_pct')} %, ale DIP filter zatiaľ "
                                 "nepotvrdzuje kvalitné dokúpenie. Skôr manuálne overiť než automaticky DCA."),
                        source="portfolio", account=account)
        except Exception:
            pass

    # Profit-taking: veľké otvorené zisky sú opačná výnimka.
    for sym, h in holdings.items():
        pnl_pct = _num_or_none(h.get("pnl_pct"))
        if pnl_pct is not None and pnl_pct >= 150:
            add("profit", sym, "Profit-taking kontrola",
                f"Otvorený zisk je približne {pnl_pct:+.1f} %. Zváž, či nechceš aspoň skontrolovať graf a plán.",
                "buy", 20,
                summary=(f"{sym} má otvorený zisk približne {pnl_pct:+.1f} %. Nie je to predajný signál, "
                         "ale zaslúži si kontrolu plánu a prípadné uzamknutie časti zisku."),
                source="portfolio", pnl_pct=pnl_pct)

    # Earnings najbližších 14 dní pre relevantné tickery.
    try:
        cal = get_earnings_calendar_view(days=14, refresh=refresh)
        for item in cal.get("items", []):
            days = item.get("days")
            sev = "counter" if item.get("in_portfolio") and days <= 1 else "watch"
            add("earnings", item.get("ticker"), "Earnings v kalendári",
                f"Earnings {item.get('date')} ({days} dní). Pri držanom titule očakávaj vyššiu volatilitu.",
                sev, 15 if item.get("in_portfolio") else 45,
                summary=(f"{item.get('ticker')} reportuje earnings {item.get('date')} "
                         f"({days} dní). Pred výsledkami môžu technické signály rýchlo stratiť platnosť."),
                source="calendar",
                date=item.get("date"), days=days, in_portfolio=item.get("in_portfolio"))
    except Exception:
        pass

    # Scanner: silné nové príležitosti mimo portfólia + zlé grafy pri držaných tituloch.
    try:
        for row in _scanner_candidate_rows(30):
            sym = str(row.get("ticker") or "").upper()
            sig = row.get("recent_signal") or {}
            in_port = sym in holdings
            dip_total = _num_or_none(row.get("dip_total"))
            health = row.get("chart_health") or {}
            daily_status = ((health.get("daily") or {}).get("status") or "").lower()
            weekly_status = ((health.get("weekly") or {}).get("status") or "").lower()
            if in_port and ("bad" in {daily_status, weekly_status}):
                add("broken", sym, "Graf potrebuje kontrolu",
                    "Držaný titul má chart health Bad na daily alebo weekly grafe.",
                    "counter", 25,
                    summary=(f"{sym} držíš v portfóliu, ale daily alebo weekly graf vyzerá poškodený. "
                             "Najprv over, či nejde o zmenu trendu alebo value trap."),
                    source="scanner")
            elif not in_port and sig and (dip_total is not None and dip_total >= DIP_STRONG_THRESHOLD):
                score = sig.get("score") or row.get("setup_score")
                add("opportunity", sym, "Nová príležitosť",
                    f"Scanner našiel {score}/4 a DIP {dip_total}. Nie je v portfóliu.",
                    "watch", 40,
                    summary=(f"{sym} nie je v portfóliu, ale scanner našiel technický signál {score}/4 "
                             f"a silný DIP ranking {dip_total}. Stojí za rýchly Verdikt."),
                    source="scanner", score=score, dip_total=dip_total)
    except Exception:
        pass

    # Inbox je určený pre človeka: jeden ticker = jeden riadok, aj keď má viac
    # dôvodov. Detailné dôvody ostávajú ako badges/reasons, aby sa nestratila
    # informácia typu "DCA áno, ale graf varuje".
    per_ticker: dict[str, dict[str, dict]] = {}
    for item in events:
        ticker = item.get("ticker")
        kind = item.get("kind")
        if not ticker or not kind:
            continue
        bucket = per_ticker.setdefault(ticker, {})
        prev = bucket.get(kind)
        if prev is None or item.get("priority", 99) < prev.get("priority", 99):
            bucket[kind] = item

    kind_labels = {
        "dca": "DCA",
        "broken": "Pozor",
        "profit": "Profit",
        "earnings": "Earnings",
        "opportunity": "Nové",
    }
    positive_kinds = {"dca", "profit", "opportunity"}
    risk_kinds = {"broken"}
    watch_kinds = {"earnings"}

    def merge_ticker_items(ticker: str, items: list[dict]) -> dict:
        items = sorted(items, key=lambda x: (x.get("priority", 99), x.get("kind", "")))
        kinds = [item.get("kind") for item in items if item.get("kind")]
        has_positive = any(kind in positive_kinds for kind in kinds)
        has_risk = any(kind in risk_kinds for kind in kinds)
        has_watch = any(kind in watch_kinds for kind in kinds)
        primary = items[0]

        if has_positive and has_risk:
            severity = "mixed"
            title = "Zmiešaný signál"
            summary = (f"{ticker}: pozitívny dôvod aj varovanie naraz. "
                       "Najprv otvor Verdikt/graf, potom rieš DCA alebo vstup.")
        elif has_risk:
            severity = "counter"
            title = primary.get("title") or "Pozor"
            summary = primary.get("summary") or primary.get("detail") or ""
        elif has_positive:
            severity = "buy"
            title = primary.get("title") or "Stojí za kontrolu"
            summary = primary.get("summary") or primary.get("detail") or ""
        elif has_watch:
            severity = "watch"
            title = primary.get("title") or "Sledovať"
            summary = primary.get("summary") or primary.get("detail") or ""
        else:
            severity = primary.get("severity") or "watch"
            title = primary.get("title") or "Sledovať"
            summary = primary.get("summary") or primary.get("detail") or ""

        if len(items) > 1 and not (has_positive and has_risk):
            labels = " + ".join(kind_labels.get(kind, kind or "?") for kind in kinds)
            title = f"{title} · {labels}"
            if not summary:
                summary = f"{ticker} má viac dôvodov na kontrolu: {labels}."

        reasons = [{
            "kind": item.get("kind"),
            "label": kind_labels.get(item.get("kind"), item.get("kind")),
            "title": item.get("title"),
            "detail": item.get("detail"),
            "summary": item.get("summary"),
            "severity": item.get("severity"),
            "priority": item.get("priority"),
            "source": item.get("source"),
        } for item in items]
        detail = " | ".join(
            f"{kind_labels.get(item.get('kind'), item.get('kind'))}: {item.get('detail') or item.get('summary') or ''}"
            for item in items
        )
        return {
            "id": f"inbox:{ticker}:{now.date().isoformat()}:{'+'.join(kinds)}",
            "kind": primary.get("kind"),
            "kinds": kinds,
            "ticker": ticker,
            "title": title,
            "detail": detail,
            "summary": summary,
            "severity": severity,
            "priority": min(item.get("priority", 99) for item in items),
            "time": now.isoformat(),
            "reasons": reasons,
        }

    ordered = sorted(
        (merge_ticker_items(ticker, list(kind_map.values())) for ticker, kind_map in per_ticker.items()),
        key=lambda x: (x.get("priority", 99), x.get("ticker", "")),
    )[:30]
    counts = {}
    for kind_map in per_ticker.values():
        for kind in kind_map:
            counts[kind] = counts.get(kind, 0) + 1
    payload = {"generated_at": now.isoformat(), "counts": counts, "tickers": len(per_ticker), "items": ordered}
    return _investor_cache_set(cache_key, payload)


WEEKLY_PLAN_CACHE_TTL = 86400


@app.get("/api/investor/plan")
def get_investor_plan(refresh: int = Query(0)):
    """Týždenný plán: ľudská syntéza NAD Investor Inboxom — nie nový analytický
    engine, len prioritizácia do 5 sekcií so šablónovými vetami (žiadny LLM):
    Pozri dnes / Možný nákup / Možné DCA / Riziko / Drž bez akcie.
    Cieľ je menej mentálneho hluku, nie viac dát. Nespúšťa nový scan."""
    cache_key = f"investor_plan:{datetime.now(timezone.utc).date().isoformat()}"
    cached = _investor_cache_get(cache_key, WEEKLY_PLAN_CACHE_TTL, refresh)
    if cached:
        return cached

    inbox = get_investor_inbox(refresh=refresh)
    items = inbox.get("items") or []
    holdings: dict = {}
    try:
        holdings = _get_portfolio_holdings()
    except Exception:
        pass

    # Pozri dnes: top 3-7 podľa počtu dôvodov + závažnosti. Vety sa REUSUJÚ
    # z inbox merge (vrátane "zmiešaný signál") — jeden zdroj formulácií.
    sev_weight = {"mixed": 3.0, "counter": 2.5, "buy": 2.0, "watch": 1.0}

    def item_score(it: dict) -> float:
        return len(it.get("reasons") or []) * 2 + sev_weight.get(it.get("severity"), 1.0)

    ranked = sorted(items, key=item_score, reverse=True)
    focus = ranked[:7]
    if len(focus) > 3:
        top = item_score(focus[0])
        focus = [it for i, it in enumerate(focus) if i < 3 or item_score(it) >= top * 0.5]
    focus_rows = [{
        "ticker": it.get("ticker"),
        "summary": it.get("summary") or it.get("detail") or "",
        "kinds": it.get("kinds") or [],
        "severity": it.get("severity"),
    } for it in focus]

    inbox_tickers = {it.get("ticker") for it in items}

    # Možné DCA: kvalitný dip bez zlomeného grafu (tie ostávajú v Pozri dnes
    # so zmiešaným summary). Riziká: earnings/zlomený graf na držaných tituloch.
    dca_rows: list = []
    risk_rows: list = []
    used_tickers = {str(row.get("ticker") or "").upper() for row in focus_rows if row.get("ticker")}
    for it in items:
        kinds = set(it.get("kinds") or [])
        t = str(it.get("ticker") or "").upper()
        if not t or t in used_tickers:
            continue
        if "dca" in kinds and "broken" not in kinds:
            dca_rows.append({"ticker": t,
                             "summary": f"{t} je DCA kandidát so zdravým grafom — strata v pásme, DIP drží."})
            used_tickers.add(t)
            continue
        if kinds & {"broken", "earnings"}:
            reason = next((r for r in (it.get("reasons") or []) if r.get("kind") in ("broken", "earnings")), None)
            risk_rows.append({"ticker": t,
                              "summary": (reason or {}).get("summary") or (reason or {}).get("detail")
                                         or it.get("summary") or ""})
            used_tickers.add(t)

    # Možný nákup: PRIENIK buy signál + DIP kvalita + zdravý graf (nie len jedno
    # kritérium), len nedržané tituly. Z posledného scanner behu (cache).
    buy_rows: list = []
    try:
        dip_min = float(_dash_settings().get("dca_dip_min") or 90)
        scan = enrich_scanner_payload(load_scanner_cache()) or {}
        for row in (scan.get("results") or []):
            t = str(row.get("ticker") or "").upper()
            if not t or t in holdings or t in used_tickers:
                continue
            sig = row.get("recent_signal") or {}
            if sig.get("tier") != "buy":
                continue
            dip = row.get("dip_total")
            if dip is None or float(dip) < dip_min:
                continue
            ch = row.get("chart_health") or {}
            dstat = str((ch.get("daily") or {}).get("status") or "").lower()
            wstat = str((ch.get("weekly") or {}).get("status") or "").lower()
            if "bad" in (dstat, wstat):
                continue
            buy_rows.append({
                "ticker": t,
                "summary": f"{t} má buy signál {sig.get('score', '?')}/4 aj DIP kvalitu ({int(float(dip))}) — stojí za detail.",
            })
            used_tickers.add(t)
    except Exception:
        pass
    buy_rows = buy_rows[:6]

    quiet = sorted(t for t in holdings.keys() if t not in inbox_tickers)

    focus_syms = [r["ticker"] for r in focus_rows if r.get("ticker")]
    if focus_syms:
        extra = f" (+{len(focus_syms) - 3} ďalšie)" if len(focus_syms) > 3 else ""
        headline = "Tento týždeň rieš hlavne " + ", ".join(focus_syms[:3]) + extra + "."
    else:
        headline = "Pokojný týždeň — nič nevyžaduje tvoju akciu."

    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "headline": headline,
        "focus": focus_rows,
        "buy_candidates": buy_rows,
        "dca": dca_rows,
        "risks": risk_rows,
        "quiet": {"count": len(quiet), "tickers": quiet},
    }
    return _investor_cache_set(cache_key, payload)


ASSISTANT_EXPORT_SCHEMA_VERSION = "1.2"


def _assistant_snapshot(account: str) -> dict:
    """Return the existing processed portfolio snapshot without triggering eToro."""
    try:
        return _positions_cache.get(account) or cache_read(_portfolio_disk_path(account)) or {}
    except Exception:
        return {}


def _assistant_iso_timestamp(value) -> str | None:
    try:
        return datetime.fromtimestamp(float(value), timezone.utc).isoformat()
    except (TypeError, ValueError, OSError, OverflowError):
        return None


def _assistant_number(value):
    value = _num_or_none(value)
    return value if value is not None else None


def _assistant_chart_state(row: dict) -> dict:
    health = row.get("chart_health") or {}
    daily = health.get("daily") or {}
    weekly = health.get("weekly") or {}
    return {
        "daily_state": daily.get("status"),
        "weekly_state": weekly.get("status"),
        "daily_reason": daily.get("reason"),
        "weekly_reason": weekly.get("reason"),
    }


def _assistant_compact(data: dict) -> dict:
    """Keep export factual and compact: omit unavailable fields instead of null noise."""
    out = {}
    for key, value in data.items():
        if isinstance(value, dict):
            value = _assistant_compact(value)
        elif isinstance(value, list):
            value = [_assistant_compact(item) if isinstance(item, dict) else item for item in value]
        if value is not None and value != [] and value != {}:
            out[key] = value
    return out


def _assistant_days_since(timestamp: str | None) -> int | None:
    try:
        then = datetime.fromisoformat(str(timestamp).replace("Z", "+00:00"))
        return max(0, (datetime.now(timezone.utc).date() - then.date()).days)
    except (TypeError, ValueError):
        return None


def _assistant_export_position(symbol: str, lots: list[dict], scanner_row: dict,
                               dip: dict, earnings: dict, settings: dict,
                               total_equity: float) -> dict:
    amount = sum(float(p.get("amount") or 0) for p in lots)
    pnl = sum(float(p.get("pnl") or 0) for p in lots)
    current_rate = next((p.get("currentRate") for p in reversed(lots) if p.get("currentRate") is not None), None)
    last_buy = max(lots, key=lambda p: str(p.get("openDateTime") or ""), default={})
    open_lots = [{
        "opened_at": p.get("openDateTime") or None,
        "entry_price": _assistant_number(p.get("openRate")),
        "invested": _assistant_number(p.get("amount")),
        "profit_loss": _assistant_number(p.get("pnl")),
        "profit_loss_pct": _assistant_number(p.get("pnlPct")),
    } for p in lots[:20]]
    signal = scanner_row.get("recent_signal") or {}
    chart_state = _assistant_chart_state(scanner_row)
    last_entry = _assistant_number(last_buy.get("openRate"))
    current_price = _assistant_number(current_rate)
    total_pnl_pct = round(pnl / amount * 100, 2) if amount else None
    last_lot_pnl_pct = _assistant_number(last_buy.get("pnlPct"))
    change_from_last_entry = round((current_price - last_entry) / last_entry * 100, 2) if current_price and last_entry else None
    dip_total = _assistant_number((dip or {}).get("total"))
    trigger_from_last_entry = change_from_last_entry is not None and change_from_last_entry <= -float(settings["dca_loss_pct"])
    trigger_from_total = total_pnl_pct is not None and total_pnl_pct <= -float(settings["dca_loss_pct"])
    dip_ok = dip_total is not None and dip_total >= float(settings["dca_dip_min"])
    blocking = [
        name for name, state in (("daily_bad", chart_state.get("daily_state")), ("weekly_bad", chart_state.get("weekly_state")))
        if str(state or "").lower() == "bad"
    ]
    dca_status = "not_eligible"
    if trigger_from_total and dip_ok:
        dca_status = "conditional" if blocking else "eligible"
    return _assistant_compact({
        "ticker": symbol,
        "name": next((p.get("name") for p in lots if p.get("name")), symbol),
        "asset_type": next((p.get("type") for p in lots if p.get("type")), None),
        "position": {
            "invested": round(amount, 2),
            "current_value": round(amount + pnl, 2),
            "current_price": current_price,
            "profit_loss": round(pnl, 2),
            "profit_loss_pct": total_pnl_pct,
            "invested_weight_pct": round(amount / total_equity * 100, 2) if total_equity else None,
            "market_value_weight_pct": round((amount + pnl) / total_equity * 100, 2) if total_equity else None,
            "last_buy": {
                "opened_at": last_buy.get("openDateTime") or None,
                "entry_price": last_entry,
                "profit_loss_pct": last_lot_pnl_pct,
                "days_since": _assistant_days_since(last_buy.get("openDateTime")),
            },
            "open_lots": open_lots,
        },
        "signals": {
            "dip_rank": _assistant_number((dip or {}).get("rank")),
            "dip_score": _assistant_number((dip or {}).get("total")),
            "fa_score": _assistant_number((dip or {}).get("fa")),
            "ta_score": _assistant_number((dip or {}).get("ta")),
            "dip_strength": (dip or {}).get("label"),
            "signal_strength": signal.get("score"),
            "signal_tier": signal.get("tier"),
            **chart_state,
        },
        "earnings": _assistant_compact({
            "date": (earnings or {}).get("date"),
            "days_remaining": (earnings or {}).get("days"),
            "confirmed": bool((earnings or {}).get("date")),
        }),
        "dca_context": {
            "status": dca_status,
            "portfolio_pnl_pct": total_pnl_pct,
            "last_lot_pnl_pct": last_lot_pnl_pct,
            "change_from_last_entry_pct": change_from_last_entry,
            "dca_drawdown_from_last_entry_pct": change_from_last_entry if change_from_last_entry is not None and change_from_last_entry < 0 else None,
            "trigger_met_from_last_entry": trigger_from_last_entry,
            "trigger_met_on_total_position": trigger_from_total,
            "dip_score_threshold_met": dip_ok,
            "positive_conditions": [name for name, met in (("portfolio_drawdown", trigger_from_total), ("dip_score", dip_ok)) if met],
            "blocking_conditions": blocking,
        },
    })


@app.get("/api/assistant/export")
def get_assistant_export():
    """Read-only AI handoff. Protected by normal Basic Auth, never public-token auth.
    It reuses dashboard snapshots and scanner/DIP cache; it never explicitly requests refresh=1."""
    now = datetime.now(timezone.utc)
    settings = _dash_settings()
    snapshots = {account: _assistant_snapshot(account) for account in ("1", "2")}
    positions_by_symbol: dict[str, list[dict]] = {}
    all_position_symbols: set[str] = set()
    excluded_crypto_symbols: set[str] = set()
    excluded_other_symbols: set[str] = set()
    orders = []
    summary = {"cash": 0.0, "invested": 0.0, "equity": 0.0, "total_pnl": 0.0, "daily_pnl": 0.0}
    source_times = {}
    for account, snapshot in snapshots.items():
        source_times[f"portfolio_account_{account}"] = _assistant_iso_timestamp(snapshot.get("ts"))
        for key in summary:
            summary[key] += float((snapshot.get("summary") or {}).get(key) or 0)
        for position in snapshot.get("data") or []:
            symbol = str(position.get("symbol") or "").upper()
            asset_type = str(position.get("type") or "")
            if symbol:
                all_position_symbols.add(symbol)
            if asset_type not in {"Stock", "ETF"}:
                if asset_type == "Crypto":
                    excluded_crypto_symbols.add(symbol)
                else:
                    excluded_other_symbols.add(symbol)
                continue
            if symbol:
                positions_by_symbol.setdefault(symbol, []).append(position)
        for order in snapshot.get("orders") or []:
            if str(order.get("type") or "") not in {"Stock", "ETF"}:
                continue
            orders.append({
                "ticker": str(order.get("symbol") or "").upper(),
                "kind": order.get("kind"),
                "side": "buy" if order.get("isBuy", True) else "sell",
                "limit_price": _assistant_number(order.get("rate")),
                "invested": _assistant_number(order.get("amount")),
                "opened_at": order.get("openDateTime") or None,
            })

    dip_raw = load_dip_scores()
    dip_meta = dip_raw.get("_meta") if isinstance(dip_raw.get("_meta"), dict) else {}
    dip_scores = {str(k).upper(): v for k, v in dip_raw.items() if not str(k).startswith("_") and isinstance(v, dict)}
    raw_scan = load_scanner_cache()
    scanner_rows = raw_scan.get("results") if isinstance(raw_scan, dict) else []
    scanner_rows = scanner_rows if isinstance(scanner_rows, list) else []
    scanner_by_symbol = {str(row.get("ticker") or "").upper(): row for row in scanner_rows if isinstance(row, dict)}

    try:
        inbox = get_investor_inbox(refresh=0)
    except Exception:
        inbox = {"items": [], "generated_at": None}
    try:
        earnings = get_earnings_calendar_view(days=30, refresh=0)
    except Exception:
        earnings = {"items": [], "generated_at": None}
    earnings_by_symbol = {
        str(item.get("ticker") or "").upper(): item
        for item in (earnings.get("items") or []) if isinstance(item, dict)
    }

    total_equity = summary["equity"]
    portfolio = []
    for symbol in sorted(positions_by_symbol):
        row = _assistant_export_position(symbol, positions_by_symbol[symbol], scanner_by_symbol.get(symbol) or {},
                                         dip_scores.get(symbol) or {}, earnings_by_symbol.get(symbol) or {}, settings, total_equity)
        portfolio.append(row)

    portfolio_by_symbol = {row["ticker"]: row for row in portfolio}

    def attention_item(inbox_item: dict) -> dict | None:
        ticker = str(inbox_item.get("ticker") or "").upper()
        position = portfolio_by_symbol.get(ticker)
        if not position:
            return None
        kinds = set(inbox_item.get("kinds") or [])
        reasons = inbox_item.get("reasons") or []
        broken_titles = {str(reason.get("title") or "").lower() for reason in reasons if isinstance(reason, dict)}
        has_chart_break = any("graf" in title for title in broken_titles)
        has_dca_quality_warning = "broken" in kinds and not has_chart_break
        dca = position.get("dca_context") or {}
        types = []
        if dca.get("status") in {"eligible", "conditional"}:
            types.append("dca_candidate")
        if has_chart_break:
            types.append("broken_chart")
        elif has_dca_quality_warning:
            types.append("dca_quality_warning")
        if "earnings" in kinds:
            types.append("earnings")
        if "profit" in kinds:
            types.append("profit_review")
        if not types:
            return None
        action_type = "chart_review" if "broken_chart" in types else (
            "dca_review" if {"dca_candidate", "dca_quality_warning"} & set(types) else (
                "earnings_review" if "earnings" in types else "profit_review"
            )
        )
        status = "conditional" if dca.get("status") == "conditional" else (
            "confirmed" if dca.get("status") == "eligible" else "review")
        if "broken_chart" in types:
            severity = "warning"
        elif "earnings" in types:
            severity = "watch"
        else:
            severity = "info"
        return _assistant_compact({
            "ticker": ticker,
            "types": types,
            "action_type": action_type,
            "priority": inbox_item.get("priority"),
            "status": status,
            "severity": severity,
            "facts": {
                "position_pnl_pct": position["position"].get("profit_loss_pct"),
                "last_lot_pnl_pct": (position.get("dca_context") or {}).get("last_lot_pnl_pct"),
                "dip_score": (position.get("signals") or {}).get("dip_score"),
                "daily_state": (position.get("signals") or {}).get("daily_state"),
                "weekly_state": (position.get("signals") or {}).get("weekly_state"),
                "earnings_date": (position.get("earnings") or {}).get("date"),
            },
            "positive_conditions": (position.get("dca_context") or {}).get("positive_conditions"),
            "blocking_conditions": (position.get("dca_context") or {}).get("blocking_conditions"),
        })

    attention_items = [item for item in (attention_item(item) for item in (inbox.get("items") or [])) if item]
    attention_tickers = {item["ticker"] for item in attention_items}
    for ticker, position in portfolio_by_symbol.items():
        dca = position.get("dca_context") or {}
        if ticker in attention_tickers or dca.get("status") not in {"eligible", "conditional"}:
            continue
        attention_items.append(_assistant_compact({
            "ticker": ticker, "types": ["dca_candidate"], "priority": 50,
            "status": dca.get("status"), "severity": "watch" if dca.get("status") == "conditional" else "info",
            "facts": {"position_pnl_pct": position["position"].get("profit_loss_pct"), "dip_score": position["signals"].get("dip_score")},
            "positive_conditions": dca.get("positive_conditions"), "blocking_conditions": dca.get("blocking_conditions"),
        }))
    attention_items.sort(key=lambda item: (item.get("priority", 99), item.get("ticker", "")))

    def dip_export_row(symbol: str, dip: dict) -> dict:
        return _assistant_compact({
            "rank": _assistant_number(dip.get("rank")), "ticker": symbol,
            "company": dip.get("company"), "in_portfolio": symbol in positions_by_symbol,
            "dip_score": _assistant_number(dip.get("total")), "fa_score": _assistant_number(dip.get("fa")),
            "ta_score": _assistant_number(dip.get("ta")), "dip_strength": dip.get("label"),
        })

    ranked_dip = sorted(dip_scores.items(), key=lambda item: (_assistant_number(item[1].get("rank")) or 999999, item[0]))
    dip_scan = [dip_export_row(symbol, dip) for symbol, dip in ranked_dip[:100]]
    new_candidates_priority = []
    new_candidates_watch = []
    new_candidates_low_confidence = []
    selected_priority_symbols = set()
    for symbol, scan in scanner_by_symbol.items():
        if symbol in positions_by_symbol:
            continue
        signal = scan.get("recent_signal") or {}
        if not signal:
            continue
        dip = dip_scores.get(symbol) or {}
        candidate = _assistant_compact({
            "ticker": symbol, "company": scan.get("name") or dip.get("company"),
            "dip_rank": _assistant_number(dip.get("rank")), "dip_score": _assistant_number(dip.get("total")),
            "fa_score": _assistant_number(dip.get("fa")), "ta_score": _assistant_number(dip.get("ta")),
            "signal_strength": signal.get("score"), "signal_tier": signal.get("tier"),
            "last_price": _assistant_number(scan.get("last") or scan.get("price")),
            "chart_state": _assistant_chart_state(scan),
            "passes_dip_threshold": _assistant_number(dip.get("total")) is not None and _assistant_number(dip.get("total")) >= float(settings["dca_dip_min"]),
        })
        if candidate.get("passes_dip_threshold"):
            new_candidates_priority.append(candidate)
            selected_priority_symbols.add(symbol)
        elif candidate.get("dip_score") is not None:
            new_candidates_watch.append(candidate)
        else:
            new_candidates_low_confidence.append(candidate)
    for candidates in (new_candidates_priority, new_candidates_watch, new_candidates_low_confidence):
        candidates.sort(key=lambda item: (-(item.get("dip_score") or 0), item.get("ticker", "")))

    top_ranked_not_selected = []
    for symbol, dip in ranked_dip:
        if symbol in positions_by_symbol or symbol in selected_priority_symbols:
            continue
        scan = scanner_by_symbol.get(symbol)
        signal = (scan or {}).get("recent_signal") or {}
        dip_score = _assistant_number(dip.get("total"))
        if not scan:
            reason = "not_in_scanner_cache"
        elif not signal:
            reason = "no_scanner_signal"
        elif dip_score is None:
            reason = "missing_dip_score"
        else:
            reason = "dip_score_below_threshold"
        top_ranked_not_selected.append(_assistant_compact({
            "ticker": symbol, "company": dip.get("company"), "dip_rank": _assistant_number(dip.get("rank")),
            "dip_score": dip_score, "reason": reason,
        }))
        if len(top_ranked_not_selected) >= 25:
            break

    cash_reserved_for_orders = sum(float(order.get("invested") or 0) for order in orders if order.get("side") == "buy")
    watchlist = _read_watchlist_file()
    source_times.update({
        "dip_import": dip_meta.get("updated_at"),
        "scanner": raw_scan.get("finished_at") or raw_scan.get("generated_at") if isinstance(raw_scan, dict) else None,
        "investor_inbox": inbox.get("generated_at"),
        "earnings": earnings.get("generated_at"),
    })
    field_definitions = {
        "dip_score": "Combined imported DIP opportunity score; higher means stronger ranking, not an automatic buy.",
        "fa_score": "Fundamental-analysis component of the imported DIP score.",
        "ta_score": "Technical-analysis component of the imported DIP score.",
        "signal_strength": "Number of fulfilled dashboard setup conditions, typically from 0 to 4.",
        "current_value": "Invested value plus open P/L. It is the dashboard market-value approximation for concentration review.",
        "invested_weight_pct": "Original invested value divided by total dashboard equity.",
        "market_value_weight_pct": "Current value divided by total dashboard equity, including asset classes excluded from this analysis; use this for concentration review.",
        "dca_context": "Explicit DCA evidence. eligible means current conditions pass; conditional means positive conditions exist but chart blockers require manual review.",
        "change_from_last_entry_pct": "Signed price change from the latest open lot entry. Positive means current price is above that entry; negative means below it.",
        "dca_drawdown_from_last_entry_pct": "Negative-only price change from the latest open lot entry. Omitted when the current price is at or above that entry.",
        "attention_items": "One normalized list of held tickers that need review. It replaces duplicated weekly-plan and inbox text in this export.",
        "passes_dip_threshold": "Whether the imported DIP score meets the configured DCA minimum. It separates priority candidates from scanner-only watch candidates.",
    }
    return {
        "schema_version": ASSISTANT_EXPORT_SCHEMA_VERSION,
        "generated_at": now.isoformat(),
        "currency": "USD",
        "strategy": {
            "name": "DIP Strategy v2",
            "dca_trigger_pct": -float(settings["dca_loss_pct"]),
            "dca_min_dip_score": settings["dca_dip_min"],
            "holding_horizon": "1-3 years",
            "buy_only": True,
            "notes": ["DIP signal is not an automatic buy.", "Verify weekly and daily chart before DCA."],
        },
        "analysis_scope": {
            "include_asset_types": ["Stock", "ETF"],
            "exclude_crypto_from_dip_analysis": True,
            "exclude_crypto_from_export": True,
        },
        "field_definitions": field_definitions,
        "data_quality": {
            "source_timestamps": _assistant_compact(source_times),
            "portfolio_source": "Processed eToro snapshot cache; this export does not force a refresh.",
            "limitations": [
                "Daily P/L is a dashboard approximation based on live price and previous market close.",
                "Scanner and DIP data can be older than the portfolio snapshot; inspect source_timestamps.",
                "Dashboard recommendations are interpretive aids, not trading instructions.",
            ],
        },
        "portfolio_summary": {
            "account_value": round(summary["equity"], 2), "invested_value": round(summary["invested"], 2),
            "cash_available": round(summary["cash"], 2),
            "cash_reserved_for_orders": round(cash_reserved_for_orders, 2),
            "cash_free_after_orders": round(max(0.0, summary["cash"] - cash_reserved_for_orders), 2),
            "profit_loss": round(summary["total_pnl"], 2), "daily_profit_loss": round(summary["daily_pnl"], 2),
            "positions_count_total": len(all_position_symbols), "positions_count_exported": len(portfolio),
            "positions_count_excluded_crypto": len(excluded_crypto_symbols), "positions_count_excluded_other": len(excluded_other_symbols),
            "pending_orders_count": len(orders),
        },
        "attention_items": attention_items,
        "earnings_held_positions": [
            _assistant_compact({"ticker": item.get("ticker"), "date": item.get("date"), "days_remaining": item.get("days"),
                                "position_pnl_pct": item.get("pnl_pct")})
            for item in (earnings.get("items") or []) if item.get("in_portfolio") and str(item.get("ticker") or "").upper() in positions_by_symbol
        ],
        "watchlist": [{"ticker": item.get("symbol"), "name": item.get("name"), "added_at": item.get("addedAt")} for item in watchlist],
        "positions": portfolio,
        "pending_orders": orders,
        "dip_ranking_top_100": {"meta": dip_meta, "items": dip_scan},
        "new_candidates_priority": new_candidates_priority[:50],
        "new_candidates_watch": new_candidates_watch[:50],
        "new_candidates_low_confidence": new_candidates_low_confidence[:50],
        "top_ranked_not_selected": top_ranked_not_selected,
        "analysis_request": {
            "task": "Analyze portfolio and DIP scan.",
            "questions": [
                "Which existing positions are suitable for DCA?",
                "Which positions may be value traps or need chart review?",
                "Which new candidates deserve manual chart review?",
                "What should be prioritized this week?",
            ],
            "response_language": "sk",
            "desired_style": "direct, skeptical, evidence-based",
            "requested_output_contract": ["separate facts from interpretation", "flag conflicting signals and missing data", "do not invent unavailable data"],
        },
    }


@app.get("/api/scanner/notes")
def get_scanner_notes():
    if SCANNER_NOTES_FILE.exists():
        try:
            data = json.loads(SCANNER_NOTES_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict) and "content" in data:
                return data
        except Exception:
            pass
    return {"content": ""}


@app.post("/api/scanner/notes")
async def save_scanner_note(request: Request):
    body = await request.json()
    content = str(body.get("content", ""))
    SCANNER_NOTES_FILE.write_text(
        json.dumps({"content": content, "updated_at": datetime.now(timezone.utc).isoformat()}, ensure_ascii=False),
        encoding="utf-8",
    )
    return {"ok": True}


@app.post("/api/admin/backfill-regime-context")
def backfill_regime_context(target: str = Query("log"), ticker: str = Query(None),
                            limit: int = Query(10), force: int = Query(0)):
    """Doplní regime kontext do historických signálov, ktorým chýba (boli zalogované
    pred SIGNAL_CONTEXT_VERSION). Idempotentné a inkrementálne: spúšťaj opakovane,
    `limit` tickerov na volanie (HMM fit per signál je pamäťovo náročný). Reuse
    build_signal_context — žiadny look-ahead. Tag context_source='backfill'.
    target=log|archive|both. Mutuje len chýbajúci `context`, ostatné polia nechá."""
    targets = ["log", "archive"] if target == "both" else [target]
    summary = {"target": target, "tickers_processed": 0, "backfilled": 0,
               "errors": 0, "remaining_tickers": 0, "per_ticker": {}}
    for tgt in targets:
        if tgt == "log":
            store = load_signals_log(); save = save_signals_log
        elif tgt == "archive":
            store = load_signals_archive()
            save = lambda d: SIGNALS_ARCHIVE.write_text(
                json.dumps(d, indent=2, ensure_ascii=False), encoding="utf-8")
        else:
            continue
        # tickery s aspoň jedným signálom bez kontextu
        pending = [t for t, ents in store.items() if isinstance(ents, dict) and
                   any(isinstance(e, dict) and (force or not e.get("context")) for e in ents.values())]
        if ticker:
            pending = [t for t in pending if t == ticker.upper()]
        summary["remaining_tickers"] += max(0, len(pending) - limit) if not ticker else 0
        changed = False
        for t in pending[:limit]:
            res = _backfill_ticker_context(t, store[t], force=bool(force))
            summary["per_ticker"][t] = res
            summary["tickers_processed"] += 1
            summary["backfilled"] += res.get("backfilled", 0)
            summary["errors"] += res.get("errors", 0)
            if res.get("backfilled"):
                changed = True
        if changed:
            try:
                save(store)
            except Exception as e:
                summary["save_error"] = str(e)
    return summary


# ── News sentiment (Alpha Vantage) ────────────────────────────────────────────
# Lazy, cache-first: free tier = 25 req/deň, preto 12h TTL na persistent disku
# a stale fallback pri chybe/limite. Nikdy nesmie zhodiť scanner.

NEWS_CACHE_DIR = DATA_ROOT / "news_cache"
NEWS_CACHE_TTL_H = 12
NEWS_RELEVANCE_MIN = 0.15
NEWS_MAX_ITEMS = 10

# ── ApeWisdom Reddit mentions ──────────────────────────────────────────────────

APE_CACHE_FILE = NEWS_CACHE_DIR / "_apewisdom.json"
APE_CACHE_TTL_H = 6


def _ape_cache_read() -> dict | None:
    if APE_CACHE_FILE.exists():
        try:
            return json.loads(APE_CACHE_FILE.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None


_APE_BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36"
)
_APE_FETCH_LOCK = asyncio.Lock()


def _ape_merge_pages(pages: list) -> dict:
    merged: dict = {}
    for page in pages:
        for item in (page.get("results") or []):
            t = (item.get("ticker") or "").upper()
            if not t:
                continue
            merged[t] = {
                "ticker": t,
                "name": item.get("name"),
                "mentions": item.get("mentions"),
                "upvotes": item.get("upvotes"),
                "rank": item.get("rank"),
                "rank_24h_ago": item.get("rank_24h_ago"),
                "mentions_24h_ago": item.get("mentions_24h_ago"),
            }
    return merged


async def _ape_fetch_server() -> dict | None:
    """Stiahne top 5 strán ApeWisdom zo servera s browser User-Agent."""
    import httpx
    pages = []
    total_pages = 5
    try:
        async with httpx.AsyncClient(
            headers={"User-Agent": _APE_BROWSER_UA},
            timeout=15,
            follow_redirects=True,
        ) as client:
            for p in range(1, min(total_pages, 5) + 1):
                r = await client.get(
                    f"https://apewisdom.io/api/v1.0/filter/all-stocks/page/{p}"
                )
                if r.status_code != 200:
                    print(f"[ape] HTTP {r.status_code} page {p}")
                    break
                data = r.json()
                if data.get("pages"):
                    total_pages = data["pages"]
                pages.append(data)
    except Exception as exc:
        print(f"[ape] fetch error: {exc}")
        return None
    if not pages:
        return None
    return _ape_merge_pages(pages)


@app.get("/api/reddit/mentions")
async def get_reddit_mentions(tickers: str = Query("")):
    """Vráti Reddit mentions pre zoznam tickerov. Ak je cache stará, obnoví ju server-side."""
    cached = _ape_cache_read()
    age_h: float | None = None
    if cached:
        try:
            age_h = (datetime.now(timezone.utc) -
                     datetime.fromisoformat(cached["fetched_at"])).total_seconds() / 3600
        except Exception:
            pass

    # Obnov cache ak je stará alebo prázdna (non-blocking pre prvé volanie — lock zabraňuje súbežným fetchom)
    if cached is None or (age_h is not None and age_h >= APE_CACHE_TTL_H):
        if not _APE_FETCH_LOCK.locked():
            async with _APE_FETCH_LOCK:
                merged = await _ape_fetch_server()
                if merged is not None:
                    payload = {
                        "fetched_at": datetime.now(timezone.utc).isoformat(),
                        "data": merged,
                    }
                    NEWS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
                    APE_CACHE_FILE.write_text(
                        json.dumps(payload, ensure_ascii=False), encoding="utf-8"
                    )
                    cached = payload
                    age_h = 0.0

    ticker_list = [t.strip().upper() for t in tickers.split(",") if t.strip()]
    data = {
        t: cached["data"].get(t)
        for t in ticker_list
        if cached and cached.get("data", {}).get(t)
    } if cached else {}
    return {
        "data": data,
        "fetched_at": cached.get("fetched_at") if cached else None,
        "age_h": round(age_h, 1) if age_h is not None else None,
        "stale": age_h is not None and age_h >= APE_CACHE_TTL_H,
        "empty": not cached or not cached.get("data"),
    }



def _news_cache_path(ticker: str) -> Path:
    return NEWS_CACHE_DIR / f"{ticker.upper()}.json"


def _news_cache_read(ticker: str) -> dict | None:
    p = _news_cache_path(ticker)
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None


def _news_scrub_error(msg: str) -> str:
    """AV vkladá API kľúč do rate-limit hlášky — nikdy ho nepustiť do UI/cache."""
    key = os.getenv("ALPHA_VANTAGE_API_KEY", "")
    text = str(msg)
    if key:
        text = text.replace(key, "***")
    return text[:200]


# ── News clustering ────────────────────────────────────────────────────────
# Viac vydavateľov často prevezme tú istú správu (wire story) takmer s rovnakým
# titulkom. Bez klastrovania by taká udalosť vážila v priemernom sentimente
# N-krát namiesto raz. Lacná heuristika (Jaccard prienik tokenov titulku) —
# žiadne NLP knižnice, žiadny nový RAM footprint.
_NEWS_STOPWORDS = {
    "the", "a", "an", "to", "of", "in", "on", "for", "and", "or", "is", "are",
    "as", "at", "by", "its", "it", "with", "after", "before", "amid", "says",
    "say", "will", "has", "have", "had", "from", "into", "over", "up", "down",
    "than", "that", "this", "be", "was", "were", "vs",
}


def _news_title_tokens(title: str) -> set:
    words = re.findall(r"[a-z0-9]+", (title or "").lower())
    return {w for w in words if len(w) > 2 and w not in _NEWS_STOPWORDS}


def _news_cluster_items(items: list[dict], threshold: float = 0.5) -> list[dict]:
    """Zoskupí články pokrývajúce tú istú udalicu — Jaccard prienik tokenov titulku
    >= threshold. Union-find, O(n^2) na n<=NEWS_MAX_ITEMS (lacné, žiadny externý
    volanie). Pridá cluster_id/cluster_size/cluster_primary; poradie a obsah
    zoznamu inak nemení — čisto anotácia nad už vybranými top-N článkami."""
    n = len(items)
    if n <= 1:
        for it in items:
            it["cluster_id"] = 0
            it["cluster_size"] = 1
            it["cluster_primary"] = True
        return items
    token_sets = [_news_title_tokens(it.get("title")) for it in items]
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    for i in range(n):
        if not token_sets[i]:
            continue
        for j in range(i + 1, n):
            if not token_sets[j]:
                continue
            inter = token_sets[i] & token_sets[j]
            uni = token_sets[i] | token_sets[j]
            if uni and len(inter) / len(uni) >= threshold:
                union(i, j)

    groups: dict[int, list[int]] = {}
    for i in range(n):
        groups.setdefault(find(i), []).append(i)
    # zoradenie klastrov podľa najsilnejšieho člena — len pre stabilné cluster_id,
    # neovplyvňuje poradie v zozname (to ostáva podľa items ako prišli)
    ordered = sorted(groups.values(),
                      key=lambda idxs: max(items[i].get("relevance") or 0 for i in idxs),
                      reverse=True)
    for cid, idxs in enumerate(ordered):
        primary = max(idxs, key=lambda i: items[i].get("relevance") or 0)
        for i in idxs:
            items[i]["cluster_id"] = cid
            items[i]["cluster_size"] = len(idxs)
            items[i]["cluster_primary"] = (i == primary)
    return items


def _news_parse_feed(ticker: str, data: dict) -> list[dict]:
    """Normalizuje surovú AV NEWS_SENTIMENT odpoveď na zoznam položiek."""
    # AV vracia 200 aj pri rate-limite — limit hláška je v "Note"/"Information"
    if "feed" not in data:
        msg = data.get("Note") or data.get("Information") or data.get("Error Message") or "unexpected response"
        raise RuntimeError(_news_scrub_error(msg))

    items = []
    for art in data.get("feed", []):
        # ticker-specific sentiment, nie overall (článok môže hodnotiť iný titul)
        ts = next((t for t in art.get("ticker_sentiment", [])
                   if t.get("ticker", "").upper() == ticker.upper()), None)
        if not ts:
            continue
        relevance = _num_or_none(ts.get("relevance_score")) or 0.0
        if relevance < NEWS_RELEVANCE_MIN:
            continue
        # time_published: "20260610T143000" → ISO
        tp_raw = art.get("time_published", "")
        try:
            tp = datetime.strptime(tp_raw, "%Y%m%dT%H%M%S").replace(tzinfo=timezone.utc).isoformat()
        except Exception:
            tp = None
        items.append({
            "title": art.get("title"),
            "url": art.get("url"),
            "source": art.get("source"),
            "time_published": tp,
            "sentiment_label": ts.get("ticker_sentiment_label"),
            "sentiment_score": _num_or_none(ts.get("ticker_sentiment_score")),
            "relevance": round(relevance, 3),
        })
    # aktuálnosť + relevancia: novšie a relevantnejšie hore
    items.sort(key=lambda a: (a.get("time_published") or "", a.get("relevance") or 0), reverse=True)
    items = items[:NEWS_MAX_ITEMS]
    return _news_cluster_items(items)


def _news_fetch_av(ticker: str) -> list[dict]:
    """Stiahne a normalizuje news sentiment z Alpha Vantage pre jeden ticker."""
    api_key = os.getenv("ALPHA_VANTAGE_API_KEY", "")
    if not api_key:
        raise RuntimeError("ALPHA_VANTAGE_API_KEY nie je nastavený")
    resp = requests.get(
        "https://www.alphavantage.co/query",
        params={"function": "NEWS_SENTIMENT", "tickers": ticker.upper(),
                "limit": 50, "apikey": api_key},
        timeout=20,
    )
    resp.raise_for_status()
    return _news_parse_feed(ticker, resp.json())


def _news_save_cache(ticker: str, items: list[dict]) -> dict:
    payload = {
        "ticker": ticker,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "items": items,
    }
    NEWS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    _news_cache_path(ticker).write_text(
        json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return payload


@app.get("/api/news/clientkey")
def get_news_clientkey():
    """Vydá AV kľúč prehliadaču — fallback keď je Render zdieľaná IP rate-limitnutá.
    Endpoint je za basic auth ako všetko ostatné; kľúč nikdy necachovať na disk."""
    key = os.getenv("ALPHA_VANTAGE_API_KEY", "")
    if not key:
        raise HTTPException(404, "ALPHA_VANTAGE_API_KEY nie je nastavený")
    return {"key": key}


@app.get("/api/news/summary")
def get_news_summary(tickers: str = Query("")):
    """Agregovaný sentiment z disk cache — žiadne AV requesty, len prečíta
    čo už bolo stiahnuté cez 📰. Relevance-weighted priemer sentiment skóre.

    Počíta sa len z cluster_primary článkov (jeden reprezentant na udalosť) —
    inak by udalosť pokrytá 5 vydavateľmi vážila v priemere 5x viac než udalosť
    s jediným článkom. Staré cache záznamy bez cluster_primary poľa (pred
    zavedením clusteringu) sa berú ako primary, kým sa cache neobnoví (12h TTL)."""
    out = {}
    for t in tickers.split(","):
        t = t.strip().upper()
        if not t or len(t) > 12:
            continue
        cached = _news_cache_read(t)
        items = (cached or {}).get("items") or []
        primary_items = [i for i in items if i.get("cluster_primary", True)]
        scored = [(i.get("sentiment_score"), i.get("relevance") or 0.0)
                  for i in primary_items if i.get("sentiment_score") is not None]
        if not scored:
            continue
        wsum = sum(r for _, r in scored) or 1.0
        avg = sum(s * r for s, r in scored) / wsum
        out[t] = {"avg": round(avg, 3), "n": len(primary_items), "n_articles": len(items),
                  "fetched_at": cached.get("fetched_at")}
    return {"summary": out}


# ── Earnings calendar (Alpha Vantage EARNINGS_CALENDAR, 1 request pre celý trh) ──

EARNINGS_CACHE_FILE = NEWS_CACHE_DIR / "_earnings_calendar.json"
EARNINGS_CACHE_TTL_H = 24


def _earnings_parse_csv(text: str) -> dict:
    """CSV: symbol,name,reportDate,... → {SYMBOL: najbližší reportDate}."""
    if not text or text.lstrip().startswith("{"):
        # AV vracia JSON namiesto CSV pri chybe/rate-limite
        try:
            data = json.loads(text)
            msg = data.get("Note") or data.get("Information") or data.get("Error Message") or "unexpected response"
        except Exception:
            msg = "unexpected response"
        raise RuntimeError(_news_scrub_error(msg))
    dates: dict = {}
    lines = text.strip().splitlines()
    if not lines or "symbol" not in lines[0].lower():
        raise RuntimeError("unexpected CSV format")
    for line in lines[1:]:
        parts = line.split(",")
        if len(parts) < 3:
            continue
        sym, report = parts[0].strip().upper(), parts[2].strip()
        if not sym or not report:
            continue
        if sym not in dates or report < dates[sym]:
            dates[sym] = report
    if not dates:
        raise RuntimeError("empty earnings calendar")
    return dates


def _earnings_cache_read() -> dict | None:
    if EARNINGS_CACHE_FILE.exists():
        try:
            return json.loads(EARNINGS_CACHE_FILE.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None


def _earnings_save_cache(dates: dict) -> dict:
    payload = {"fetched_at": datetime.now(timezone.utc).isoformat(), "dates": dates}
    NEWS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    EARNINGS_CACHE_FILE.write_text(json.dumps(payload), encoding="utf-8")
    return payload


def _earnings_fetch_finnhub() -> dict:
    """Finnhub /calendar/earnings — JSON, 60 req/min free tier, pokrýva 3 mesiace dopredu."""
    api_key = os.getenv("FINNHUB_API_KEY", "")
    if not api_key:
        raise RuntimeError("FINNHUB_API_KEY nie je nastavený")
    today = datetime.now(timezone.utc).date()
    resp = requests.get(
        "https://finnhub.io/api/v1/calendar/earnings",
        params={"from": today.isoformat(),
                "to": (today + timedelta(days=90)).isoformat(),
                "token": api_key},
        timeout=30,
    )
    resp.raise_for_status()
    dates: dict = {}
    for item in (resp.json().get("earningsCalendar") or []):
        sym = (item.get("symbol") or "").upper()
        report = item.get("date") or ""
        if not sym or not report:
            continue
        if sym not in dates or report < dates[sym]:
            dates[sym] = report
    if not dates:
        raise RuntimeError("empty earnings calendar (finnhub)")
    return dates


_earnings_bulk_failed_at = 0.0   # 2026-07-09: backoff keď OBAJA provideri zlyhali,
                                 # ale cache má (staršie) dates — bez toho by sa
                                 # Finnhub+AV skúšali odznova pri každom volaní
                                 # (get_chart, inbox, calendar widget...) donekonečna
EARNINGS_BULK_FAIL_BACKOFF = 900  # 15 min


@app.get("/api/earnings")
def get_earnings_calendar(refresh: int = Query(0)):
    global _earnings_bulk_failed_at
    cached = _earnings_cache_read()
    if cached and not refresh:
        try:
            fetched = datetime.fromisoformat(cached["fetched_at"])
            age_h = (datetime.now(timezone.utc) - fetched).total_seconds() / 3600
            ttl = 1 if cached.get("error") else EARNINGS_CACHE_TTL_H
            if age_h < ttl:
                return {**cached, "stale": False}
        except Exception:
            pass
    if not refresh and cached and cached.get("dates") and time.time() < _earnings_bulk_failed_at + EARNINGS_BULK_FAIL_BACKOFF:
        return {**cached, "stale": True}
    # Primárne Finnhub (60 req/min, bez zdieľaného IP limitu), fallback Alpha Vantage
    try:
        payload = _earnings_save_cache(_earnings_fetch_finnhub())
        return {**payload, "stale": False}
    except Exception as e:
        print(f"[earnings] finnhub failed, fallback AV: {e}")
    api_key = os.getenv("ALPHA_VANTAGE_API_KEY", "")
    try:
        if not api_key:
            raise RuntimeError("ALPHA_VANTAGE_API_KEY nie je nastavený")
        resp = requests.get(
            "https://www.alphavantage.co/query",
            params={"function": "EARNINGS_CALENDAR", "horizon": "3month", "apikey": api_key},
            timeout=30,
        )
        resp.raise_for_status()
        payload = _earnings_save_cache(_earnings_parse_csv(resp.text))
        return {**payload, "stale": False}
    except Exception as e:
        _earnings_bulk_failed_at = time.time()
        err = _news_scrub_error(str(e))
        if cached and cached.get("dates"):
            return {**cached, "stale": True, "error": err}
        payload = {"fetched_at": datetime.now(timezone.utc).isoformat(),
                   "dates": {}, "error": err}
        try:
            NEWS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
            EARNINGS_CACHE_FILE.write_text(json.dumps(payload), encoding="utf-8")
        except Exception:
            pass
        return {**payload, "stale": False}


@app.post("/api/earnings/ingest")
async def ingest_earnings_calendar(request: Request):
    """Surové CSV stiahnuté prehliadačom (browser-direct fallback ako pri news)."""
    text = (await request.body()).decode("utf-8", errors="replace")
    try:
        dates = _earnings_parse_csv(text)
    except Exception as e:
        return {"dates": {}, "stale": False,
                "fetched_at": datetime.now(timezone.utc).isoformat(),
                "error": _news_scrub_error(str(e))}
    payload = _earnings_save_cache(dates)
    return {**payload, "stale": False}


_earnings_sym_failed_at: dict[str, float] = {}   # sym → ts neúspechu (1h backoff)

# ── Yahoo quoteSummary (cookie + crumb session, ako yfinance ale pod kontrolou) ──
_YAHOO_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
             "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")
_yahoo_auth: dict = {"cookie": None, "crumb": None, "expiry": 0.0, "failed_until": 0.0}
_yahoo_auth_lock = threading.Lock()
_YAHOO_AUTH_BACKOFF = 600  # 10 min — Yahoo z Render IP prakticky nikdy neprejde
                          # (CLAUDE.md), bez backoffu každý fail-soft caller
                          # (insights, earnings, profile fallback) zopakuje
                          # 2 blokujúce HTTP volania nadarmo.


def _yahoo_get_auth(force: bool = False) -> dict | None:
    """Session cookie z fc.yahoo.com + crumb token. Cache 30 min."""
    with _yahoo_auth_lock:
        if not force and _yahoo_auth["crumb"] and time.time() < _yahoo_auth["expiry"]:
            return dict(_yahoo_auth)
        if not force and time.time() < _yahoo_auth.get("failed_until", 0):
            return None   # známy nedávny fail — neskúšaj znova naslepo
        try:
            s = requests.Session()
            s.headers["User-Agent"] = _YAHOO_UA
            # fc.yahoo.com vracia 404, ale nastaví session cookie — neraise-ovať
            s.get("https://fc.yahoo.com", timeout=10, allow_redirects=True)
            cookie = "; ".join(f"{c.name}={c.value}" for c in s.cookies)
            r = s.get("https://query1.finance.yahoo.com/v1/test/getcrumb", timeout=10)
            crumb = r.text.strip()
            if not cookie or not crumb or "<" in crumb:
                raise RuntimeError(f"bad crumb/cookie (crumb={crumb[:30]!r})")
            _yahoo_auth.update({"cookie": cookie, "crumb": crumb, "expiry": time.time() + 1800, "failed_until": 0.0})
            return dict(_yahoo_auth)
        except Exception as e:
            print(f"[yahoo] auth failed: {e}")
            _yahoo_auth.update({"cookie": None, "crumb": None, "failed_until": time.time() + _YAHOO_AUTH_BACKOFF})
            return None


def _yahoo_quote_summary(symbol: str, modules: str) -> dict | None:
    """GET v10/finance/quoteSummary — vráti result[0] alebo None (fail-soft)."""
    auth = _yahoo_get_auth()
    if not auth:
        return None
    url = f"https://query1.finance.yahoo.com/v10/finance/quoteSummary/{urllib.parse.quote(symbol)}"
    for attempt in (1, 2):
        try:
            r = requests.get(url,
                             params={"modules": modules, "crumb": auth["crumb"]},
                             headers={"User-Agent": _YAHOO_UA, "Cookie": auth["cookie"]},
                             timeout=15)
            if r.status_code in (401, 403) and attempt == 1:
                auth = _yahoo_get_auth(force=True)
                if not auth:
                    return None
                continue
            r.raise_for_status()
            res = (r.json().get("quoteSummary") or {}).get("result") or []
            return res[0] if res else None
        except Exception as e:
            if attempt == 2:
                print(f"[yahoo] quoteSummary {symbol} failed: {e}")
                return None
    return None


def _yraw(v):
    """Yahoo čísla sú {raw: 1.23, fmt: '1.23'} objekty."""
    return v.get("raw") if isinstance(v, dict) else v


# ── Ticker insights — insider transakcie + EPS história (Yahoo) ──────────────
YAHOO_INSIGHTS_DIR = DATA_ROOT / "yahoo_insights"
INSIGHTS_TTL_H = 12
INSIGHTS_SCHEMA_VERSION = 4


def _insights_parse(qs: dict) -> dict:
    out: dict = {}
    # Insider transakcie za 90 dní
    txs = (qs.get("insiderTransactions") or {}).get("transactions") or []
    cutoff = (datetime.now(timezone.utc) - timedelta(days=90)).timestamp()
    buys = sells = 0
    buy_val = sell_val = 0.0
    recent = []
    for t in txs:
        ts = _yraw(t.get("startDate"))
        if not ts or ts < cutoff:
            continue
        text = (t.get("transactionText") or "").lower()
        kind = "buy" if ("purchase" in text or "buy" in text) else ("sell" if "sale" in text else None)
        if not kind:
            continue   # awards/grants/exercise bez kontextu nezapočítavaj
        val = float(_yraw(t.get("value")) or 0)
        if kind == "buy":
            buys += 1; buy_val += val
        else:
            sells += 1; sell_val += val
        if len(recent) < 6:
            recent.append({"date": datetime.fromtimestamp(ts, timezone.utc).date().isoformat(),
                           "name": t.get("filerName"), "relation": t.get("filerRelation"),
                           "type": kind, "shares": _yraw(t.get("shares")), "value": val or None})
    out["insider"] = {"buys_90d": buys, "sells_90d": sells,
                      "net_value_90d": round(buy_val - sell_val, 0), "recent": recent}
    # EPS história — posledné 4 kvartály (beat/miss)
    hist = (qs.get("earningsHistory") or {}).get("history") or []
    eps_hist = []
    for h in hist:
        a, e = _yraw(h.get("epsActual")), _yraw(h.get("epsEstimate"))
        sp = _yraw(h.get("surprisePercent"))
        eps_hist.append({"quarter": h.get("quarter", {}).get("fmt") if isinstance(h.get("quarter"), dict) else h.get("quarter"),
                         "actual": a, "estimate": e,
                         "surprise_pct": round(sp * 100, 1) if sp is not None else None,
                         "beat": (a is not None and e is not None and a >= e)})
    out["eps_history"] = eps_hist
    # Odhad na aktuálny kvartál + rast
    for tr in (qs.get("earningsTrend") or {}).get("trend") or []:
        if tr.get("period") == "0q":
            out["eps_next"] = {"avg": _yraw((tr.get("earningsEstimate") or {}).get("avg")),
                               "analysts": _yraw((tr.get("earningsEstimate") or {}).get("numberOfAnalysts")),
                               "growth": _yraw(tr.get("growth"))}
            break
    # Earnings dátum z calendarEvents (bonus pre earnings reťazec)
    cal = ((qs.get("calendarEvents") or {}).get("earnings") or {}).get("earningsDate") or []
    dates = [d for d in (_yraw(x) for x in cal) if d]
    if dates:
        out["earnings_date"] = datetime.fromtimestamp(min(dates), timezone.utc).date().isoformat()
    trends = (qs.get("recommendationTrend") or {}).get("trend") or []
    if trends:
        tr = trends[0]
        out["analyst_consensus"] = {
            "period": tr.get("period"),
            "strong_buy": int(tr.get("strongBuy") or 0),
            "buy": int(tr.get("buy") or 0),
            "hold": int(tr.get("hold") or 0),
            "sell": int(tr.get("sell") or 0),
            "strong_sell": int(tr.get("strongSell") or 0),
        }
    fd = qs.get("financialData") or {}
    target_mean = _yraw(fd.get("targetMeanPrice"))
    target_low = _yraw(fd.get("targetLowPrice"))
    target_high = _yraw(fd.get("targetHighPrice"))
    if any(v is not None for v in (target_mean, target_low, target_high)):
        out["price_target"] = {
            "mean": target_mean,
            "low": target_low,
            "high": target_high,
            "median": _yraw(fd.get("targetMedianPrice")),
            "current": _yraw(fd.get("currentPrice")),
        }
    ks = qs.get("defaultKeyStatistics") or {}
    short_float = _yraw(ks.get("shortPercentOfFloat"))
    if short_float is not None:
        out["short_interest"] = {
            "percent_float": round(float(short_float) * 100, 2),
            "short_ratio": _yraw(ks.get("shortRatio")),
            "shares_short": _yraw(ks.get("sharesShort")),
            "source": "yahoo",
        }
    return out


def _scrub_token(msg: str) -> str:
    """requests chyby obsahujú celú URL vrátane ?token=/?apiKey= — maskuj kľúč.
    apiKey pridané 2026-07-09: Massive volania (raise_for_status()) mohli
    dostať MASSIVE_API_KEY do error stringu, ktorý sa ukladal aj do
    company_profiles/{SYM}.json na disku."""
    return re.sub(r"(token|apiKey)=[^&\s]+", r"\1=***", str(msg), flags=re.I)


def _insights_fetch_finnhub(sym: str) -> dict:
    """Insider transakcie + EPS surprises z Finnhubu (free tier, funguje z Renderu)."""
    api_key = os.getenv("FINNHUB_API_KEY", "")
    if not api_key:
        raise RuntimeError("FINNHUB_API_KEY nie je nastavený")
    today = datetime.now(timezone.utc).date()
    out: dict = {}
    # Insider transakcie 90 d — SEC Form 4 kódy: P = purchase, S = sale
    r = requests.get("https://finnhub.io/api/v1/stock/insider-transactions",
                     params={"symbol": sym, "from": (today - timedelta(days=90)).isoformat(),
                             "to": today.isoformat(), "token": api_key}, timeout=15)
    r.raise_for_status()
    buys = sells = 0
    buy_val = sell_val = 0.0
    recent = []
    for t in (r.json().get("data") or []):
        code = (t.get("transactionCode") or "").upper()
        kind = "buy" if code == "P" else ("sell" if code == "S" else None)
        if not kind:
            continue   # granty, exercise, gifty atď. nezapočítavaj
        shares = abs(float(t.get("change") or 0))
        price = float(t.get("transactionPrice") or 0)
        val = shares * price
        if kind == "buy":
            buys += 1; buy_val += val
        else:
            sells += 1; sell_val += val
        if len(recent) < 6:
            recent.append({"date": t.get("transactionDate"), "name": t.get("name"),
                           "relation": None, "type": kind, "shares": shares, "value": val or None})
    out["insider"] = {"buys_90d": buys, "sells_90d": sells,
                      "net_value_90d": round(buy_val - sell_val, 0), "recent": recent}
    # EPS surprises — posledné 4 kvartály (API vracia najnovší prvý → otoč)
    r = requests.get("https://finnhub.io/api/v1/stock/earnings",
                     params={"symbol": sym, "token": api_key}, timeout=15)
    r.raise_for_status()
    eps_hist = []
    for h in (r.json() or [])[:4][::-1]:
        a, e = h.get("actual"), h.get("estimate")
        sp = h.get("surprisePercent")
        eps_hist.append({"quarter": h.get("period"), "actual": a, "estimate": e,
                         "surprise_pct": round(sp, 1) if sp is not None else None,
                         "beat": (a is not None and e is not None and a >= e)})
    out["eps_history"] = eps_hist
    try:
        r = requests.get("https://finnhub.io/api/v1/stock/recommendation",
                         params={"symbol": sym, "token": api_key}, timeout=15)
        r.raise_for_status()
        recs = r.json() or []
        if recs:
            tr = recs[0]
            out["analyst_consensus"] = {
                "period": tr.get("period"),
                "strong_buy": int(tr.get("strongBuy") or 0),
                "buy": int(tr.get("buy") or 0),
                "hold": int(tr.get("hold") or 0),
                "sell": int(tr.get("sell") or 0),
                "strong_sell": int(tr.get("strongSell") or 0),
            }
    except Exception as e:
        print(f"[insights] recommendation {sym} failed: {_scrub_token(e)}")
    try:
        r = requests.get("https://finnhub.io/api/v1/stock/price-target",
                         params={"symbol": sym, "token": api_key}, timeout=15)
        r.raise_for_status()
        pt = r.json() or {}
        if any(pt.get(k) is not None for k in ("targetMean", "targetLow", "targetHigh")):
            out["price_target"] = {
                "mean": pt.get("targetMean"),
                "low": pt.get("targetLow"),
                "high": pt.get("targetHigh"),
                "median": pt.get("targetMedian"),
                "updated_at": pt.get("lastUpdated"),
            }
    except Exception as e:
        print(f"[insights] price target {sym} failed: {_scrub_token(e)}")
    try:
        r = requests.get("https://finnhub.io/api/v1/stock/metric",
                         params={"symbol": sym, "metric": "all", "token": api_key}, timeout=15)
        r.raise_for_status()
        metric = (r.json() or {}).get("metric") or {}
        short_float = next((metric.get(k) for k in (
            "shortPercentOfFloat", "shortInterestPercent", "shortInterestPct"
        ) if metric.get(k) is not None), None)
        short_ratio = next((metric.get(k) for k in (
            "shortRatio", "shortInterestRatio"
        ) if metric.get(k) is not None), None)
        if short_float is not None or short_ratio is not None:
            out["short_interest"] = {
                "percent_float": round(float(short_float), 2) if short_float is not None else None,
                "short_ratio": short_ratio,
                "source": "finnhub",
            }
    except Exception as e:
        print(f"[insights] short interest {sym} failed: {_scrub_token(e)}")
    if not (buys or sells or eps_hist):
        raise RuntimeError("prázdna odpoveď (insider aj EPS)")
    return out


def _fmp_price_target_raw(sym: str) -> tuple[dict | list | None, str, int | None]:
    """Volá FMP price-target endpoint — najprv stable, potom v3 ako fallback.
    Vracia (raw_payload, endpoint_used, http_status). Vždy bez throw."""
    api_key = os.getenv("FMP_API_KEY", "").strip()
    if not api_key:
        return (None, "no_key", None)
    urls = [
        "https://financialmodelingprep.com/stable/price-target-consensus",
        "https://financialmodelingprep.com/api/v3/price-target-consensus",
        "https://financialmodelingprep.com/api/v4/price-target-consensus",
    ]
    last_status = None
    for url in urls:
        try:
            r = requests.get(url, params={"symbol": sym, "apikey": api_key}, timeout=10)
            last_status = r.status_code
            if r.status_code != 200:
                continue
            data = r.json()
            if isinstance(data, dict) and data.get("Error Message"):
                continue
            if isinstance(data, list) and not data:
                continue
            return (data, url, r.status_code)
        except Exception as e:
            print(f"[insights] FMP {url} {sym} failed: {e}")
    return (None, "all_failed", last_status)


def _fmp_price_target(sym: str) -> dict | None:
    """Cieľová cena analytikov z FMP — kompatibilná s out['price_target']."""
    data, _url, _status = _fmp_price_target_raw(sym)
    if data is None:
        return None
    if isinstance(data, list):
        data = data[0] if data else {}
    mean = data.get("targetConsensus") or data.get("priceTargetAverage") or data.get("targetMean")
    high = data.get("targetHigh") or data.get("priceTargetHigh")
    low  = data.get("targetLow")  or data.get("priceTargetLow")
    med  = data.get("targetMedian")
    try:
        if not mean or float(mean) <= 0:
            return None
    except (TypeError, ValueError):
        return None
    return {"mean": mean, "low": low, "high": high, "median": med, "source": "fmp"}


# ── Fair value snapshot (lazy card in Analytics; Finnhub free tier, 24h disk cache) ──
FAIR_VALUE_DIR = DATA_ROOT / "fair_value"
FAIR_VALUE_TTL_H = 24
FAIR_VALUE_SCHEMA_VERSION = 1
PEERS_DIR = DATA_ROOT / "peers"
PEERS_TTL_H = 24
PEERS_SCHEMA_VERSION = 2
MEMO_DIR = DATA_ROOT / "ticker_memos"
MEMO_TTL_H = 24
MEMO_SCHEMA_VERSION = 2


def _validate_ticker_symbol(symbol: str) -> str:
    sym = str(symbol or "").upper().strip()
    if not re.fullmatch(r"[A-Z0-9.\-]{1,20}", sym):
        raise HTTPException(status_code=400, detail="Neplatný ticker")
    return sym


def _read_ticker_json_cache(path: Path, schema_version: int, ttl_h: int) -> dict | None:
    with _cache_file_lock(path):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
            fetched_at = payload.get("fetched_at") or payload.get("generated_at")
            age_h = (datetime.now(timezone.utc) - datetime.fromisoformat(fetched_at)).total_seconds() / 3600
            if payload.get("schema_version") == schema_version and age_h < ttl_h:
                return payload
        except Exception:
            pass
    return None


def _write_ticker_json_cache(path: Path, payload: dict) -> None:
    tmp = path.with_name(
        f"{path.name}.{threading.get_ident()}.{_time_module.time_ns()}.tmp"
    )
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with _cache_file_lock(path):
            tmp.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
            os.replace(tmp, path)
    finally:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass


def _fv_num(value) -> float | None:
    try:
        value = float(value)
        return value if math.isfinite(value) else None
    except (TypeError, ValueError):
        return None


def _fv_pct(value) -> float | None:
    """Finnhub metriky sa podľa poľa vracajú ako 12.4 alebo 0.124."""
    value = _fv_num(value)
    if value is None:
        return None
    return value * 100 if abs(value) <= 1 else value


def _fv_dcf_value(fcf_per_share: float, growth_pct: float | None) -> dict | None:
    """Jednoduchý 5r DCF v troch scenároch; len pre kladný FCF na akciu."""
    if fcf_per_share is None or fcf_per_share <= 0:
        return None

    def calc(growth: float, discount: float) -> float | None:
        terminal_growth = 0.025
        if discount <= terminal_growth:
            return None
        flow = fcf_per_share
        present = 0.0
        for year in range(1, 6):
            flow *= 1 + growth
            present += flow / ((1 + discount) ** year)
        terminal = flow * (1 + terminal_growth) / (discount - terminal_growth)
        return present + terminal / ((1 + discount) ** 5)

    base_growth = max(0.02, min((growth_pct or 5.0) / 100, 0.10))
    values = {
        "low": calc(max(0.01, base_growth - 0.025), 0.11),
        "base": calc(base_growth, 0.10),
        "high": calc(min(0.12, base_growth + 0.025), 0.09),
    }
    if not all(v is not None and math.isfinite(v) and v > 0 for v in values.values()):
        return None
    return {key: round(value, 2) for key, value in values.items()}


def _build_fair_value_payload(sym: str, quote: dict, metric: dict, price_target: dict) -> dict:
    """Skladá transparentné modely bez automatického investičného odporúčania."""
    current = _fv_num(quote.get("c"))
    eps = _fv_num(metric.get("epsAnnual") or metric.get("epsNormalizedAnnual"))
    book = _fv_num(metric.get("bookValuePerShareAnnual"))
    fcf_per_share = _fv_num(metric.get("freeCashFlowPerShareAnnual"))
    growth = _fv_pct(metric.get("epsGrowth5Y") or metric.get("epsGrowthTTMYoy") or metric.get("epsGrowth3Y"))
    peg = _fv_num(metric.get("pegTTM") or metric.get("pegAnnual"))
    models: dict[str, dict] = {}
    excluded_models: list[dict] = []

    target_mean = _fv_num(price_target.get("targetMean"))
    target_low = _fv_num(price_target.get("targetLow"))
    target_high = _fv_num(price_target.get("targetHigh"))
    if target_mean and target_mean > 0:
        models["analyst_target"] = {
            "label": "Cieľ analytikov", "value": round(target_mean, 2),
            "low": round(target_low, 2) if target_low and target_low > 0 else None,
            "high": round(target_high, 2) if target_high and target_high > 0 else None,
            "note": "Priemerný 12-mesačný cieľ analytikov; nie je vlastný model dashboardu.",
        }

    price_to_book = current / book if current and book and book > 0 else None
    if eps and eps > 0 and book and book > 0 and (price_to_book is None or price_to_book <= 12):
        models["graham"] = {
            "label": "Graham number", "value": round(math.sqrt(22.5 * eps * book), 2),
            "note": "Len pre ziskové firmy s kladnou účtovnou hodnotou; pre rastové firmy môže byť príliš konzervatívne.",
        }
    elif eps and eps > 0 and book and book > 0 and price_to_book and price_to_book > 12:
        excluded_models.append({"label": "Graham number", "reason": "Nevhodný pre asset-light rastovú firmu s vysokým P/B."})

    if eps and eps > 0 and growth is not None and 5 <= growth <= 35 and peg is not None and 0 < peg <= 2.5:
        models["lynch"] = {
            "label": "Lynch / PEG", "value": round(eps * growth, 2), "peg": round(peg, 2) if peg else None,
            "growth_pct": round(growth, 1),
            "note": "Orientačný Lynchov vzťah P/E ≈ rast EPS; platí len pre stabilne rastúce ziskové firmy.",
        }
    elif eps and eps > 0 and peg is not None and peg > 2.5:
        excluded_models.append({"label": "Lynch / PEG", "reason": f"PEG {peg:.2f} je mimo použiteľného rozsahu pre tento zjednodušený model."})

    dcf = _fv_dcf_value(fcf_per_share, growth)
    if dcf:
        models["dcf"] = {
            "label": "DCF light", **dcf,
            "growth_pct": round(growth, 1) if growth is not None else 5.0,
            "note": "5-ročný model voľného cash flow na akciu. Scenáre: konzervatívny / stredný / optimistický.",
        }

    # Analyst target is external context, not a compatible intrinsic-value model.
    # Never let it turn a Graham/Lynch outlier into a misleading giant range.
    intrinsic_values = []
    intrinsic_model_count = 0
    for key in ("graham", "lynch"):
        value = _fv_num((models.get(key) or {}).get("value"))
        if value and value > 0:
            intrinsic_values.append(value)
            intrinsic_model_count += 1
    if dcf:
        intrinsic_values.extend((dcf["low"], dcf["base"], dcf["high"]))
        intrinsic_model_count += 1
    range_available = bool(dcf) or intrinsic_model_count >= 2
    fair_low = min(intrinsic_values) if range_available and intrinsic_values else None
    fair_high = max(intrinsic_values) if range_available and intrinsic_values else None
    midpoint = sorted(intrinsic_values)[len(intrinsic_values) // 2] if range_available and intrinsic_values else None
    potential = ((midpoint / current - 1) * 100) if current and midpoint else None
    status = "insufficient_models" if models else "unavailable"
    if current and fair_low and fair_high:
        status = "below_range" if current < fair_low else "above_range" if current > fair_high else "within_range"
    return {
        "ticker": sym, "schema_version": FAIR_VALUE_SCHEMA_VERSION,
        "current_price": round(current, 4) if current else None,
        "models": models,
        "summary": {
            "fair_low": round(fair_low, 2) if fair_low else None,
            "fair_high": round(fair_high, 2) if fair_high else None,
            "midpoint": round(midpoint, 2) if midpoint else None,
            "potential_pct": round(potential, 1) if potential is not None else None,
            "status": status,
            "model_count": len(models),
            "range_model_count": intrinsic_model_count,
            "note": "Pásmo modelov, nie cieľová cena ani investičné odporúčanie.",
        },
        "excluded_models": excluded_models,
        "inputs": {
            "eps_annual": round(eps, 4) if eps is not None else None,
            "book_value_per_share": round(book, 4) if book is not None else None,
            "free_cash_flow_per_share": round(fcf_per_share, 4) if fcf_per_share is not None else None,
            "eps_growth_pct": round(growth, 1) if growth is not None else None,
        },
    }


def _get_fair_value_payload(sym: str, refresh: int = 0) -> dict:
    """Shared cached fair-value loader for the route and investment memo."""
    fpath = FAIR_VALUE_DIR / f"{sym}.json"
    if not refresh:
        cached = _read_ticker_json_cache(fpath, FAIR_VALUE_SCHEMA_VERSION, FAIR_VALUE_TTL_H)
        if cached is not None:
            return cached
    api_key = os.getenv("FINNHUB_API_KEY", "").strip()
    if not api_key:
        return {"ticker": sym, "schema_version": FAIR_VALUE_SCHEMA_VERSION, "error": "FINNHUB_API_KEY nie je nastavený"}
    try:
        base = {"symbol": sym, "token": api_key}
        quote = requests.get("https://finnhub.io/api/v1/quote", params=base, timeout=12)
        metric = requests.get("https://finnhub.io/api/v1/stock/metric", params={**base, "metric": "all"}, timeout=15)
        quote.raise_for_status(); metric.raise_for_status()
        # Finnhub free plans can deny price-target while still allowing quote + metric.
        # Analyst target is an optional fourth model, never a reason to hide the rest.
        target_payload = {}
        fair_value_sources = ["finnhub"]
        try:
            target = requests.get("https://finnhub.io/api/v1/stock/price-target", params=base, timeout=12)
            if target.ok:
                target_payload = target.json() or {}
            else:
                fmp_target = _fmp_price_target(sym)
                if fmp_target:
                    target_payload = {
                        "targetMean": fmp_target.get("mean"), "targetLow": fmp_target.get("low"),
                        "targetHigh": fmp_target.get("high"), "targetMedian": fmp_target.get("median"),
                    }
                    fair_value_sources.append("fmp")
        except Exception:
            pass
        payload = _build_fair_value_payload(sym, quote.json() or {}, (metric.json() or {}).get("metric") or {}, target_payload)
        if not payload["models"]:
            payload["error"] = "Pre tento titul nie sú dostupné vhodné vstupy pre modely férovej hodnoty."
        payload["sources"] = fair_value_sources
        payload["fetched_at"] = datetime.now(timezone.utc).isoformat()
        try:
            _write_ticker_json_cache(fpath, payload)
        except Exception as e:
            payload.setdefault("warnings", []).append(
                f"Diskovú cache sa nepodarilo zapísať: {_scrub_token(e)}"
            )
        return payload
    except Exception as e:
        return {"ticker": sym, "schema_version": FAIR_VALUE_SCHEMA_VERSION,
                "error": f"Dáta férovej hodnoty nie sú dostupné: {_scrub_token(e)}"}


@app.get("/api/ticker/fair-value/{symbol}")
def get_ticker_fair_value(symbol: str, refresh: int = Query(0)):
    """Lazy free valuation snapshot. Neovplyvňuje scanner, signály ani verdikt."""
    return _get_fair_value_payload(_validate_ticker_symbol(symbol), refresh)


def _first_number(data: dict, *keys: str) -> float | None:
    for key in keys:
        value = _fv_num(data.get(key))
        if value is not None:
            return value
    return None


def _peer_row_template(sym: str) -> dict:
    return {
        "ticker": sym, "name": None, "current_price": None, "market_cap": None,
        "trailing_pe": None, "forward_pe": None, "ev_to_ebitda": None,
        "price_to_sales": None, "revenue_growth_pct": None,
        "gross_margin_pct": None, "operating_margin_pct": None,
        "debt_to_equity": None, "dividend_yield_pct": None,
        "analyst_target": None, "analyst_upside_pct": None,
        "sources": [], "insufficient_data": True,
    }


def _peer_row_finnhub(sym: str, api_key: str) -> dict:
    row = _peer_row_template(sym)
    base = {"symbol": sym, "token": api_key}
    profile = requests.get("https://finnhub.io/api/v1/stock/profile2",
                           params=base, timeout=8)
    profile.raise_for_status()
    prof = profile.json() or {}
    metric_response = requests.get("https://finnhub.io/api/v1/stock/metric",
                                   params={**base, "metric": "all"}, timeout=8)
    metric_response.raise_for_status()
    metric = (metric_response.json() or {}).get("metric") or {}
    row.update({
        "name": prof.get("name") or None,
        "market_cap": (_fv_num(prof.get("marketCapitalization")) * 1e6
                       if _fv_num(prof.get("marketCapitalization")) is not None else None),
        "trailing_pe": _first_number(metric, "peTTM", "peBasicExclExtraTTM", "peNormalizedAnnual"),
        "forward_pe": _first_number(metric, "forwardPE"),
        "ev_to_ebitda": _first_number(metric, "evEbitdaTTM", "evEbitdaAnnual"),
        "price_to_sales": _first_number(metric, "psTTM", "psAnnual"),
        "revenue_growth_pct": _fv_pct(_first_number(metric, "revenueGrowthTTMYoy", "revenueGrowth3Y")),
        "gross_margin_pct": _fv_pct(_first_number(metric, "grossMarginTTM", "grossMarginAnnual")),
        "operating_margin_pct": _fv_pct(_first_number(metric, "operatingMarginTTM", "operatingMarginAnnual")),
        "debt_to_equity": _first_number(metric, "totalDebtToEquityAnnual", "totalDebt/totalEquityAnnual"),
        # Finnhub reports this field in percentage points, not as a fraction.
        "dividend_yield_pct": _first_number(metric, "dividendYieldIndicatedAnnual", "dividendYieldTTM"),
        "analyst_target": _first_number(metric, "priceTargetMean", "targetPrice"),
        "sources": ["finnhub"],
    })
    try:
        target_response = requests.get("https://finnhub.io/api/v1/stock/price-target",
                                       params=base, timeout=8)
        if target_response.ok:
            target_payload = target_response.json() or {}
            row["analyst_target"] = _first_number(target_payload, "targetMean", "targetMedian")
    except Exception:
        pass
    target = row["analyst_target"]
    try:
        quote = requests.get("https://finnhub.io/api/v1/quote", params=base, timeout=8)
        quote.raise_for_status()
        current = _fv_num((quote.json() or {}).get("c"))
        row["current_price"] = current
        if target is not None and current and current > 0:
            row["analyst_upside_pct"] = (target / current - 1) * 100
    except Exception:
        pass
    return row


def _peer_row_yahoo(sym: str) -> dict:
    row = _peer_row_template(sym)
    qs = _yahoo_quote_summary(
        sym, "price,summaryDetail,defaultKeyStatistics,financialData,assetProfile"
    )
    if not qs:
        return row
    price = qs.get("price") or {}
    summary = qs.get("summaryDetail") or {}
    stats = qs.get("defaultKeyStatistics") or {}
    financial = qs.get("financialData") or {}
    current = _fv_num(_yraw(financial.get("currentPrice"))) or _fv_num(_yraw(price.get("regularMarketPrice")))
    target = _fv_num(_yraw(financial.get("targetMeanPrice")))
    row.update({
        "name": price.get("longName") or price.get("shortName") or None,
        "current_price": current,
        "market_cap": _fv_num(_yraw(price.get("marketCap"))) or _fv_num(_yraw(summary.get("marketCap"))),
        "trailing_pe": _fv_num(_yraw(summary.get("trailingPE"))),
        "forward_pe": _fv_num(_yraw(stats.get("forwardPE"))) or _fv_num(_yraw(summary.get("forwardPE"))),
        "ev_to_ebitda": _fv_num(_yraw(stats.get("enterpriseToEbitda"))),
        "price_to_sales": _fv_num(_yraw(summary.get("priceToSalesTrailing12Months"))),
        "revenue_growth_pct": _fv_pct(_yraw(financial.get("revenueGrowth"))),
        "gross_margin_pct": _fv_pct(_yraw(financial.get("grossMargins"))),
        "operating_margin_pct": _fv_pct(_yraw(financial.get("operatingMargins"))),
        "debt_to_equity": _fv_num(_yraw(financial.get("debtToEquity"))),
        "dividend_yield_pct": _fv_pct(_yraw(summary.get("dividendYield"))),
        "analyst_target": target,
        "analyst_upside_pct": ((target / current - 1) * 100 if target and current and current > 0 else None),
        "sources": ["yahoo"],
    })
    return row


def _peer_row_yfinance(sym: str) -> dict:
    """Last-resort yfinance snapshot; called lazily and sequentially."""
    row = _peer_row_template(sym)
    info = yf.Ticker(sym).get_info() or {}
    current = _first_number(info, "currentPrice", "regularMarketPrice", "previousClose")
    target = _first_number(info, "targetMeanPrice")
    row.update({
        "name": info.get("longName") or info.get("shortName") or None,
        "current_price": current,
        "market_cap": _first_number(info, "marketCap"),
        "trailing_pe": _first_number(info, "trailingPE"),
        "forward_pe": _first_number(info, "forwardPE"),
        "ev_to_ebitda": _first_number(info, "enterpriseToEbitda"),
        "price_to_sales": _first_number(info, "priceToSalesTrailing12Months"),
        "revenue_growth_pct": _fv_pct(info.get("revenueGrowth")),
        "gross_margin_pct": _fv_pct(info.get("grossMargins")),
        "operating_margin_pct": _fv_pct(info.get("operatingMargins")),
        "debt_to_equity": _first_number(info, "debtToEquity"),
        "dividend_yield_pct": _fv_pct(info.get("dividendYield")),
        "analyst_target": target,
        "analyst_upside_pct": ((target / current - 1) * 100 if target and current and current > 0 else None),
        "sources": ["yfinance"] if info else [],
    })
    return row


def _finalize_peer_row(row: dict) -> dict:
    numeric_fields = (
        "current_price", "market_cap", "trailing_pe", "forward_pe", "ev_to_ebitda",
        "price_to_sales", "revenue_growth_pct", "gross_margin_pct",
        "operating_margin_pct", "debt_to_equity", "dividend_yield_pct",
        "analyst_target", "analyst_upside_pct",
    )
    for key in numeric_fields:
        value = _fv_num(row.get(key))
        row[key] = round(value, 2) if value is not None else None
    available = sum(row.get(key) is not None for key in numeric_fields)
    row["insufficient_data"] = available < 4
    return row


def _yfinance_peer_symbols(sym: str, warnings: list[str]) -> list[str]:
    """Resolve sector context only; Yahoo's screener filters are not stable."""
    try:
        info = yf.Ticker(sym).get_info() or {}
        sector, industry = info.get("sector"), info.get("industry")
        if not sector and not industry:
            warnings.append("yfinance nevrátil sektor ani odvetvie pre automatický výber peers.")
            return []
        context = " / ".join(value for value in (sector, industry) if value)
        warnings.append(
            f"Sektorový kontext yfinance ({context}) je dostupný, ale spoľahlivý automatický "
            "zoznam peers vyžaduje FINNHUB_API_KEY alebo parameter peers."
        )
        return []
    except Exception as e:
        warnings.append(f"yfinance sektorový kontext nie je dostupný: {_scrub_token(e)}")
        return []


def _parse_peer_query(raw: str, sym: str, warnings: list[str]) -> list[str]:
    peers = []
    invalid = []
    for item in str(raw or "").split(","):
        ticker = item.upper().strip()
        if not ticker:
            continue
        if not re.fullmatch(r"[A-Z0-9.\-]{1,20}", ticker):
            invalid.append(ticker[:20])
            continue
        if ticker != sym and ticker not in peers:
            peers.append(ticker)
    if invalid:
        warnings.append("Neplatné tickery v parametri peers boli vynechané.")
    if len(peers) > 6:
        warnings.append("Parameter peers bol skrátený na maximálne 6 tickerov.")
    return peers[:6]


def _get_peers_payload(sym: str, peers: str = "", refresh: int = 0) -> dict:
    warnings: list[str] = []
    explicit = _parse_peer_query(peers, sym, warnings)
    identity = ",".join(explicit) if explicit else "auto"
    suffix = hashlib.blake2b(identity.encode("utf-8"), digest_size=6).hexdigest()
    fpath = PEERS_DIR / f"{sym}_{suffix}.json"
    if not refresh:
        cached = _read_ticker_json_cache(fpath, PEERS_SCHEMA_VERSION, PEERS_TTL_H)
        if cached is not None:
            if warnings:
                cached = {**cached, "warnings": list(dict.fromkeys([
                    *cached.get("warnings", []), *warnings
                ]))}
            return cached

    peer_symbols = explicit
    resolver = "explicit" if explicit else None
    api_key = os.getenv("FINNHUB_API_KEY", "").strip()
    if not peer_symbols and api_key:
        try:
            response = requests.get("https://finnhub.io/api/v1/stock/peers",
                                    params={"symbol": sym, "token": api_key}, timeout=8)
            response.raise_for_status()
            values = response.json() or []
            peer_symbols = [str(x).upper().strip() for x in values
                            if str(x).upper().strip() != sym and
                            re.fullmatch(r"[A-Z0-9.\-]{1,20}", str(x).upper().strip())][:6]
            if peer_symbols:
                resolver = "finnhub"
        except Exception as e:
            warnings.append(f"Finnhub výber peers zlyhal: {_scrub_token(e)}")
    if not peer_symbols:
        peer_symbols = _yfinance_peer_symbols(sym, warnings)
        if peer_symbols:
            resolver = "yfinance_sector_industry"
    if not peer_symbols:
        warnings.append("Automatický zoznam peers nie je dostupný; použite parameter peers.")

    rows = []
    for ticker in [sym, *peer_symbols]:
        row = None
        if api_key:
            try:
                row = _finalize_peer_row(_peer_row_finnhub(ticker, api_key))
            except Exception as e:
                warnings.append(f"Finnhub dáta pre {ticker} nie sú dostupné: {_scrub_token(e)}")
        if row is None or row.get("insufficient_data", True):
            try:
                yahoo_row = _peer_row_yahoo(ticker)
                if row is None:
                    row = yahoo_row
                else:
                    for key, value in yahoo_row.items():
                        if key == "sources":
                            row[key] = list(dict.fromkeys([*row.get(key, []), *value]))
                        elif row.get(key) is None and value is not None:
                            row[key] = value
            except Exception as e:
                warnings.append(f"Yahoo dáta pre {ticker} nie sú dostupné: {_scrub_token(e)}")
        if row is None or not row.get("sources"):
            try:
                row = _peer_row_yfinance(ticker)
            except Exception as e:
                warnings.append(f"yfinance dáta pre {ticker} nie sú dostupné: {_scrub_token(e)}")
        rows.append(_finalize_peer_row(row or _peer_row_template(ticker)))

    generated_at = datetime.now(timezone.utc).isoformat()
    sources = list(dict.fromkeys(
        [resolver] + [source for row in rows for source in row.get("sources", [])]
    )) if resolver else list(dict.fromkeys(source for row in rows for source in row.get("sources", [])))
    payload = {
        "ticker": sym, "schema_version": PEERS_SCHEMA_VERSION,
        "resolver": resolver, "peers": peer_symbols, "companies": rows,
        "sources": sources, "generated_at": generated_at, "fetched_at": generated_at,
        "insufficient_data": not peer_symbols or all(row["insufficient_data"] for row in rows),
        "warnings": list(dict.fromkeys(warnings)),
    }
    try:
        _write_ticker_json_cache(fpath, payload)
    except Exception as e:
        payload["warnings"].append(f"Diskovú cache sa nepodarilo zapísať: {_scrub_token(e)}")
    return payload


@app.get("/api/ticker/peers/{symbol}")
def get_ticker_peers(symbol: str, refresh: int = Query(0), peers: str = Query("")):
    """Lazy peer comparison. Purely orientational; never feeds trading signals."""
    return _get_peers_payload(_validate_ticker_symbol(symbol), peers, refresh)


def _memo_cache_path(sym: str, peers: str) -> Path:
    identity = ",".join(_parse_peer_query(peers, sym, [])) if peers else "auto"
    suffix = hashlib.blake2b(identity.encode("utf-8"), digest_size=6).hexdigest()
    return MEMO_DIR / f"{sym}_{suffix}.json"


@app.get("/api/ticker/memo/{symbol}")
def get_ticker_memo(symbol: str, refresh: int = Query(0), peers: str = Query("")):
    """Aggregated, cached and fail-soft investment context for AI analysis."""
    sym = _validate_ticker_symbol(symbol)
    fpath = _memo_cache_path(sym, peers)
    if not refresh:
        cached = _read_ticker_json_cache(fpath, MEMO_SCHEMA_VERSION, MEMO_TTL_H)
        if cached is not None:
            return cached

    warnings: list[str] = []

    def load_block(label: str, loader, fallback: dict) -> dict:
        try:
            value = loader()
            if not isinstance(value, dict):
                raise RuntimeError("provider vrátil neplatný formát")
            if value.get("error"):
                warnings.append(f"{label}: {value['error']}")
            return value
        except Exception as e:
            warnings.append(f"{label}: {_scrub_token(e)}")
            return fallback

    profile = load_block("Profil firmy", lambda: get_ticker_profile(sym, refresh), {"ticker": sym})
    insights = load_block("Analytický konsenzus", lambda: get_ticker_insights(sym, refresh), {"ticker": sym})
    fair_value = load_block("Férová hodnota", lambda: _get_fair_value_payload(sym, refresh), {"ticker": sym})
    peers_block = load_block("Peers", lambda: _get_peers_payload(sym, peers, refresh), {"ticker": sym, "companies": []})
    warnings.extend(peers_block.get("warnings") or [])
    target_row = next((row for row in peers_block.get("companies", [])
                       if row.get("ticker") == sym), _peer_row_template(sym))
    quote = {
        "ticker": sym, "name": target_row.get("name") or profile.get("name"),
        "current_price": target_row.get("current_price") or fair_value.get("current_price"),
        "market_cap": target_row.get("market_cap") or profile.get("market_cap"),
        "analyst_target": target_row.get("analyst_target") or (insights.get("price_target") or {}).get("mean"),
        "analyst_upside_pct": target_row.get("analyst_upside_pct"),
    }
    if quote["analyst_upside_pct"] is None:
        current, target = _fv_num(quote["current_price"]), _fv_num(quote["analyst_target"])
        if current and current > 0 and target is not None:
            quote["analyst_upside_pct"] = round((target / current - 1) * 100, 2)
    fundamentals = {key: target_row.get(key) for key in (
        "trailing_pe", "forward_pe", "ev_to_ebitda", "price_to_sales",
        "revenue_growth_pct", "gross_margin_pct", "operating_margin_pct",
        "debt_to_equity", "dividend_yield_pct",
    )}
    fundamentals["insufficient_data"] = sum(value is not None for value in fundamentals.values()) < 4
    generated_at = datetime.now(timezone.utc).isoformat()
    payload = {
        "ticker": sym, "schema_version": MEMO_SCHEMA_VERSION,
        "company": {"profile": profile, "quote": quote},
        "fundamentals": fundamentals,
        "analyst": {
            "consensus": insights.get("analyst_consensus"),
            "price_target": insights.get("price_target"),
            "source": insights.get("source"),
        },
        "fair_value": fair_value,
        "peers": peers_block,
        "field_definitions": {
            "valuation_multiples": "P/E, EV/EBITDA a P/S sú orientačné násobky ocenenia; nižšie číslo nemusí samo osebe znamenať lacnú akciu.",
            "growth_and_margins": "Rast tržieb a marže sú percentá za obdobie dostupné u poskytovateľa dát.",
            "debt_to_equity": "Pomer dlhu k vlastnému kapitálu; medzi odvetviami nie je priamo porovnateľný.",
            "analyst_upside_pct": "Rozdiel medzi priemerným cieľom analytikov a aktuálnou cenou, ak sú obe hodnoty dostupné.",
            "fair_value": "Orientačné pásmo modelov, nie cieľová cena ani investičné odporúčanie.",
        },
        "analysis_request": (
            "Údaje sú orientačné. Priprav investično-bankový memo pohľad pre horizont 1–3+ rokov: "
            "téza, porovnanie peers, ocenenie, katalyzátory a riziká. Chýbajúce údaje výslovne označ; "
            "nevymýšľaj ich a nevydávaj výstup za pokyn na obchodovanie."
        ),
        "sources": list(dict.fromkeys(
            [source for source in [profile.get("source"), insights.get("source"),
                                   *fair_value.get("sources", []),
                                   *peers_block.get("sources", [])] if source]
        )),
        "generated_at": generated_at, "fetched_at": generated_at,
        "insufficient_data": fundamentals["insufficient_data"] or peers_block.get("insufficient_data", False),
        "warnings": list(dict.fromkeys(warnings)),
    }
    try:
        _write_ticker_json_cache(fpath, payload)
    except Exception as e:
        payload["warnings"].append(f"Diskovú cache sa nepodarilo zapísať: {_scrub_token(e)}")
    return payload


@app.get("/api/diagnostics/fmp/{symbol}")
def diag_fmp(symbol: str):
    """Vráti surovú FMP odpoveď pre debugovanie price-target chyby. Kľúč nikdy v odpovedi."""
    sym = symbol.upper()
    api_key = os.getenv("FMP_API_KEY", "").strip()
    if not api_key:
        return {"ticker": sym, "error": "FMP_API_KEY nie je nastavený"}

    def _scrub(text: str) -> str:
        return str(text).replace(api_key, "***")[:500]

    attempts = []
    for url in (
        "https://financialmodelingprep.com/stable/price-target-consensus",
        "https://financialmodelingprep.com/api/v3/price-target-consensus",
        "https://financialmodelingprep.com/api/v4/price-target-consensus",
    ):
        rec = {"url": url.rsplit("/", 1)[-1] + (" (stable)" if "/stable/" in url else
               " (v3)" if "/v3/" in url else " (v4)")}
        try:
            r = requests.get(url, params={"symbol": sym, "apikey": api_key}, timeout=10)
            rec["status"] = r.status_code
            rec["body"] = _scrub(r.text)
        except Exception as e:
            rec["status"] = "exception"
            rec["body"] = _scrub(e)
        attempts.append(rec)

    data, url_used, status = _fmp_price_target_raw(sym)
    return {
        "ticker": sym,
        "endpoint": url_used,
        "http_status": status,
        "attempts": attempts,
        "raw": data,
        "parsed": _fmp_price_target(sym),
    }


# ── Ticker → SPDR sektorový ETF (Finnhub profile2) ────────────────────────────
# Finnhub vracia hrubú finnhubIndustry; mapujeme ju kľúčovými slovami na 11 SPDR
# sektorov, aby RS karta vedela porovnať ticker aj voči vlastnému sektoru, nie len
# voči QQQ/SPY. Sektor sa mení zriedka → cache 90 dní na disku.
_FINNHUB_INDUSTRY_TO_ETF = [
    ("semiconduct", "XLK"), ("software", "XLK"), ("technology", "XLK"), ("hardware", "XLK"),
    ("electronic", "XLK"), ("it services", "XLK"),
    ("media", "XLC"), ("communication", "XLC"), ("telecom", "XLC"), ("entertainment", "XLC"),
    ("retail", "XLY"), ("automobile", "XLY"), ("auto", "XLY"), ("hotel", "XLY"),
    ("restaurant", "XLY"), ("apparel", "XLY"), ("leisure", "XLY"), ("luxury", "XLY"),
    ("consumer discretionary", "XLY"),
    ("food", "XLP"), ("beverage", "XLP"), ("tobacco", "XLP"), ("household", "XLP"),
    ("consumer staple", "XLP"), ("consumer product", "XLP"),
    ("pharmaceutical", "XLV"), ("biotechnology", "XLV"), ("health", "XLV"), ("medical", "XLV"),
    ("life science", "XLV"),
    ("bank", "XLF"), ("insurance", "XLF"), ("financial", "XLF"), ("capital market", "XLF"),
    ("aerospace", "XLI"), ("machinery", "XLI"), ("transportation", "XLI"), ("airline", "XLI"),
    ("industrial", "XLI"), ("logistics", "XLI"), ("construction", "XLI"),
    ("oil", "XLE"), ("gas", "XLE"), ("energy", "XLE"), ("petroleum", "XLE"),
    ("chemical", "XLB"), ("metal", "XLB"), ("mining", "XLB"), ("material", "XLB"),
    ("paper", "XLB"), ("steel", "XLB"),
    ("utilit", "XLU"),
    ("real estate", "XLRE"), ("reit", "XLRE"),
]
SECTOR_CACHE_FILE = DATA_ROOT / "_ticker_sectors.json"
_sector_cache_lock = threading.Lock()
_SECTOR_TTL_DAYS = 90


def _industry_to_etf(industry: str) -> str | None:
    low = (industry or "").lower()
    for needle, etf in _FINNHUB_INDUSTRY_TO_ETF:
        if needle in low:
            return etf
    return None


# Finnhub profile2 'exchange' (hrubý text) → Google Finance burzový kód.
# Poradie záleží: špecifické varianty (ARCA/AMERICAN) pred generickým NYSE.
_FINNHUB_EXCH_TO_GOOGLE = [
    ("ARCA", "NYSEARCA"),
    ("NYSE AMERICAN", "NYSEAMERICAN"),
    ("NYSE MKT", "NYSEAMERICAN"),
    ("AMEX", "NYSEAMERICAN"),
    ("NASDAQ", "NASDAQ"),
    ("NEW YORK STOCK EXCHANGE", "NYSE"),
    ("NYSE", "NYSE"),
    ("TORONTO", "TSE"),
    ("XETRA", "ETR"),
    ("FRANKFURT", "FRA"),
    ("LONDON", "LON"),
    ("EURONEXT PARIS", "EPA"),
    ("EURONEXT AMSTERDAM", "AMS"),
    ("EURONEXT BRUSSELS", "EBR"),
    ("SWISS", "SWX"),
    ("TOKYO", "TYO"),
    ("HONG KONG", "HKG"),
    ("AUSTRALIAN", "ASX"),
]


def _finnhub_exch_to_google(raw: str) -> str | None:
    up = (raw or "").upper()
    for needle, code in _FINNHUB_EXCH_TO_GOOGLE:
        if needle in up:
            return code
    return None


def _ticker_sector_etf(sym: str) -> tuple[str | None, str | None]:
    """Vráti (etf, finnhub_industry) pre ticker. Cache 90d na disku, fail-soft.
    Pri fetchi uloží aj surovú burzu ('exchange') pre Google Finance odkazy."""
    etf, industry, _exch = _finnhub_profile_cached(sym)
    return etf, industry


def _finnhub_profile_cached(sym: str) -> tuple[str | None, str | None, str | None]:
    """Spoločný resolver na Finnhub profile2 — vracia (etf, industry, exchange_raw).
    Jedno volanie pokrýva sektorové mapovanie aj burzu pre Google odkaz."""
    sym = sym.upper()
    with _sector_cache_lock:
        cache = {}
        if SECTOR_CACHE_FILE.exists():
            try:
                cache = json.loads(SECTOR_CACHE_FILE.read_text(encoding="utf-8"))
            except Exception:
                cache = {}
        hit = cache.get(sym)
        if hit:
            try:
                age = (datetime.now(timezone.utc)
                       - datetime.fromisoformat(hit["fetched_at"])).days
                # exchange pribudol neskôr — keď ho starý záznam nemá, refetchni
                if age < _SECTOR_TTL_DAYS and "exchange" in hit:
                    return hit.get("etf"), hit.get("industry"), hit.get("exchange")
            except Exception:
                pass
    api_key = os.getenv("FINNHUB_API_KEY", "")
    if not api_key:
        return None, None, None
    try:
        r = requests.get("https://finnhub.io/api/v1/stock/profile2",
                         params={"symbol": sym, "token": api_key}, timeout=15)
        r.raise_for_status()
        prof = r.json() or {}
        industry = prof.get("finnhubIndustry") or ""
        exchange = prof.get("exchange") or ""
        etf = _industry_to_etf(industry)
        with _sector_cache_lock:
            cache = {}
            if SECTOR_CACHE_FILE.exists():
                try:
                    cache = json.loads(SECTOR_CACHE_FILE.read_text(encoding="utf-8"))
                except Exception:
                    cache = {}
            cache[sym] = {"etf": etf, "industry": industry, "exchange": exchange,
                          "fetched_at": datetime.now(timezone.utc).isoformat()}
            try:
                SECTOR_CACHE_FILE.write_text(json.dumps(cache), encoding="utf-8")
            except Exception:
                pass
        return etf, industry, exchange
    except Exception as e:
        print(f"[sector] profile2 {sym}: {_scrub_token(e)}")
        return None, None, None


@app.get("/api/ticker/exchanges")
def get_ticker_exchanges(tickers: str = Query(""), limit: int = Query(60)):
    """Mapuje tickery na Google Finance burzový kód (NASDAQ/NYSE/…). Z 90d cache;
    chýbajúce dorieši cez Finnhub profile2 so stropom `limit` volaní na request,
    aby veľký skener nezahltil free tier (60 req/min). Neznáma burza → None."""
    syms = [t.strip().upper() for t in tickers.split(",") if t.strip()][:300]
    out: dict[str, str | None] = {}
    budget = max(0, limit)
    # Najprv lacné cache hity bez fetchu
    with _sector_cache_lock:
        cache = {}
        if SECTOR_CACHE_FILE.exists():
            try:
                cache = json.loads(SECTOR_CACHE_FILE.read_text(encoding="utf-8"))
            except Exception:
                cache = {}
    misses = []
    for sym in syms:
        hit = cache.get(sym)
        if hit and "exchange" in hit:
            out[sym] = _finnhub_exch_to_google(hit.get("exchange"))
        else:
            misses.append(sym)
    # Zvyšok dorieš v rámci rozpočtu (každé je jeden profile2 fetch)
    for sym in misses:
        if budget <= 0:
            out[sym] = None
            continue
        _etf, _ind, exch = _finnhub_profile_cached(sym)
        out[sym] = _finnhub_exch_to_google(exch)
        budget -= 1
    return {"exchanges": out}


@app.get("/api/ticker/rs/{symbol}")
def get_ticker_rs(symbol: str):
    """Relatívna sila tickera voči QQQ, SPY a vlastnému SPDR sektoru (1M/3M).
    Interpretačná vrstva pre Predictive — NEOVPLYVŇUJE C1–C4. Benchmarky ťahajú
    zo zdieľanej _yf cache; sektor z Finnhub profile2 (cache 90d)."""
    sym = symbol.upper()

    def perf(df, days):
        c = df["Close"].astype(float)
        if len(c) <= days or not float(c.iloc[-days - 1]):
            return None
        return (float(c.iloc[-1]) - float(c.iloc[-days - 1])) / float(c.iloc[-days - 1]) * 100

    try:
        tdf = _yf_download_cached(sym, "6mo", "1d")
    except Exception as e:
        return {"ticker": sym, "error": str(e)}

    # Vlastný sektor tickera (fail-soft — RS funguje aj bez neho)
    sector_etf, sector_industry = _ticker_sector_etf(sym)

    benches = {}
    bench_list = ["QQQ", "SPY"]
    if sector_etf and sector_etf not in bench_list:
        bench_list.append(sector_etf)
    for b in bench_list:
        try:
            benches[b] = _yf_download_cached(b, "6mo", "1d")
        except Exception:
            pass
    periods = {}
    for plabel, days in (("1m", 21), ("3m", 63)):
        tp = perf(tdf, days)
        if tp is None:
            continue
        entry = {"ticker_perf": round(tp, 2)}
        for b, bdf in benches.items():
            bp = perf(bdf, days)
            if bp is not None:
                entry[f"vs_{b}"] = round(tp - bp, 2)
        periods[plabel] = entry
    if not periods:
        return {"ticker": sym, "error": "nedostatok histórie"}
    out = {"ticker": sym, "periods": periods}
    if sector_etf:
        out["sector_etf"] = sector_etf
        out["sector_name"] = _SECTOR_ETFS.get(sector_etf)
        out["sector_industry"] = sector_industry
    return out


@app.get("/api/ticker/insights/{symbol}")
def get_ticker_insights(symbol: str, refresh: int = Query(0)):
    """Insider transakcie + EPS beat/miss história + earnings dátum (Yahoo, 12h cache)."""
    sym = symbol.upper()
    fpath = YAHOO_INSIGHTS_DIR / f"{sym}.json"
    cached = None
    if fpath.exists():
        try:
            cached = json.loads(fpath.read_text(encoding="utf-8"))
            age_h = (datetime.now(timezone.utc)
                     - datetime.fromisoformat(cached["fetched_at"])).total_seconds() / 3600
            schema_ok = cached.get("schema_version") == INSIGHTS_SCHEMA_VERSION
            if schema_ok and age_h < (1 if cached.get("error") else INSIGHTS_TTL_H) and not refresh:
                return cached
        except Exception:
            cached = None
    core = None
    errs = []
    # Primárne Finnhub (z Renderu funguje), Yahoo quoteSummary ako fallback
    try:
        core = {**_insights_fetch_finnhub(sym), "source": "finnhub"}
    except Exception as e:
        errs.append(f"finnhub: {_scrub_token(e)}")
    # Finnhub can return insider/EPS data while omitting analyst targets.
    # Enrich only missing sections from Yahoo instead of discarding good data.
    core_target = (core or {}).get("price_target") or {}
    try:
        core_target_valid = float(core_target.get("mean")) > 0
    except (TypeError, ValueError):
        core_target_valid = False
    if core is not None and (
        not core.get("analyst_consensus") or not core_target_valid
    ):
        qs = _yahoo_quote_summary(sym, "recommendationTrend,financialData")
        if qs is not None:
            yahoo_extra = _insights_parse(qs)
            enriched = False
            if not core.get("analyst_consensus") and yahoo_extra.get("analyst_consensus"):
                core["analyst_consensus"] = yahoo_extra["analyst_consensus"]
                enriched = True
            yahoo_target = yahoo_extra.get("price_target") or {}
            try:
                yahoo_target_valid = float(yahoo_target.get("mean")) > 0
            except (TypeError, ValueError):
                yahoo_target_valid = False
            if not core_target_valid and yahoo_target_valid:
                core["price_target"] = yahoo_target
                enriched = True
            if enriched:
                core["source"] = "finnhub+yahoo"
        # FMP ako záloha keď Yahoo price target tiež chýba
        if not (core or {}).get("price_target"):
            fmp_pt = _fmp_price_target(sym)
            if fmp_pt:
                core["price_target"] = fmp_pt
                if "fmp" not in core.get("source", ""):
                    core["source"] = core.get("source", "finnhub") + "+fmp"
    if core is None:
        qs = _yahoo_quote_summary(
            sym,
            "insiderTransactions,earningsHistory,earningsTrend,calendarEvents,"
            "recommendationTrend,financialData,defaultKeyStatistics",
        )
        if qs is not None:
            core = {**_insights_parse(qs), "source": "yahoo"}
        else:
            errs.append("yahoo unavailable")
    if core is None:
        if cached and not cached.get("error"):
            return {**cached, "stale": True}
        payload = {"ticker": sym, "schema_version": INSIGHTS_SCHEMA_VERSION,
                   "error": " | ".join(errs),
                   "fetched_at": datetime.now(timezone.utc).isoformat()}
    else:
        payload = {"ticker": sym, "schema_version": INSIGHTS_SCHEMA_VERSION, **core,
                   "fetched_at": datetime.now(timezone.utc).isoformat()}
        # doplň earnings dátum do zdieľaného kalendára (rovnako ako per-symbol Finnhub)
        if payload.get("earnings_date"):
            try:
                ec = _earnings_cache_read()
                if ec is not None and not (ec.get("dates") or {}).get(sym):
                    ec.setdefault("dates", {})[sym] = payload["earnings_date"]
                    EARNINGS_CACHE_FILE.write_text(json.dumps(ec), encoding="utf-8")
            except Exception:
                pass
    try:
        YAHOO_INSIGHTS_DIR.mkdir(parents=True, exist_ok=True)
        fpath.write_text(json.dumps(payload), encoding="utf-8")
    except Exception:
        pass
    return payload

# ── Company profile — karta "O firme" v Analytike ────────────────────────────
COMPANY_PROFILE_DIR = DATA_ROOT / "company_profiles"
COMPANY_PROFILE_TTL_H = 24 * 30   # popis biznisu je prakticky statický
COMPANY_PROFILE_SCHEMA_VERSION = 1


def _profile_fetch_massive(sym: str) -> dict | None:
    """Massive /v3/reference/tickers — jediný náš zdroj s plným 'description'."""
    api_key = os.getenv("MASSIVE_API_KEY", "").strip()
    if not api_key:
        return None
    r = requests.get(f"https://api.massive.com/v3/reference/tickers/{sym}",
                     params={"apiKey": api_key}, timeout=20)
    r.raise_for_status()
    res = (r.json() or {}).get("results") or {}
    if not isinstance(res, dict) or not res.get("name"):
        return None
    addr = res.get("address") or {}
    return {
        "name": res.get("name"),
        "description": res.get("description"),
        "industry": res.get("sic_description"),
        "employees": res.get("total_employees"),
        "listed_since": res.get("list_date"),
        "market_cap": res.get("market_cap"),
        "website": res.get("homepage_url"),
        "exchange": res.get("primary_exchange"),
        "hq": ", ".join(x for x in (addr.get("city"), addr.get("state")) if x) or None,
    }


def _profile_fetch_yahoo(sym: str) -> dict | None:
    """Yahoo assetProfile — longBusinessSummary + sektor/odvetvie (z Render IP
    zvyčajne neprejde, ale lokálne/fallbackovo áno)."""
    qs = _yahoo_quote_summary(sym, "assetProfile")
    prof = (qs or {}).get("assetProfile") or {}
    if not prof:
        return None
    return {
        "name": None,
        "description": prof.get("longBusinessSummary"),
        "industry": ", ".join(x for x in (prof.get("sector"), prof.get("industry")) if x) or None,
        "employees": _yraw(prof.get("fullTimeEmployees")),
        "listed_since": None,
        "market_cap": None,
        "website": prof.get("website"),
        "exchange": None,
        "hq": ", ".join(x for x in (prof.get("city"), prof.get("country")) if x) or None,
    }


def _profile_fetch_finnhub(sym: str) -> dict | None:
    """Finnhub profile2 — bez popisu, ale meno/odvetvie/IPO/market cap má vždy."""
    api_key = os.getenv("FINNHUB_API_KEY", "").strip()
    if not api_key:
        return None
    r = requests.get("https://finnhub.io/api/v1/stock/profile2",
                     params={"symbol": sym, "token": api_key}, timeout=15)
    r.raise_for_status()
    prof = r.json() or {}
    if not prof.get("name"):
        return None
    mcap = prof.get("marketCapitalization")
    return {
        "name": prof.get("name"),
        "description": None,
        "industry": prof.get("finnhubIndustry"),
        "employees": None,
        "listed_since": prof.get("ipo"),
        "market_cap": float(mcap) * 1e6 if mcap else None,   # Finnhub udáva milióny USD
        "website": prof.get("weburl"),
        "exchange": prof.get("exchange"),
        "hq": prof.get("country"),
    }


@app.get("/api/ticker/profile/{symbol}")
def get_ticker_profile(symbol: str, refresh: int = Query(0)):
    """Firemný profil pre kartu "O firme": popis biznisu, odvetvie, zamestnanci,
    market cap. Massive reference → Yahoo assetProfile → Finnhub profile2;
    prvý zdroj je základ, ďalšie len dopĺňajú chýbajúce polia, stop pri popise.
    30d disk cache (1h negative), stale fallback. Čisto interpretačná vrstva."""
    sym = re.sub(r"[^A-Za-z0-9.\-]", "", symbol.upper())
    if not sym:
        raise HTTPException(status_code=400, detail="Invalid symbol")
    fpath = COMPANY_PROFILE_DIR / f"{sym}.json"
    cached = None
    if fpath.exists():
        try:
            cached = json.loads(fpath.read_text(encoding="utf-8"))
            age_h = (datetime.now(timezone.utc)
                     - datetime.fromisoformat(cached["fetched_at"])).total_seconds() / 3600
            schema_ok = cached.get("schema_version") == COMPANY_PROFILE_SCHEMA_VERSION
            if schema_ok and age_h < (1 if cached.get("error") else COMPANY_PROFILE_TTL_H) and not refresh:
                return cached
        except Exception:
            cached = None
    core = None
    errs = []
    for source, fetcher in (("massive", _profile_fetch_massive),
                            ("yahoo", _profile_fetch_yahoo),
                            ("finnhub", _profile_fetch_finnhub)):
        try:
            got = fetcher(sym)
        except Exception as e:
            errs.append(f"{source}: {_scrub_token(e)}")
            continue
        if not got:
            continue
        if core is None:
            core = {**got, "source": source}
        else:
            filled = [k for k, v in got.items() if not core.get(k) and v]
            for k in filled:
                core[k] = got[k]
            if filled:
                core["source"] += "+" + source
        if core.get("description"):
            break
    if core is None:
        if cached and not cached.get("error"):
            return {**cached, "stale": True}
        payload = {"ticker": sym, "schema_version": COMPANY_PROFILE_SCHEMA_VERSION,
                   "error": " | ".join(errs) or "no data",
                   "fetched_at": datetime.now(timezone.utc).isoformat()}
    else:
        payload = {"ticker": sym, "schema_version": COMPANY_PROFILE_SCHEMA_VERSION, **core,
                   "fetched_at": datetime.now(timezone.utc).isoformat()}
    try:
        COMPANY_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
        fpath.write_text(json.dumps(payload), encoding="utf-8")
    except Exception:
        pass
    return payload


def _earnings_next_date(symbol: str) -> str | None:
    """Najbližší earnings dátum pre symbol: bulk kalendár → Finnhub per-symbol
    (bulk free kalendár veľké tituly občas vynecháva) → yfinance fallback."""
    sym = symbol.upper()
    try:
        payload = get_earnings_calendar(refresh=0)
        d = (payload.get("dates") or {}).get(sym)
        if d:
            return d
    except Exception:
        pass
    api_key = os.getenv("FINNHUB_API_KEY", "")
    last_fail = _earnings_sym_failed_at.get(sym, 0)
    if api_key and (time.time() - last_fail) > 3600:
        try:
            today = datetime.now(timezone.utc).date()
            resp = requests.get(
                "https://finnhub.io/api/v1/calendar/earnings",
                params={"symbol": sym, "from": today.isoformat(),
                        "to": (today + timedelta(days=120)).isoformat(),
                        "token": api_key},
                timeout=15,
            )
            resp.raise_for_status()
            ds = sorted(i.get("date") for i in (resp.json().get("earningsCalendar") or []) if i.get("date"))
            if ds:
                # doplň do zdieľanej cache (fetched_at nechaj — nepredlžuj TTL bulku)
                cached = _earnings_cache_read()
                if cached is not None:
                    cached.setdefault("dates", {})[sym] = ds[0]
                    try:
                        EARNINGS_CACHE_FILE.write_text(json.dumps(cached), encoding="utf-8")
                    except Exception:
                        pass
                return ds[0]
            _earnings_sym_failed_at[sym] = time.time()
        except Exception as e:
            _earnings_sym_failed_at[sym] = time.time()
            print(f"[earnings] per-symbol fetch failed {sym}: {e}")
    # Yahoo calendarEvents raw (vlastný UA + crumb — yfinance .calendar na Renderi padá)
    qs = _yahoo_quote_summary(sym, "calendarEvents")
    if qs:
        cal_ev = ((qs.get("calendarEvents") or {}).get("earnings") or {}).get("earningsDate") or []
        ds = [d for d in (_yraw(x) for x in cal_ev) if d]
        if ds:
            return datetime.fromtimestamp(min(ds), timezone.utc).date().isoformat()
    try:
        cal = yf.Ticker(sym).calendar
        if isinstance(cal, dict):
            ed = cal.get("Earnings Date", [])
            if not isinstance(ed, list):
                ed = [ed]
            ds = sorted(str(pd.Timestamp(v).date()) for v in ed if v)
            if ds:
                return ds[0]
    except Exception:
        pass
    return None


@app.get("/api/earnings/{symbol}")
def get_earnings_for_symbol(symbol: str):
    """Per-ticker earnings dátum pre Predictive kartu (bulk → Finnhub symbol → yf)."""
    return {"ticker": symbol.upper(), "date": _earnings_next_date(symbol)}


# ── Market context (Scanner) — QQQ/SPY trend, VIX, breadth, sektory ──────────
MARKET_CTX_FILE = DATA_ROOT / "_market_context.json"
MARKET_CTX_TTL_H = 6
_market_ctx_lock = threading.Lock()
_market_breadth_running = False
MASSIVE_MARKET_DIR = DATA_ROOT / "massive_market"
MASSIVE_MARKET_DIR.mkdir(parents=True, exist_ok=True)
SP500_UNIVERSE_FILE = MASSIVE_MARKET_DIR / "_sp500_universe.json"
SP500_UNIVERSE_TTL_DAYS = 7
_massive_market_lock = threading.Lock()

_SECTOR_ETFS = {"XLK": "Tech", "XLC": "Komunikácie", "XLY": "Spotreba cykl.",
                "XLP": "Spotreba defenz.", "XLV": "Zdravotníctvo", "XLF": "Financie",
                "XLI": "Priemysel", "XLE": "Energie", "XLB": "Materiály",
                "XLU": "Utility", "XLRE": "Reality"}


def _massive_market_file(date_str: str) -> Path:
    return MASSIVE_MARKET_DIR / f"{date_str}.json"


def _sp500_universe() -> list[str]:
    cached = None
    if SP500_UNIVERSE_FILE.exists():
        try:
            cached = json.loads(SP500_UNIVERSE_FILE.read_text(encoding="utf-8"))
            fetched = datetime.fromisoformat(cached["fetched_at"])
            if (datetime.now(timezone.utc) - fetched).days < SP500_UNIVERSE_TTL_DAYS:
                return cached.get("tickers") or []
        except Exception:
            cached = None
    try:
        from bs4 import BeautifulSoup
        response = requests.get(
            "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies",
            headers={"User-Agent": "TradingDashboard/1.0 (personal market research)"},
            timeout=20,
        )
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        table = soup.select_one("table.wikitable")
        tickers = []
        for row in table.select("tbody tr")[1:] if table else []:
            cell = row.find("td")
            ticker = cell.get_text(strip=True).upper().replace(".", "-") if cell else ""
            if ticker and re.fullmatch(r"[A-Z0-9\-]+", ticker):
                tickers.append(ticker)
        tickers = sorted(set(tickers))
        if len(tickers) < 450:
            raise RuntimeError("S&P 500 universe response is incomplete")
        payload = {
            "source": "wikipedia",
            "fetched_at": datetime.now(timezone.utc).isoformat(),
            "tickers": tickers,
        }
        SP500_UNIVERSE_FILE.write_text(json.dumps(payload), encoding="utf-8")
        return tickers
    except Exception as exc:
        print(f"[massive] S&P 500 universe: {type(exc).__name__}")
        return (cached or {}).get("tickers") or []


def _massive_candidate_dates() -> list[str]:
    day = datetime.now(timezone.utc).date()
    dates = []
    while len(dates) < 5:
        day -= timedelta(days=1)
        if day.weekday() < 5:
            dates.append(day.isoformat())
    return dates


def _massive_fetch_grouped_day(date_str: str) -> dict | None:
    api_key = os.getenv("MASSIVE_API_KEY", "").strip()
    if not api_key:
        return None
    response = requests.get(
        f"https://api.massive.com/v2/aggs/grouped/locale/us/market/stocks/{date_str}",
        params={"adjusted": "true", "include_otc": "false", "apiKey": api_key},
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    rows = payload.get("results") or []
    if not rows:
        return None
    tracked_universe = set(NASDAQ100_TICKERS) | set(_sp500_universe())
    by_ticker = {}
    for row in rows:
        ticker = str(row.get("T") or "").upper()
        if ticker in tracked_universe:
            by_ticker[ticker] = {
                key: row.get(key) for key in ("o", "h", "l", "c", "v", "vw", "n", "t")
            }
    return {
        "schema_version": 2,
        "provider": "massive",
        "date": date_str,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "market_count": len(rows),
        "rows": by_ticker,
    }


def _massive_load_snapshot(refresh: bool = False) -> dict | None:
    """One grouped EOD request per market day, persisted on Render disk."""
    with _massive_market_lock:
        candidates = _massive_candidate_dates()
        if not refresh:
            for date_str in candidates:
                path = _massive_market_file(date_str)
                if path.exists():
                    try:
                        cached = json.loads(path.read_text(encoding="utf-8"))
                        if cached.get("schema_version") == 2 and cached.get("rows"):
                            return cached
                    except Exception:
                        pass
        for date_str in candidates:
            path = _massive_market_file(date_str)
            if path.exists() and not refresh:
                continue
            try:
                snapshot = _massive_fetch_grouped_day(date_str)
                if snapshot:
                    path.write_text(json.dumps(snapshot), encoding="utf-8")
                    return snapshot
            except requests.HTTPError as exc:
                status = exc.response.status_code if exc.response is not None else None
                print(f"[massive] grouped day {date_str}: HTTP {status}")
                if status == 429:
                    break
            except Exception as exc:
                print(f"[massive] grouped day {date_str}: {type(exc).__name__}")
                break
    return None


def _percentile_rank(values: list[float], value: float) -> int | None:
    clean = sorted(v for v in values if math.isfinite(v))
    if not clean or not math.isfinite(value):
        return None
    return round(sum(1 for v in clean if v <= value) / len(clean) * 100)


def _massive_universe_context(
    snapshot: dict,
    tickers: list[str],
    universe_name: str,
) -> dict | None:
    rows = snapshot.get("rows") or {}
    universe_rows = {ticker: rows[ticker] for ticker in tickers if ticker in rows}
    if not universe_rows:
        return None

    advances = declines = flat = above_vwap = 0
    up_volume = down_volume = 0.0
    transaction_values = []
    ticker_context = {}
    for ticker, row in universe_rows.items():
        open_ = float(row.get("o") or 0)
        close = float(row.get("c") or 0)
        volume = float(row.get("v") or 0)
        vwap = float(row.get("vw") or 0)
        transactions = float(row.get("n") or 0)
        change_pct = ((close / open_) - 1) * 100 if open_ > 0 else None
        vs_vwap_pct = ((close / vwap) - 1) * 100 if vwap > 0 else None
        if change_pct is not None:
            if change_pct > 0.01:
                advances += 1
                up_volume += volume
            elif change_pct < -0.01:
                declines += 1
                down_volume += volume
            else:
                flat += 1
        if vs_vwap_pct is not None and vs_vwap_pct >= 0:
            above_vwap += 1
        if transactions > 0:
            transaction_values.append(transactions)
        ticker_context[ticker] = {
            "change_pct": round(change_pct, 2) if change_pct is not None else None,
            "vs_vwap_pct": round(vs_vwap_pct, 2) if vs_vwap_pct is not None else None,
            "volume": int(volume),
            "transactions": int(transactions),
            "vwap": round(vwap, 4) if vwap else None,
        }

    for item in ticker_context.values():
        item["activity_percentile"] = _percentile_rank(
            transaction_values, float(item.get("transactions") or 0)
        )

    coverage = len(universe_rows)
    advance_pct = advances / coverage * 100
    above_vwap_pct = above_vwap / coverage * 100
    volume_ratio = up_volume / down_volume if down_volume > 0 else None
    score = advance_pct * 0.55 + above_vwap_pct * 0.45
    if volume_ratio is not None:
        score += max(-10, min(10, (volume_ratio - 1) * 12))
    score = round(max(0, min(100, score)), 1)
    if score >= 62:
        state, label = "bullish", "Bullish"
    elif score < 42:
        state, label = "defensive", "Defenzívny"
    else:
        state, label = "neutral", "Neutrálny"
    return {
        "provider": "massive",
        "universe_name": universe_name,
        "date": snapshot.get("date"),
        "market_count": snapshot.get("market_count"),
        "coverage": coverage,
        "universe": len(tickers),
        "state": state,
        "label": label,
        "score": score,
        "advancers": advances,
        "decliners": declines,
        "flat": flat,
        "advance_pct": round(advance_pct, 1),
        "above_vwap_pct": round(above_vwap_pct, 1),
        "up_down_volume_ratio": round(volume_ratio, 2) if volume_ratio is not None else None,
        "up_volume": round(up_volume),
        "down_volume": round(down_volume),
        "tickers": ticker_context,
        "interpretation_only": True,
    }


def _massive_market_context(refresh: bool = False) -> dict | None:
    if not ENABLE_MASSIVE_MARKET:
        return None
    snapshot = _massive_load_snapshot(refresh=refresh)
    if not snapshot:
        return None
    nasdaq = _massive_universe_context(snapshot, NASDAQ100_TICKERS, "Nasdaq-100")
    sp500_tickers = _sp500_universe() if ENABLE_MASSIVE_SP500 else []
    sp500 = _massive_universe_context(snapshot, sp500_tickers, "S&P 500") if sp500_tickers else None
    if not nasdaq and not sp500:
        return None
    return {
        "provider": "massive",
        "date": snapshot.get("date"),
        "market_count": snapshot.get("market_count"),
        "stored_tickers": len(snapshot.get("rows") or {}),
        "nasdaq100": nasdaq,
        "sp500": sp500,
        "interpretation_only": True,
    }


def _mc_trend(df: pd.DataFrame) -> dict:
    """Trend v štýle signal_tier: EMA10 vs EMA20 (+ close pod EMA20 pre down)."""
    close = df["Close"].astype(float)
    c = float(close.iloc[-1])
    e10 = float(close.ewm(span=10, adjust=False).mean().iloc[-1])
    e20 = float(close.ewm(span=20, adjust=False).mean().iloc[-1])
    trend = "up" if e10 > e20 else ("down" if c < e20 else "side")
    perf_1m = None
    if len(close) > 21:
        prev = float(close.iloc[-22])
        if prev:
            perf_1m = round((c - prev) / prev * 100, 2)
    return {"trend": trend, "close": round(c, 2), "perf_1m": perf_1m}


# ── FRED makro dáta (Federal Reserve) ─────────────────────────────────────────
# Vypĺňa medzeru "inflačné dáta nemáme": výnosová krivka, inflácia, fed funds,
# nezamestnanosť → reálny makro kontext k trendovo-volatilitnému kvadrantu.
# Vyžaduje FRED_API_KEY (zadarmo); bez kľúča je fail-soft (pole sa vynechá).
# Interpretačná vrstva — NEOVPLYVŇUJE C1–C4.
FRED_CACHE_FILE = DATA_ROOT / "_fred_macro.json"
_fred_lock = threading.Lock()
_FRED_TTL_H = 12


def _fred_series_latest(series_id: str, api_key: str, limit: int = 1) -> list[dict]:
    r = requests.get(
        "https://api.stlouisfed.org/fred/series/observations",
        params={"series_id": series_id, "api_key": api_key, "file_type": "json",
                "sort_order": "desc", "limit": limit},
        timeout=15,
    )
    r.raise_for_status()
    return [o for o in (r.json().get("observations") or [])
            if o.get("value") not in (".", "", None)]


def _fred_macro_label(m: dict) -> tuple[str, str]:
    infl = m.get("inflation_yoy")
    curve = m.get("yield_curve") or {}
    if curve.get("inverted"):
        return "Inverzná krivka", "10Y < 2Y — historický predstih recesie (6–18 mes.)"
    if infl is not None and infl >= 4:
        return "Vysoká inflácia", "inflácia ≥ 4 % — tlak na sadzby, opatrnosť"
    if infl is not None and 2 <= infl < 4:
        return "Goldilocks", "inflácia 2–4 % + pozitívna krivka — zdravé prostredie"
    if infl is not None and infl < 2:
        return "Dezinflácia", "inflácia < 2 % — priestor na uvoľnenie politiky"
    return "Neutrál", "bez výrazného makro extrému"


def _fred_macro(refresh: bool = False) -> dict | None:
    api_key = os.getenv("FRED_API_KEY", "").strip()
    if not api_key:
        return None
    with _fred_lock:
        if not refresh and FRED_CACHE_FILE.exists():
            try:
                cached = json.loads(FRED_CACHE_FILE.read_text(encoding="utf-8"))
                age_h = (datetime.now(timezone.utc)
                         - datetime.fromisoformat(cached["fetched_at"])).total_seconds() / 3600
                if age_h < _FRED_TTL_H:
                    return cached
            except Exception:
                pass
    out: dict = {}
    try:
        t10y2y = _fred_series_latest("T10Y2Y", api_key)
        if t10y2y:
            spread = float(t10y2y[0]["value"])
            out["yield_curve"] = {"spread_10y2y": round(spread, 2),
                                  "inverted": spread < 0, "date": t10y2y[0]["date"]}
        dgs10 = _fred_series_latest("DGS10", api_key)
        if dgs10:
            out["treasury_10y"] = round(float(dgs10[0]["value"]), 2)
        dff = _fred_series_latest("DFF", api_key)
        if dff:
            out["fed_funds"] = round(float(dff[0]["value"]), 2)
        unrate = _fred_series_latest("UNRATE", api_key)
        if unrate:
            out["unemployment"] = round(float(unrate[0]["value"]), 1)
        cpi = _fred_series_latest("CPIAUCSL", api_key, limit=13)  # YoY potrebuje 13 mes.
        if len(cpi) >= 13 and float(cpi[12]["value"]):
            out["inflation_yoy"] = round((float(cpi[0]["value"]) / float(cpi[12]["value"]) - 1) * 100, 1)
            out["inflation_date"] = cpi[0]["date"]
    except Exception as e:
        print(f"[fred] {_scrub_token(e)}")
        if not out:
            return None
    out["label"], out["note"] = _fred_macro_label(out)
    out["fetched_at"] = datetime.now(timezone.utc).isoformat()
    out["interpretation_only"] = True
    with _fred_lock:
        try:
            FRED_CACHE_FILE.write_text(json.dumps(out), encoding="utf-8")
        except Exception:
            pass
    return out


def _mc_compute_core() -> dict:
    out: dict = {}
    for key, sym in (("qqq", "QQQ"), ("spy", "SPY")):
        try:
            out[key] = _mc_trend(_yf_download_cached(sym, "6mo", "1d"))
        except Exception as e:
            out[key] = None
            print(f"[marketctx] {sym}: {e}")
    try:
        vix = _yf_download_cached("^VIX", "3mo", "1d")
        v = float(vix["Close"].astype(float).iloc[-1])
        level = "pokoj" if v < 15 else "normál" if v < 20 else "zvýšený" if v < 30 else "stres"
        out["vix"] = {"value": round(v, 1), "level": level}
    except Exception as e:
        out["vix"] = None
        print(f"[marketctx] VIX: {e}")
    sectors = []
    for etf, name in _SECTOR_ETFS.items():
        try:
            df = _yf_download_cached(etf, "3mo", "1d")
            close = df["Close"].astype(float)
            if len(close) > 21 and float(close.iloc[-22]):
                sectors.append({"etf": etf, "name": name,
                                "perf_1m": round((float(close.iloc[-1]) - float(close.iloc[-22]))
                                                 / float(close.iloc[-22]) * 100, 2)})
        except Exception:
            continue
    sectors.sort(key=lambda s: s["perf_1m"], reverse=True)
    out["sectors"] = sectors
    out["macro"] = _fred_macro()   # FRED makro vrstva (fail-soft bez FRED_API_KEY)
    return out


def _mc_regime_quadrant(data: dict) -> dict | None:
    """Trhový režim ako kvadrant trend × volatilita. Interpretačná nadstavba nad
    market-context dátami (nie inflačný Goldilocks — inflačné dáta nemáme).
    NEOVPLYVŇUJE C1–C4. Breadth zjemňuje rozhodnutie, ale nie je povinný."""
    qqq = data.get("qqq") or {}
    spy = data.get("spy") or {}
    vix = data.get("vix") or {}
    breadth = data.get("breadth") or {}
    trends = [t for t in (qqq.get("trend"), spy.get("trend")) if t]
    if not trends or vix.get("value") is None:
        return None
    up = sum(1 for t in trends if t == "up")
    down = sum(1 for t in trends if t == "down")
    b50 = breadth.get("above_ema50_pct")
    risk_on = up > down and (b50 is None or b50 >= 50)
    risk_off = down > up or (b50 is not None and b50 < 40)
    calm = vix["value"] < 20
    if risk_on and calm:
        q, label, note = "goldilocks", "Goldilocks", "rast + pokoj — ideálne prostredie pre DIP vstupy"
    elif risk_on and not calm:
        q, label, note = "overheat", "Prehriatie", "rast + nervozita — selektívne, menšie pozície"
    elif risk_off and not calm:
        q, label, note = "riskoff", "Risk-off", "pokles + stres — defenzíva, počkať na stabilizáciu"
    elif risk_off and calm:
        q, label, note = "lull", "Útlm", "pokles + pokoj — opatrné hľadanie dna"
    else:
        q, label, note = "neutral", "Neutrál", "zmiešané signály — bez jasného režimu"
    return {"quadrant": q, "label": label, "note": note}


def _mc_breadth_worker():
    """% Nasdaq-100 titulov nad EMA50/EMA200 — sekvenčne, šetrne k pamäti."""
    global _market_breadth_running
    above50 = above200 = total = total200 = 0
    for sym in NASDAQ100_TICKERS:
        try:
            df = _yf_download_cached(sym, "1y", "1d")
            close = df["Close"].astype(float)
            if len(close) < 60:
                continue
            c = float(close.iloc[-1])
            total += 1
            if c > float(close.ewm(span=50, adjust=False).mean().iloc[-1]):
                above50 += 1
            if len(close) >= 150:
                total200 += 1
                if c > float(close.ewm(span=200, adjust=False).mean().iloc[-1]):
                    above200 += 1
            del df, close
        except Exception:
            continue
    breadth = {"above_ema50_pct": round(above50 / total * 100, 1) if total else None,
               "above_ema200_pct": round(above200 / total200 * 100, 1) if total200 else None,
               "coverage": total, "universe": len(NASDAQ100_TICKERS)}
    try:
        with _market_ctx_lock:
            data = json.loads(MARKET_CTX_FILE.read_text(encoding="utf-8")) if MARKET_CTX_FILE.exists() else {}
            data["breadth"] = breadth
            data["market_regime"] = _mc_regime_quadrant(data)   # breadth dopresnil kvadrant
            MARKET_CTX_FILE.write_text(json.dumps(data), encoding="utf-8")
    except Exception as e:
        print(f"[marketctx] breadth save: {e}")
    gc.collect()
    _market_breadth_running = False


@app.get("/api/market/context")
def get_market_context(refresh: int = Query(0)):
    """Kontext trhu pre Scanner — neovplyvňuje C1–C4 scoring, len interpretácia."""
    global _market_breadth_running
    cached = None
    with _market_ctx_lock:
        if MARKET_CTX_FILE.exists():
            try:
                cached = json.loads(MARKET_CTX_FILE.read_text(encoding="utf-8"))
            except Exception:
                cached = None
    if cached and not refresh:
        try:
            age_h = (datetime.now(timezone.utc)
                     - datetime.fromisoformat(cached["fetched_at"])).total_seconds() / 3600
            if age_h < MARKET_CTX_TTL_H:
                if not cached.get("massive"):
                    cached["massive"] = _massive_market_context(refresh=False)
                return cached
        except Exception:
            pass
    data = _mc_compute_core()
    data["fetched_at"] = datetime.now(timezone.utc).isoformat()
    # starý breadth nechaj, kým worker dopočíta nový
    data["breadth"] = (cached or {}).get("breadth")
    data["massive"] = _massive_market_context(refresh=bool(refresh))
    data["market_regime"] = _mc_regime_quadrant(data)
    with _market_ctx_lock:
        try:
            MARKET_CTX_FILE.write_text(json.dumps(data), encoding="utf-8")
        except Exception:
            pass
    if ENABLE_MARKET_BREADTH and not _market_breadth_running:
        _market_breadth_running = True
        threading.Thread(target=_mc_breadth_worker, daemon=True).start()
    return data


@app.get("/api/market/massive")
def get_massive_market_context(refresh: int = Query(0)):
    data = _massive_market_context(refresh=bool(refresh))
    if not data:
        raise HTTPException(status_code=503, detail="Massive market snapshot is unavailable")
    return data


@app.post("/api/news/{ticker}/ingest")
async def ingest_ticker_news(ticker: str, request: Request):
    """Prijme surovú AV odpoveď stiahnutú prehliadačom (klientova IP obchádza
    rate-limit zdieľanej Render IP), rozparsuje ju rovnakou logikou a uloží do cache."""
    ticker = ticker.strip().upper()
    if not ticker or len(ticker) > 12:
        raise HTTPException(400, "Neplatný ticker")
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(400, "Neplatné JSON telo")
    try:
        items = _news_parse_feed(ticker, data if isinstance(data, dict) else {})
    except Exception as e:
        return {"ticker": ticker, "items": [], "stale": False,
                "fetched_at": datetime.now(timezone.utc).isoformat(),
                "error": _news_scrub_error(str(e))}
    payload = _news_save_cache(ticker, items)
    return {**payload, "stale": False}


@app.get("/api/news/{ticker}")
def get_ticker_news(ticker: str, refresh: int = Query(0)):
    ticker = ticker.strip().upper()
    if not ticker or len(ticker) > 12:
        raise HTTPException(400, "Neplatný ticker")

    cached = _news_cache_read(ticker)
    if cached and not refresh:
        age_h = None
        try:
            fetched = datetime.fromisoformat(cached["fetched_at"])
            age_h = (datetime.now(timezone.utc) - fetched).total_seconds() / 3600
        except Exception:
            pass
        # error záznam (negatívna cache, napr. rate-limit) drží len 1h,
        # plnohodnotné dáta 12h — nech reload stránky nepáli requesty do vyčerpaného limitu
        ttl = 1 if cached.get("error") else NEWS_CACHE_TTL_H
        if age_h is not None and age_h < ttl:
            return {**cached, "stale": False}

    try:
        items = _news_fetch_av(ticker)
        payload = _news_save_cache(ticker, items)
        return {**payload, "stale": False}
    except Exception as e:
        err = _news_scrub_error(str(e))
        # stale fallback — radšej staré správy než chyba; cache s dátami neprepisujeme
        if cached and cached.get("items"):
            return {**cached, "stale": True, "error": err}
        # negatívna cache (1h TTL) — nech opakované otvorenie nepáli requesty
        payload = {"ticker": ticker, "items": [],
                   "fetched_at": datetime.now(timezone.utc).isoformat(), "error": err}
        try:
            NEWS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
            _news_cache_path(ticker).write_text(
                json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass
        return {**payload, "stale": False}


@app.get("/api/checklist")
def run_checklist(tickers: str = "", days: int = 10):
    """Scan list of tickers for recent buy signals. Fetches live data for each."""
    ticker_list = [t.strip().upper() for t in tickers.split(",") if t.strip()]
    if not ticker_list:
        return {"results": []}

    results = []
    slog = load_signals_log()

    for ticker in ticker_list:
        try:
            raw_d = _yf_download_cached(ticker, "6mo", "1d")
            if len(raw_d) < 20:
                results.append({"ticker": ticker, "error": "Nedostatok dát"})
                continue

            df_d = add_indicators(raw_d)

            # Weekly data for bias (Donchian 20w + SMA50w + EMA10/20)
            raw_w = _yf_download_cached(ticker, "1y", "1wk")
            weekly_bullish = False
            weekly_trend = None
            if len(raw_w) >= 20:
                df_w = add_indicators(raw_w)
                weekly_trend = _weekly_trend(df_w)
                weekly_bullish = bool(_weekly_trend_to_bullish(weekly_trend))

            # Score closed candles
            today_date = pd.Timestamp.now("UTC").tz_localize(None).normalize().tz_localize(None)
            cutoff     = today_date - pd.Timedelta(days=days)
            ticker_slog = slog.get(ticker, {})
            recent_signal = None
            all_signals   = []
            latest_closed_date = latest_closed_daily_date(df_d)
            zscore_values = rolling_zscore(df_d["Close"]).values

            df_score = df_d.reset_index()
            for i in range(5, len(df_score)):
                row      = df_score.iloc[i]
                row_date = pd.Timestamp(row.iloc[0])
                if row_date.tzinfo: row_date = row_date.tz_localize(None)
                if row_date.date() >= today_date.date():
                    continue

                date_key = str(row_date.date())
                close = float(row["Close"])
                zscore = float(zscore_values[i]) if i < len(zscore_values) else 0.0
                sc, details = score_signal_day(row, zscore)

                if sc >= 2:
                    tier = signal_tier(sc, details["trend"])
                    if date_key not in ticker_slog:
                        entry = {
                            "score": sc,
                            "tier": tier,
                            "close": round(close, 2),
                            "details": details,
                            "rules_version": SIGNAL_RULES_VERSION,
                        }
                        if latest_closed_date is not None and row_date.normalize() == latest_closed_date:
                            entry["context"] = build_signal_context(
                                df_d, row_date, details, zscore, weekly_bullish
                            )
                        ticker_slog[date_key] = entry
                    sig = {"date": date_key, "score": sc, "tier": tier, "close": round(close, 2)}
                    all_signals.append(sig)
                    if row_date.date() >= cutoff.date():
                        if recent_signal is None or date_key > recent_signal["date"]:
                            recent_signal = sig

            if ticker_slog:
                slog[ticker] = ticker_slog

            setup = build_setup_assessment(df_d, weekly_bullish, recent_signal, len(all_signals))
            results.append({
                "ticker":         ticker,
                "weekly_bullish": weekly_bullish,
                "weekly_trend":   weekly_trend,
                "recent_signal":  recent_signal,
                "signal_count":   len(all_signals),
                "last_close":     round(float(df_d.iloc[-1]["Close"]), 2),
                **setup,
                "error":          None,
            })
        except Exception as e:
            results.append({"ticker": ticker, "error": str(e)[:60],
                            "recent_signal": None, "weekly_bullish": False,
                            "signal_count": 0, "last_close": None})

    save_signals_log(slog)
    return {"results": results}



@app.get("/api/export")
def export_snapshot(ticker: str = "AAPL", period: str = "2y"):
    """Return a self-contained HTML snapshot for archiving."""
    import json as _json
    from fastapi.responses import HTMLResponse

    try:
        raw = yf.download(ticker, period=period, interval="1wk",
                          auto_adjust=True, progress=False)
        if raw.empty:
            raise HTTPException(404, f"No data for {ticker}")
        if isinstance(raw.columns, pd.MultiIndex):
            raw.columns = raw.columns.get_level_values(0)
        raw = raw.dropna(subset=["Open", "High", "Low", "Close"])

        today = pd.Timestamp.now("UTC").tz_localize(None).normalize().tz_localize(None)
        this_monday = today - pd.Timedelta(days=today.weekday())
        last_ts = pd.Timestamp(raw.index[-1])
        if last_ts.tz is not None:
            last_ts = last_ts.tz_localize(None)
        current_week_open = last_ts >= this_monday

        df      = add_indicators(raw)
        candles = df_to_candles(df)
        pred    = predict_next_candle(df)
        backtest = run_backtest(df)

        pred_candle = {
            "time":  candles[-1]["time"] + 7 * 24 * 3600,
            "open":  pred["open"], "high": pred["high"],
            "low":   pred["low"],  "close": pred["close"],
        }
        overlay = []
        for r in backtest.get("detail", []):
            try:
                ts = int(pd.Timestamp(r["date"]).timestamp())
                overlay.append({"time": ts,
                    "pred_open": r["pred_open"], "pred_high": r["pred_high"],
                    "pred_low":  r["pred_low"],  "pred_close": r["pred_close"],
                    "actual_close": r["actual_close"], "correct": r["correct"]})
            except Exception:
                pass

        snap_date = pd.Timestamp.now("UTC").tz_localize(None).strftime("%Y-%m-%d %H:%M UTC")
        data_json = _json.dumps({
            "ticker": ticker.upper(), "snap_date": snap_date,
            "current_week_open": current_week_open,
            "candles": candles, "prediction": pred,
            "pred_candle": pred_candle,
            "backtest": {
                "total": backtest["total_predictions"],
                "direction_accuracy": backtest["direction_accuracy"],
                "avg_error_pct": backtest["avg_error_pct"],
                "indicator_hit_rate": backtest["indicator_hit_rate"],
                "overlay": overlay,
            }
        })

        html = (BASE_DIR / "static" / "index.html").read_text(encoding="utf-8")

        inject = "<script>\nconst SNAPSHOT_DATA = " + data_json + ";\nconst SNAPSHOT_MODE = true;\n</script>"
        html = html.replace("</head>", inject + "\n</head>")

        old_fetch = "    const res = await fetch(`/api/chart?ticker=${encodeURIComponent(ticker)}&period=${period}`);"
        new_fetch = (
            "    if (typeof SNAPSHOT_MODE !== 'undefined' && SNAPSHOT_MODE) {\n"
            "      lastData = SNAPSHOT_DATA;\n"
            "      document.getElementById('tickerInput').value = SNAPSHOT_DATA.ticker;\n"
            "      renderCharts(SNAPSHOT_DATA);\n"
            "      renderSidebar(SNAPSHOT_DATA);\n"
            "      document.getElementById('statusMsg').textContent ="
            " 'Snapshot: ' + SNAPSHOT_DATA.ticker + ' / ' + SNAPSHOT_DATA.snap_date;\n"
            "      document.getElementById('loadBtn').disabled = false;\n"
            "      return;\n"
            "    }\n"
            "    const res = await fetch(`/api/chart?ticker=${encodeURIComponent(ticker)}&period=${period}`);"
        )
        html = html.replace(old_fetch, new_fetch)

        filename = ticker.upper() + "_" + pd.Timestamp.now("UTC").tz_localize(None).strftime("%Y%m%d_%H%M") + ".html"
        return HTMLResponse(
            content=html,
            headers={"Content-Disposition": 'attachment; filename="' + filename + '"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

# ══ END PREDICTIVE CHART routes ══════════════════════════

# Virtual trading bot bol odstr?nen? z dashboardu (2026-06-29).
# Ak sa vr?ti, bude d?va? v???? zmysel ako samostatn? projekt/slu?ba.

@app.get("/api/health")
def health():
    return {
        "status": "ok",
        "port": 8766,
        "presets_file": PRESETS_FILE,
        "data_dir": os.getenv("DATA_DIR"),
        "render": bool(os.getenv("RENDER")),
    }


def _previous_weekday_utc() -> str:
    day = datetime.now(timezone.utc).date() - timedelta(days=1)
    while day.weekday() >= 5:
        day -= timedelta(days=1)
    return day.isoformat()


@app.get("/api/diagnostics/massive")
def massive_diagnostics(
    symbol: str = Query("AAPL", min_length=1, max_length=12),
    date: str | None = Query(None),
):
    """Probe Massive free-plan capabilities without exposing or persisting its key."""
    api_key = os.getenv("MASSIVE_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(status_code=503, detail="MASSIVE_API_KEY is not configured")

    sym = re.sub(r"[^A-Za-z0-9.\-]", "", symbol.upper())
    if not sym:
        raise HTTPException(status_code=400, detail="Invalid symbol")
    probe_date = date or _previous_weekday_utc()
    try:
        datetime.strptime(probe_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Date must use YYYY-MM-DD")

    base = "https://api.massive.com"
    probes = {
        "daily_bars": f"/v2/aggs/ticker/{sym}/range/1/day/{probe_date}/{probe_date}",
        "previous_day": f"/v2/aggs/ticker/{sym}/prev",
        "daily_market": f"/v2/aggs/grouped/locale/us/market/stocks/{probe_date}",
        "ticker_details": f"/v3/reference/tickers/{sym}",
    }

    results = {}
    for name, path in probes.items():
        try:
            response = requests.get(
                base + path,
                params={"adjusted": "true", "apiKey": api_key},
                timeout=20,
            )
            payload = response.json() if response.content else {}
            rows = payload.get("results")
            if isinstance(rows, list):
                sample = rows[0] if rows else None
                count = len(rows)
            elif isinstance(rows, dict):
                sample = rows
                count = 1
            else:
                sample = None
                count = int(payload.get("resultsCount") or 0)
            results[name] = {
                "ok": response.ok and payload.get("status") not in ("ERROR", "NOT_AUTHORIZED"),
                "http_status": response.status_code,
                "status": payload.get("status"),
                "count": count,
                "sample_fields": sorted(sample.keys()) if isinstance(sample, dict) else [],
                "message": payload.get("error") or payload.get("message"),
            }
        except Exception as exc:
            # Do not return request URLs because requests may include apiKey in them.
            results[name] = {
                "ok": False,
                "http_status": None,
                "status": "exception",
                "count": 0,
                "sample_fields": [],
                "message": type(exc).__name__,
            }

    return {
        "provider": "massive",
        "symbol": sym,
        "date": probe_date,
        "configured": True,
        "probes": results,
    }


# Cache stratégia statiky: HTML nesie ?v= verzie JS/CSS, takže MUSÍ revalidovať
# (no-cache = smie cachovať, ale vždy over so serverom). JS/CSS sú immutable —
# každá zmena obsahu bumpne ?v=, stará URL sa už nikdy nezmení, prehliadač ich
# smie držať navždy bez requestu. `private` lebo obsah je za basic auth.
_STATIC_IMMUTABLE = {"Cache-Control": "private, max-age=31536000, immutable"}
_STATIC_REVALIDATE = {"Cache-Control": "private, no-cache"}

@app.get("/")
def root():
    from fastapi.responses import FileResponse
    return FileResponse(FRONTEND_DIR / "trading_dashboard.html", headers=_STATIC_REVALIDATE)

@app.get("/help")
def help_page():
    from fastapi.responses import FileResponse
    return FileResponse(FRONTEND_DIR / "help.html", headers=_STATIC_REVALIDATE)

@app.get("/dashboard.css")
def dashboard_css():
    from fastapi.responses import FileResponse
    return FileResponse(FRONTEND_DIR / "dashboard.css", media_type="text/css", headers=_STATIC_IMMUTABLE)

@app.get("/favicon.svg")
def favicon_svg():
    from fastapi.responses import FileResponse
    return FileResponse(FRONTEND_DIR / "favicon.svg", media_type="image/svg+xml", headers=_STATIC_IMMUTABLE)

@app.get("/lizard-icon.png")
def lizard_icon_png():
    from fastapi.responses import FileResponse
    return FileResponse(FRONTEND_DIR / "lizard-icon.png", media_type="image/png", headers=_STATIC_IMMUTABLE)

@app.get("/assets/help/screenshots/{fname}")
def help_screenshot(fname: str):
    from fastapi.responses import FileResponse
    allowed = {
        "charts-overview.svg",
        "portfolio-overview.svg",
        "scanner-overview.svg",
        "predictive-overview.svg",
        "verdict-overview.svg",
    }
    if fname not in allowed:
        raise HTTPException(status_code=404, detail="Neznamy help asset")
    return FileResponse(
        FRONTEND_DIR / "assets" / "help" / "screenshots" / fname,
        media_type="image/svg+xml",
        headers=_STATIC_IMMUTABLE,
    )

# Frontend moduly — whitelist namiesto path param
# validácie, nech sa nedá vyžiadať nič mimo frontend/js/.
_JS_MODULES = {
    "core.js", "live.js", "watchlist.js", "portfolio.js", "scanner.js",
    "chart_patterns.js", "predictive.js", "verdict.js", "charts.js", "main.js",
}

@app.get("/js/{fname}")
def dashboard_js_module(fname: str):
    from fastapi.responses import FileResponse
    if fname not in _JS_MODULES:
        raise HTTPException(status_code=404, detail="Neznámy modul")
    return FileResponse(FRONTEND_DIR / "js" / fname, media_type="application/javascript", headers=_STATIC_IMMUTABLE)

if __name__ == "__main__":
    # ── eToro proxy ako background thread ─────────────────────────────────────
    try:
        import etoro_proxy as _ep
        _ep.start_proxy_thread()
    except Exception as e:
        print(f"  WARN: eToro proxy thread zlyhalo: {e}")

    if os.getenv("RENDER") and (not os.getenv("DASH_USER") or not os.getenv("DASH_PASS")):
        raise RuntimeError("RENDER mode requires DASH_USER and DASH_PASS (fail-closed).")

    _PORT = int(os.getenv("PORT", 8766))
    _HOST = "0.0.0.0" if os.getenv("RENDER") else "127.0.0.1"
    print(f"  Basic Auth: {'zapnutá' if os.getenv('DASH_USER') else 'vypnutá'}")
    print(f"Trading Dashboard — http://{_HOST}:{_PORT}")
    uvicorn.run(app, host=_HOST, port=_PORT, reload=False)
